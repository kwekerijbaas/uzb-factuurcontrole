"""Genereer tarieven_uzb.xlsx vanuit een door LO geleverde tarievenlijst.

Gebruik:
  python _genereer_tarieven_uzb.py <bron_xlsx> <uzb_code> [--vanaf YYYY-MM-DD]

Voorbeeld:
  python _genereer_tarieven_uzb.py "~/Downloads/260217 tarieven 2026 Level One.xlsx" L1 --vanaf 2026-01-01

Hiermee wordt een nieuw tabblad <uzb_code> toegevoegd aan tarieven_uzb.xlsx
(bestaand bestand wordt aangevuld; bestaande rijen voor dezelfde periode worden
afgesloten via geldig_tot = vanaf-1).

Het script kan opnieuw uitgevoerd worden bij tariefswijziging — historische
rijen blijven bestaan zodat oude facturen valideerbaar blijven.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

KOLOMMEN = [
    "loonschaal",      # bv. "GTB B2"
    "contractvorm",    # "Vast" | "Flex" | "Seizoen"
    "basisloon",       # ter referentie (informatief)
    "tarief_100",      # 100% tarief
    "tarief_135",      # 135% tarief (veelal = tarief_100)
    "tarief_150",      # 150% tarief (overwerkuren 1.5)
    "tarief_200",      # 200% tarief (overwerkuren 2.0)
    "tarief_bijz_150", # bijzondere uren 1.5
    "geldig_vanaf",    # ISO datum
    "geldig_tot",      # ISO datum, leeg = nu geldig
    "opmerking",       # vrij veld
]

CONTRACTVORMEN = ["Vast", "Flex", "Seizoen"]
TARIEVEN_KOLOM_MAP = {  # bron-Excel kolom-index naar contractvorm
    6: "Vast",
    7: "Flex",
    8: "Seizoen",
}


def lees_bron_tarieven(bron_pad: Path) -> list[dict]:
    """Lees de door LO geleverde tarievenlijst en geef genormaliseerde records terug.

    De brontabel heeft per loonschaal 5 rijen (normale uren / overwerk 1.35 / 1.5 / 2.0 /
    bijzondere 1.5), elk met 3 kolommen Vast/Flex/Seizoen. Wij flatten dat naar één
    record per (loonschaal, contractvorm).
    """
    wb = load_workbook(bron_pad, data_only=True)
    ws = wb["Tarieven"]

    records: list[dict] = []
    huidige_schaal: str = ""
    huidig_basisloon: float | None = None
    pakket: dict[str, dict] = {}     # contractvorm → tarieven dict

    def flush_pakket():
        if not huidige_schaal or not pakket:
            return
        for cv, tarieven in pakket.items():
            if not tarieven.get("tarief_100"):
                continue
            records.append({
                "loonschaal":   huidige_schaal,
                "contractvorm": cv,
                "basisloon":    huidig_basisloon,
                **tarieven,
            })

    for r in range(1, ws.max_row + 1):
        schaal_cel = ws.cell(r, 2).value
        basisloon_cel = ws.cell(r, 3).value
        component = ws.cell(r, 4).value
        percentage = ws.cell(r, 5).value

        if schaal_cel and isinstance(schaal_cel, str) and "GTB" in schaal_cel:
            # Nieuwe loonschaal-rij begint
            flush_pakket()
            huidige_schaal = schaal_cel.strip()
            huidig_basisloon = basisloon_cel if isinstance(basisloon_cel, (int, float)) else None
            pakket = {cv: {} for cv in CONTRACTVORMEN}
            continue

        if not component:
            continue
        comp = str(component).strip().lower()
        try:
            pct = float(percentage)
        except (TypeError, ValueError):
            continue

        sleutel = None
        if "normale" in comp and pct == 1.0:
            sleutel = "tarief_100"
        elif "overwerk" in comp and pct == 1.35:
            sleutel = "tarief_135"
        elif "overwerk" in comp and pct == 1.5:
            sleutel = "tarief_150"
        elif "overwerk" in comp and pct == 2.0:
            sleutel = "tarief_200"
        elif "bijzondere" in comp and pct == 1.5:
            sleutel = "tarief_bijz_150"
        if not sleutel:
            continue

        for col_idx, cv in TARIEVEN_KOLOM_MAP.items():
            v = ws.cell(r, col_idx).value
            if isinstance(v, (int, float)):
                pakket[cv][sleutel] = round(float(v), 4)

    flush_pakket()
    return records


def schrijf_tabblad(uitvoer_pad: Path, uzb_code: str, records: list[dict],
                    geldig_vanaf: str, opmerking: str = ""):
    """Schrijf records naar tabblad <uzb_code> in tarieven_uzb.xlsx.

    Als het tabblad bestaat: alle bestaande open rijen (geldig_tot leeg) worden
    afgesloten op (geldig_vanaf - 1 dag). Daarna worden de nieuwe records toegevoegd.
    """
    if uitvoer_pad.exists():
        wb = load_workbook(uitvoer_pad)
    else:
        wb = Workbook()
        # Verwijder default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if uzb_code in wb.sheetnames:
        ws = wb[uzb_code]
        # Sluit bestaande open rijen af
        kop = [c.value for c in ws[1]]
        try:
            tot_idx = kop.index("geldig_tot") + 1
            vanaf_idx = kop.index("geldig_vanaf") + 1
        except ValueError:
            sys.exit(f"Tabblad {uzb_code} bestaat maar mist verwachte kolommen.")
        nieuwe_einde = (
            date.fromisoformat(geldig_vanaf) - timedelta(days=1)
        ).isoformat()
        for r in range(2, ws.max_row + 1):
            if not ws.cell(r, tot_idx).value:
                # alleen sluiten als de oude rij vóór de nieuwe vanaf ligt
                oude_vanaf = ws.cell(r, vanaf_idx).value
                if oude_vanaf and str(oude_vanaf) < geldig_vanaf:
                    ws.cell(r, tot_idx).value = nieuwe_einde
    else:
        ws = wb.create_sheet(uzb_code)
        # Schrijf header
        for c, kol in enumerate(KOLOMMEN, 1):
            cel = ws.cell(1, c, kol)
            cel.font = Font(bold=True)
            cel.fill = PatternFill("solid", fgColor="D9EAD3")
            cel.alignment = Alignment(horizontal="center")
        # Kolombreedte
        for c, kol in enumerate(KOLOMMEN, 1):
            ws.column_dimensions[get_column_letter(c)].width = max(13, len(kol) + 2)

    # Append records onderaan
    start_row = ws.max_row + 1
    for i, rec in enumerate(records):
        rij = start_row + i
        ws.cell(rij, 1, rec.get("loonschaal", ""))
        ws.cell(rij, 2, rec.get("contractvorm", ""))
        ws.cell(rij, 3, rec.get("basisloon"))
        ws.cell(rij, 4, rec.get("tarief_100"))
        ws.cell(rij, 5, rec.get("tarief_135"))
        ws.cell(rij, 6, rec.get("tarief_150"))
        ws.cell(rij, 7, rec.get("tarief_200"))
        ws.cell(rij, 8, rec.get("tarief_bijz_150"))
        ws.cell(rij, 9, geldig_vanaf)
        ws.cell(rij, 10, "")  # geldig_tot leeg = nog actueel
        ws.cell(rij, 11, opmerking)

    wb.save(uitvoer_pad)
    print(f"✓ {len(records)} regels weggeschreven naar {uitvoer_pad} → tabblad {uzb_code}")


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("bron", help="Pad naar de tarieven-Excel van het UZB")
    p.add_argument("uzb_code", help="UZB code (bv. L1, SW, CK)")
    p.add_argument("--vanaf", default=date.today().isoformat(),
                   help="Geldig-vanaf datum (ISO, default vandaag)")
    p.add_argument("--uitvoer", default=None,
                   help="Pad naar tarieven_uzb.xlsx (default: naast dit script)")
    p.add_argument("--opmerking", default="", help="Vrije tekst per rij")
    args = p.parse_args()

    bron = Path(os.path.expanduser(args.bron))
    if not bron.exists():
        sys.exit(f"Bron niet gevonden: {bron}")

    uitvoer = Path(args.uitvoer) if args.uitvoer else Path(__file__).parent / "tarieven_uzb.xlsx"

    print(f"Lezen brontarieven uit: {bron}")
    records = lees_bron_tarieven(bron)
    print(f"  {len(records)} schaal × contractvorm-records gevonden")

    schrijf_tabblad(uitvoer, args.uzb_code, records, args.vanaf, args.opmerking)


if __name__ == "__main__":
    main()
