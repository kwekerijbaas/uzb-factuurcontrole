"""Genereer Excel-rapport: 5 tabbladen volgens vast patroon.

Tabbladen:
  1. Samenvatting          — totalen + telling per status
  2. Detail per medewerker — alle medewerkers met match-status + splitsing + tarieven
  3. Aandachtspunten       — geel: contractvorm-switch / tarief-afwijking / categorie
  4. Te-veel-gefactureerd  — rood: medewerkers/uren niet in Excel
  5. Werkwijze             — instructies + legenda
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Kleurpalet
GROEN  = "C6EFCE"; GROEN_T  = "276221"
ROOD   = "FFC7CE"; ROOD_T   = "9C0006"
GEEL   = "FFEB9C"; GEEL_T   = "9C5700"
ORANJE = "FCE4D6"; ORANJE_T = "843C0C"
BLAUW  = "BDD7EE"; BLAUW_T  = "1F3864"
GRIJS  = "F2F2F2"; GRIJS_T  = "595959"
WIT    = "FFFFFF"

STATUS_KLEUR = {
    "ok":                        (GROEN,  GROEN_T,  "✓ OK"),
    "uren_afwijking":             (GEEL,   GEEL_T,   "⚠ Uren afwijking"),
    "tarief_afwijking":           (ORANJE, ORANJE_T, "⚠ Tarief afwijking"),
    "contractvorm_switch":        (ORANJE, ORANJE_T, "⇄ Contractvorm switch"),
    "ongebruikelijke_categorie":  (GEEL,   GEEL_T,   "⚡ Ongebruikelijke cat."),
    "niet_op_factuur":            (BLAUW,  BLAUW_T,  "○ Niet op factuur"),
    "niet_in_excel":              (ROOD,   ROOD_T,   "✗ TE VEEL GEFACTUREERD"),
}

CATEGORIE_LABEL = {
    "100":      "100%",
    "135":      "135%",
    "150":      "150%",
    "200":      "200%",
    "bijz_150": "Bijz 150%",
    "bereikbaarheid": "Bereikb.",
    "reiskosten":     "Reiskost.",
    "correctie":      "Correctie",
}

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cell(ws, r, c, v="", *, bold=False, bg=WIT, fg="000000",
          align="left", sz=10, wrap=False, border=True, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def _w(ws, col_letter, breedte):
    ws.column_dimensions[col_letter].width = breedte


# ─────────────────────────────────────────────────────────────────────────────
def bouw_rapport(
    resultaten: list,           # MatchResultaat
    week_doorgegeven,           # DoorgegevenWeek
    facturen: list,             # FactuurL1
    samenvatting: dict,
    output_pad: str | Path,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _bouw_samenvatting(wb, resultaten, week_doorgegeven, facturen, samenvatting)
    _bouw_detail(wb, resultaten, week_doorgegeven)
    _bouw_aandachtspunten(wb, resultaten, week_doorgegeven)
    _bouw_te_veel_gefactureerd(wb, resultaten, week_doorgegeven)
    _bouw_werkwijze(wb, week_doorgegeven)

    pad = Path(output_pad)
    pad.parent.mkdir(parents=True, exist_ok=True)
    wb.save(pad)
    return pad


# ── tabblad 1: samenvatting ────────────────────────────────────────────────
def _bouw_samenvatting(wb, resultaten, week, facturen, sv):
    ws = wb.create_sheet(f"Samenvatting Wk{week.week_nr:02d}")
    for c, br in zip("ABCDEFGH", [4, 32, 18, 18, 16, 16, 16, 30]):
        _w(ws, c, br)

    _cell(ws, 1, 1, f"Urencontrole {week.uzb_code} — Week {week.week_nr}/{week.jaar}",
          bold=True, sz=14, border=False)
    _cell(ws, 2, 1, f"Gegenereerd: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
          sz=9, fg=GRIJS_T, border=False)
    _cell(ws, 3, 1, f"Doorgegeven: {week.bestandsnaam}", sz=9, fg=GRIJS_T, border=False)
    if facturen:
        bron = ", ".join(f.factuurnummer for f in facturen if f.factuurnummer)
        _cell(ws, 4, 1, f"Facturen: {bron}", sz=9, fg=GRIJS_T, border=False)

    # Totalen
    _cell(ws, 6, 2, "Totalen", bold=True, bg=GRIJS, sz=11)
    _cell(ws, 6, 3, "Uren", bold=True, bg=GRIJS, align="right")
    _cell(ws, 6, 4, "Bedrag (excl BTW)", bold=True, bg=GRIJS, align="right")

    tot_factuur_uren = sum(f.netto_uren for f in facturen)
    tot_factuur_bedrag = sum(f.netto_bedrag for f in facturen)
    tot_factuurbedrag_incl = sum(f.factuurbedrag for f in facturen)

    _cell(ws, 7, 2, "Doorgegeven (Excel)")
    _cell(ws, 7, 3, week.totaal, align="right", fmt="0.00")
    _cell(ws, 7, 4, "", align="right")

    _cell(ws, 8, 2, "Gefactureerd")
    _cell(ws, 8, 3, round(tot_factuur_uren, 2), align="right", fmt="0.00")
    _cell(ws, 8, 4, tot_factuur_bedrag, align="right", fmt='€#,##0.00')

    _cell(ws, 9, 2, "Verschil (factuur − doorgegeven)", bold=True)
    verschil = round(tot_factuur_uren - week.totaal, 2)
    bg_v = ROOD if verschil > 0 else (BLAUW if verschil < 0 else GROEN)
    _cell(ws, 9, 3, verschil, bold=True, align="right", fmt="+0.00;-0.00;0.00", bg=bg_v)

    _cell(ws, 10, 2, "Totaal incl BTW (referentie)", sz=9, fg=GRIJS_T)
    _cell(ws, 10, 4, tot_factuurbedrag_incl, sz=9, fg=GRIJS_T,
          align="right", fmt='€#,##0.00')

    # Tellingen per status
    _cell(ws, 12, 2, "Aantal per status", bold=True, bg=GRIJS, sz=11)
    _cell(ws, 12, 3, "Aantal", bold=True, bg=GRIJS, align="right")
    _cell(ws, 12, 4, "Toelichting", bold=True, bg=GRIJS)

    rij = 13
    volgorde = [
        "ok", "uren_afwijking", "tarief_afwijking", "contractvorm_switch",
        "ongebruikelijke_categorie", "niet_op_factuur", "niet_in_excel"
    ]
    for status in volgorde:
        n = sv["tellingen"].get(status, 0)
        bg, fg, label = STATUS_KLEUR[status]
        _cell(ws, rij, 2, label, bg=bg, fg=fg, bold=True)
        _cell(ws, rij, 3, n, align="right", bg=bg, fg=fg)
        toelichting = {
            "ok":                        "Uren én tarieven kloppen",
            "uren_afwijking":             "Totaal uren factuur ≠ Excel",
            "tarief_afwijking":           "Tarief op factuur niet in tarieventabel",
            "contractvorm_switch":        "Schaal of contractvorm wisselt binnen één week",
            "ongebruikelijke_categorie":  "200% / bijzondere uren / bereikbaarheid e.d.",
            "niet_op_factuur":            "Doorgegeven, niet gefactureerd (mogelijk Payroll)",
            "niet_in_excel":              "Op factuur, niet doorgegeven — credit aanvragen",
        }[status]
        _cell(ws, rij, 4, toelichting, sz=9)
        rij += 1

    # Financiële impact
    rij += 1
    _cell(ws, rij, 2, "Financiële impact te-veel-gefactureerd", bold=True, bg=ROOD, fg=ROOD_T)
    _cell(ws, rij, 3, sv["te_veel_gefactureerd_uren"], bold=True, align="right",
          bg=ROOD, fg=ROOD_T, fmt="0.00")
    _cell(ws, rij, 4, sv["te_veel_gefactureerd_eur"], bold=True, align="right",
          bg=ROOD, fg=ROOD_T, fmt='€#,##0.00')

    # Bekende uitzonderingen — voor F4 Vast (techniekmannen)
    rij += 2
    _cell(ws, rij, 2, "Bekende terugkerende patronen", bold=True, bg=GRIJS)
    rij += 1
    f4_vast_namen = sorted({
        r.naam_factuur for r in resultaten
        if any(t.loonschaal == "GTB F4" and t.contractvorm == "Vast"
               for t in r.tarief_regels)
    })
    if f4_vast_namen:
        _cell(ws, rij, 2, "GTB F4 Vast (specialisten/techniek)", sz=9, fg=GRIJS_T)
        _cell(ws, rij, 4, ", ".join(f4_vast_namen), sz=9, fg=GRIJS_T, wrap=True)


# ── tabblad 2: detail per medewerker ───────────────────────────────────────
def _bouw_detail(wb, resultaten, week):
    ws = wb.create_sheet(f"Detail Wk{week.week_nr:02d}")

    kolommen = [
        ("Status", 18),
        ("Naam factuur", 30),
        ("Naam Excel", 30),
        ("Nr", 6),
        ("Uren factuur", 12),
        ("Uren Excel", 12),
        ("Δ Uren", 10),
        ("100%", 8), ("135%", 8), ("150%", 8), ("200%", 8), ("Bijz", 8),
        ("Loonschaal", 14),
        ("Contractvorm", 14),
        ("Tarief 100%", 12),
        ("Tarief 150%", 12),
        ("Bedrag (excl BTW)", 14),
        ("Factuurnr", 14),
        ("Notities", 60),
    ]
    for i, (lbl, br) in enumerate(kolommen, 1):
        _cell(ws, 1, i, lbl, bold=True, bg=GRIJS, sz=10, align="center")
        _w(ws, get_column_letter(i), br)

    rij = 2
    for r in sorted(resultaten, key=lambda x: (
        ["niet_in_excel", "uren_afwijking", "contractvorm_switch",
         "tarief_afwijking", "ongebruikelijke_categorie",
         "niet_op_factuur", "ok"].index(x.hoofdstatus),
        x.naam_excel or x.naam_factuur,
    )):
        bg, fg, label = STATUS_KLEUR[r.hoofdstatus]
        _cell(ws, rij,  1, label, bold=True, bg=bg, fg=fg, sz=9, align="center")
        _cell(ws, rij,  2, r.naam_factuur)
        _cell(ws, rij,  3, r.naam_excel)
        _cell(ws, rij,  4, r.nr_excel, align="center")
        _cell(ws, rij,  5, r.uren_factuur, align="right", fmt="0.00")
        _cell(ws, rij,  6, r.uren_excel, align="right", fmt="0.00")
        delta = r.uren_verschil
        bg_d = ROOD if delta > 0 else (BLAUW if delta < 0 else WIT)
        _cell(ws, rij,  7, delta, align="right", fmt="+0.00;-0.00;0.00",
              bg=bg_d, bold=(abs(delta) > 0.05))

        # Splitsing
        _cell(ws, rij,  8, r.splits_factuur.get("100", 0) or "", align="right", fmt="0.00")
        _cell(ws, rij,  9, r.splits_factuur.get("135", 0) or "", align="right", fmt="0.00")
        _cell(ws, rij, 10, r.splits_factuur.get("150", 0) or "", align="right", fmt="0.00")
        _cell(ws, rij, 11, r.splits_factuur.get("200", 0) or "", align="right", fmt="0.00")
        _cell(ws, rij, 12, r.splits_factuur.get("bijz_150", 0) or "", align="right", fmt="0.00")

        # Loonschaal: pak de eerste niet-lege
        loonschalen = sorted({(t.loonschaal, t.contractvorm)
                              for t in r.tarief_regels if t.loonschaal})
        ls_str = " / ".join(ls for ls, _ in loonschalen) if loonschalen else ""
        cv_str = " / ".join(cv for _, cv in loonschalen) if loonschalen else ""
        _cell(ws, rij, 13, ls_str, sz=9, align="center")
        _cell(ws, rij, 14, cv_str, sz=9, align="center")

        # Tarief 100% en 150% (uit tarief_regels)
        t100 = next((t.tarief_factuur for t in r.tarief_regels if t.categorie == "100"), 0)
        t150 = next((t.tarief_factuur for t in r.tarief_regels if t.categorie == "150"), 0)
        _cell(ws, rij, 15, t100 or "", align="right", fmt="€0.000")
        _cell(ws, rij, 16, t150 or "", align="right", fmt="€0.000")

        _cell(ws, rij, 17, r.bedrag_factuur or "", align="right", fmt='€#,##0.00')
        _cell(ws, rij, 18, ", ".join(r.factuurnummers), sz=9, align="center")
        _cell(ws, rij, 19, "; ".join(r.notities), sz=9, wrap=True)

        rij += 1

    ws.freeze_panes = "B2"


# ── tabblad 3: aandachtspunten ─────────────────────────────────────────────
def _bouw_aandachtspunten(wb, resultaten, week):
    ws = wb.create_sheet(f"Aandachtspunten Wk{week.week_nr:02d}")
    ws.sheet_properties.tabColor = GEEL

    kolommen = [
        ("Type", 24),
        ("Naam", 30),
        ("Nr", 6),
        ("Detail", 80),
    ]
    for i, (lbl, br) in enumerate(kolommen, 1):
        _cell(ws, 1, i, lbl, bold=True, bg=GEEL, fg=GEEL_T, sz=10, align="center")
        _w(ws, get_column_letter(i), br)

    rij = 2
    aandachten = [
        r for r in resultaten
        if r.hoofdstatus in (
            "uren_afwijking", "tarief_afwijking", "contractvorm_switch",
            "ongebruikelijke_categorie", "niet_op_factuur"
        )
    ]
    if not aandachten:
        _cell(ws, 2, 1, "Geen aandachtspunten — alle facturen kloppen.",
              sz=11, fg=GROEN_T, bold=True)
        return

    for r in sorted(aandachten, key=lambda x: x.hoofdstatus):
        bg, fg, label = STATUS_KLEUR[r.hoofdstatus]
        _cell(ws, rij, 1, label, bold=True, bg=bg, fg=fg, sz=9, align="center")
        _cell(ws, rij, 2, r.naam_excel or r.naam_factuur)
        _cell(ws, rij, 3, r.nr_excel, align="center")
        _cell(ws, rij, 4, "; ".join(r.notities), wrap=True, sz=9)
        rij += 1


# ── tabblad 4: te veel gefactureerd ────────────────────────────────────────
def _bouw_te_veel_gefactureerd(wb, resultaten, week):
    ws = wb.create_sheet(f"Te veel gefactureerd Wk{week.week_nr:02d}")
    ws.sheet_properties.tabColor = ROOD

    kolommen = [
        ("Naam factuur", 32),
        ("Uren", 10),
        ("Bedrag (excl BTW)", 16),
        ("Factuurnr", 14),
        ("Categorieën", 22),
        ("Toelichting", 60),
    ]
    for i, (lbl, br) in enumerate(kolommen, 1):
        _cell(ws, 1, i, lbl, bold=True, bg=ROOD, fg=ROOD_T, sz=10, align="center")
        _w(ws, get_column_letter(i), br)

    teveel = [r for r in resultaten if r.hoofdstatus == "niet_in_excel"
              or (r.hoofdstatus == "uren_afwijking" and r.uren_verschil > 0)]

    if not teveel:
        _cell(ws, 2, 1,
              "Geen te-veel-gefactureerde uren in deze week — geen credit-aanvraag nodig.",
              sz=11, fg=GROEN_T, bold=True)
        return

    rij = 2
    tot_uren = 0.0
    tot_bedrag = 0.0
    for r in sorted(teveel, key=lambda x: -x.bedrag_factuur):
        _cell(ws, rij, 1, r.naam_factuur, bold=True)
        if r.hoofdstatus == "niet_in_excel":
            uren = r.uren_factuur
            bedrag = r.bedrag_factuur
        else:
            uren = r.uren_verschil
            gem_t = (r.bedrag_factuur / r.uren_factuur) if r.uren_factuur else 0
            bedrag = round(uren * gem_t, 2)
        tot_uren += uren
        tot_bedrag += bedrag
        _cell(ws, rij, 2, uren, align="right", fmt="0.00", bg=ROOD, fg=ROOD_T, bold=True)
        _cell(ws, rij, 3, bedrag, align="right", fmt='€#,##0.00', bg=ROOD, fg=ROOD_T, bold=True)
        _cell(ws, rij, 4, ", ".join(r.factuurnummers), align="center", sz=9)
        cats = ", ".join(CATEGORIE_LABEL.get(c, c) for c in sorted(r.splits_factuur))
        _cell(ws, rij, 5, cats, sz=9)
        _cell(ws, rij, 6, "; ".join(r.notities), sz=9, wrap=True)
        rij += 1

    # Totaalregel
    _cell(ws, rij, 1, "TOTAAL", bold=True, bg=ROOD, fg=ROOD_T)
    _cell(ws, rij, 2, round(tot_uren, 2), bold=True, align="right",
          fmt="0.00", bg=ROOD, fg=ROOD_T)
    _cell(ws, rij, 3, round(tot_bedrag, 2), bold=True, align="right",
          fmt='€#,##0.00', bg=ROOD, fg=ROOD_T)


# ── tabblad 5: werkwijze ───────────────────────────────────────────────────
def _bouw_werkwijze(wb, week):
    ws = wb.create_sheet("Werkwijze")
    _w(ws, "A", 4)
    _w(ws, "B", 110)

    teksten = [
        ("Werkwijze — herhaalbaar wekelijks", True, 14, BLAUW_T),
        ("", False, 10, "000000"),
        ("Bestanden:", True, 11, "000000"),
        ("• Doorgegeven uren  : ~/Downloads/WK <nr> L1.xlsx", False, 10, "000000"),
        ("• Facturen          : ~/Downloads/PP_IFAC*Level One Uitzendbureau B.V.*PurchaseInvoice.pdf", False, 10, "000000"),
        ("• Tarieventabel     : config/tarieven_uzb.xlsx (tabblad L1)", False, 10, "000000"),
        ("", False, 10, "000000"),
        ("Stappen:", True, 11, "000000"),
        ("1. Plaats de doorgegeven-uren-Excel en alle bijbehorende factuur-PDF's in ~/Downloads/.", False, 10, "000000"),
        ("2. Open een terminal in de uren_controle/ map.", False, 10, "000000"),
        ("3. Run: python main.py --week <nr> --jaar <jaar>", False, 10, "000000"),
        ("   Of zonder argumenten — dan worden week en jaar gedetecteerd uit de Excel.", False, 10, "000000"),
        ("4. Resultaat: dit Excel-rapport + een mail-concept (.txt) naast in ~/Downloads/.", False, 10, "000000"),
        ("", False, 10, "000000"),
        ("Statussen + kleuren:", True, 11, "000000"),
    ]
    rij = 1
    for t, b, sz, fg in teksten:
        _cell(ws, rij, 2, t, bold=b, sz=sz, fg=fg, border=False)
        rij += 1

    rij += 1
    for status, (bg, fg, label) in STATUS_KLEUR.items():
        _cell(ws, rij, 2, f"  {label}", bold=True, sz=10, bg=bg, fg=fg)
        rij += 1

    rij += 1
    _cell(ws, rij, 2, "Bij tariefswijziging:", bold=True, sz=11, border=False)
    rij += 1
    _cell(ws, rij, 2, "  Open config/tarieven_uzb.xlsx → tabblad L1.", sz=10, border=False)
    rij += 1
    _cell(ws, rij, 2, "  Sluit oude regels af door 'geldig_tot' in te vullen.", sz=10, border=False)
    rij += 1
    _cell(ws, rij, 2, "  Voeg nieuwe regels toe met de nieuwe 'geldig_vanaf' datum.", sz=10, border=False)
    rij += 1
    _cell(ws, rij, 2,
          "  Of run config/_genereer_tarieven_uzb.py met de nieuwe LO-tarievenlijst.",
          sz=10, border=False)
