"""Lees tarieven_uzb.xlsx en bouw lookup-functies om factuur-tarieven te valideren.

Gebruik:
    db = TarievenDatabase("config/tarieven_uzb.xlsx")
    match = db.zoek("L1", tarief=28.942, percentage=100, datum="2026-04-06")
    # match → TarievenMatch(loonschaal="GTB B2", contractvorm="Flex", ...) of None
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


# Mapping factuur-percentage naar tarieventabel-kolomnaam
PCT_NAAR_KOLOM = {
    100: "tarief_100",
    135: "tarief_135",
    150: "tarief_150",     # standaard 150% = overwerk 1.5
    200: "tarief_200",
    # bijzondere uren 1.5 — apart vanwege ander tarief
    "bijz_150": "tarief_bijz_150",
}


@dataclass
class TarievenRegel:
    uzb_code:     str
    loonschaal:   str          # "GTB B2"
    contractvorm: str          # "Vast" | "Flex" | "Seizoen"
    basisloon:    float | None
    tarieven:     dict         # {"tarief_100": 28.94, ...}
    geldig_vanaf: str          # ISO datum
    geldig_tot:   str          # ISO datum, "" = nog actueel
    opmerking:    str = ""


@dataclass
class TarievenMatch:
    """Resultaat van een lookup."""
    loonschaal:   str
    contractvorm: str
    tarief_in_tabel: float
    tarief_op_factuur: float
    delta:        float        # factuur − tabel
    binnen_tolerantie: bool
    bron:         TarievenRegel


class TarievenDatabase:
    def __init__(self, pad: str | Path):
        self.pad = Path(pad)
        self.regels: dict[str, list[TarievenRegel]] = {}   # uzb → regels
        self._laad()

    def _laad(self):
        if not self.pad.exists():
            raise FileNotFoundError(f"Tarieven-bestand niet gevonden: {self.pad}")
        wb = load_workbook(self.pad, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.max_row < 2:
                continue
            kop = [str(c.value or "").strip() for c in ws[1]]
            try:
                idx = {k: kop.index(k) for k in (
                    "loonschaal", "contractvorm", "basisloon",
                    "tarief_100", "tarief_135", "tarief_150",
                    "tarief_200", "tarief_bijz_150",
                    "geldig_vanaf", "geldig_tot", "opmerking",
                )}
            except ValueError as e:
                raise ValueError(f"Tabblad {sheet} heeft niet de verwachte kolommen: {e}")

            regels: list[TarievenRegel] = []
            for r in range(2, ws.max_row + 1):
                row = [ws.cell(r, c + 1).value for c in range(len(kop))]
                if not row[idx["loonschaal"]]:
                    continue
                tarieven = {
                    "tarief_100":     row[idx["tarief_100"]],
                    "tarief_135":     row[idx["tarief_135"]],
                    "tarief_150":     row[idx["tarief_150"]],
                    "tarief_200":     row[idx["tarief_200"]],
                    "tarief_bijz_150":row[idx["tarief_bijz_150"]],
                }
                regels.append(TarievenRegel(
                    uzb_code=     sheet,
                    loonschaal=   str(row[idx["loonschaal"]]).strip(),
                    contractvorm= str(row[idx["contractvorm"]] or "").strip(),
                    basisloon=    row[idx["basisloon"]],
                    tarieven=     tarieven,
                    geldig_vanaf= str(row[idx["geldig_vanaf"]] or "")[:10],
                    geldig_tot=   str(row[idx["geldig_tot"]] or "")[:10],
                    opmerking=    str(row[idx["opmerking"]] or ""),
                ))
            self.regels[sheet] = regels

    def actief_op(self, uzb_code: str, datum_iso: str) -> list[TarievenRegel]:
        """Geef alle tarievenregels die op de gegeven datum geldig zijn."""
        if uzb_code not in self.regels:
            return []
        out = []
        for r in self.regels[uzb_code]:
            if r.geldig_vanaf and datum_iso < r.geldig_vanaf:
                continue
            if r.geldig_tot and datum_iso > r.geldig_tot:
                continue
            out.append(r)
        return out

    def zoek(
        self,
        uzb_code: str,
        tarief: float,
        percentage: int | str,
        datum_iso: str,
        tolerantie: float = 0.02,
    ) -> Optional[TarievenMatch]:
        """Zoek de loonschaal/contractvorm waarvoor het tarief op deze datum past.

        Retourneert TarievenMatch met de beste fit; None als geen enkel tarief
        binnen de tolerantie ligt.
        """
        kolom = PCT_NAAR_KOLOM.get(percentage)
        if not kolom:
            return None

        kandidaten = self.actief_op(uzb_code, datum_iso)
        beste: Optional[TarievenMatch] = None
        for r in kandidaten:
            tabel_t = r.tarieven.get(kolom)
            if tabel_t is None:
                continue
            try:
                delta = round(float(tarief) - float(tabel_t), 4)
            except (TypeError, ValueError):
                continue
            binnen = abs(delta) <= tolerantie
            if beste is None or abs(delta) < abs(beste.delta):
                beste = TarievenMatch(
                    loonschaal=        r.loonschaal,
                    contractvorm=      r.contractvorm,
                    tarief_in_tabel=   float(tabel_t),
                    tarief_op_factuur= float(tarief),
                    delta=             delta,
                    binnen_tolerantie= binnen,
                    bron=              r,
                )
        return beste

    def alle_uzbs(self) -> list[str]:
        return list(self.regels.keys())


if __name__ == "__main__":
    # Sanity-check: laad het bestand naast deze module
    pad = Path(__file__).resolve().parent.parent / "config" / "tarieven_uzb.xlsx"
    db = TarievenDatabase(pad)
    print(f"UZBs in tarievenbestand: {db.alle_uzbs()}")
    actief = db.actief_op("L1", "2026-04-07")
    print(f"L1 actief op 2026-04-07: {len(actief)} regels")

    # Test enkele bekende tarieven uit week 15
    voor_test = [
        (28.942, 100, "Karyna Adamovych B2 Flex"),
        (33.922, 150, "Karyna Adamovych B2 Flex"),
        (33.372, 100, "Patryk Kolodziej F4 Vast"),
        (38.752, 150, "Patryk Kolodziej F4 Vast"),
        (29.612, 100, "Lukasz Baldowski C2 Flex"),
        (32.742, 100, "Pawel Lisowski (verdacht — onbekende schaal)"),
    ]
    for t, pct, label in voor_test:
        m = db.zoek("L1", tarief=t, percentage=pct, datum_iso="2026-04-07")
        if m:
            status = "✓" if m.binnen_tolerantie else "✗"
            print(f"  {status} {t:>7.3f} @ {pct}% : {m.loonschaal} {m.contractvorm} "
                  f"(delta {m.delta:+.4f}) — {label}")
        else:
            print(f"  ✗ {t:>7.3f} @ {pct}% : geen match — {label}")
