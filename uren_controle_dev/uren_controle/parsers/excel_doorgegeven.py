"""Lees `WK <nr> <UZB>.xlsx` met de doorgegeven uren per medewerker per dag.

Verwachte structuur (zie WK 15 L1.xlsx):
  R1: Kwekerij Baas
  R2: Week <YYYYWW>          (← week+jaar uit tweede deel; ook in sheetnaam "Week YYYYWW")
  R3: kolom-headers ('Medewerkers', Ma..Zo, Totaal)
  R4: data-headers (06-04, 07-04, ...)
  R5: leeg
  R6: 'Kolom1'..'Kolom9' (auto-generated header)
  R7..N: data rijen — kolom 3 = ' XXX Voornaam Achternaam', kolom 4..10 = Ma..Zo, kolom 11 = Totaal
  Laatste rijen: 'Totaal' regel + lege rij
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class DoorgegevenRegel:
    nr:           str                # personeelsnummer
    naam:         str                # zoals in Excel ("Marius Mic")
    voornaam:     str
    achternaam:   str
    week_nr:      int
    jaar:         int
    dag_uren:     dict = field(default_factory=dict)   # {date_iso: float_uren}
    totaal_uren:  float = 0.0
    notitie:      str = ""

    @property
    def uren_per_dag_lijst(self) -> list[float]:
        """[ma, di, wo, do, vr, za, zo] in uren — 0 als geen uren."""
        return [self.dag_uren.get(d, 0.0) for d in sorted(self.dag_uren)][:7]


@dataclass
class DoorgegevenWeek:
    week_nr:    int
    jaar:       int
    uzb_code:   str
    bestandsnaam: str
    regels:     list[DoorgegevenRegel] = field(default_factory=list)
    totaal:     float = 0.0
    weekdatums: list[str] = field(default_factory=list)  # ISO datums Ma..Zo


# Detectie van naamregel in kolom 3: " 87 Marius Mic" of "  87 Marius Mic"
NAAM_RE = re.compile(r"^\s*(\d+)\s+(.+?)\s*$")


def _split_naam(volledig: str) -> tuple[str, str]:
    """Voornaam = alle woorden behalve laatste; achternaam = laatste woord."""
    parts = volledig.strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[:-1]), parts[-1]


def _detecteer_week_jaar(ws) -> tuple[int, int]:
    """Lees week en jaar uit sheetnaam ('Week 202615') of cel R2 ('Week 202615')."""
    bronnen = [ws.title]
    if ws.cell(2, 1).value:
        bronnen.append(str(ws.cell(2, 1).value))
    for s in bronnen:
        m = re.search(r"(\d{4})(\d{2})", s)
        if m:
            return int(m.group(2)), int(m.group(1))
    return 0, 0


def _detecteer_datums(ws, max_kolom: int = 11) -> list[str]:
    """Lees rij 4 voor de week-datums (formaat 06-04 of als datetime)."""
    _, jaar = _detecteer_week_jaar(ws)
    if not jaar:
        jaar = date.today().year
    datums: list[str] = []
    for c in range(4, max_kolom):
        v = ws.cell(4, c).value
        if v is None:
            continue
        if isinstance(v, datetime):
            datums.append(v.date().isoformat())
            continue
        s = str(v).strip()
        m = re.match(r"(\d{1,2})-(\d{1,2})", s)
        if m:
            d = int(m.group(1))
            mnd = int(m.group(2))
            datums.append(f"{jaar}-{mnd:02d}-{d:02d}")
    return datums


def detecteer_week_jaar(xlsx_pad: str | Path) -> tuple[int, int]:
    """Snelle helper: lees alleen sheetnaam/cel R2 voor (week_nr, jaar).

    Veel sneller dan `lees_doorgegeven` omdat alleen 2 cellen worden gelezen.
    Retourneert (0, 0) als niet detecteerbaar.
    """
    pad = Path(xlsx_pad)
    wb = load_workbook(pad, data_only=True, read_only=True)
    try:
        ws = wb.active
        return _detecteer_week_jaar(ws)
    finally:
        wb.close()


def lees_doorgegeven(xlsx_pad: str | Path, uzb_code: str) -> DoorgegevenWeek:
    """Parseer een doorgegeven-uren Excel naar een DoorgegevenWeek."""
    pad = Path(xlsx_pad)
    wb = load_workbook(pad, data_only=True)
    ws = wb.active

    week_nr, jaar = _detecteer_week_jaar(ws)
    datums = _detecteer_datums(ws)

    regels: list[DoorgegevenRegel] = []
    for r in range(7, ws.max_row + 1):
        cel = ws.cell(r, 3).value
        if cel is None:
            continue
        m = NAAM_RE.match(str(cel))
        if not m:
            continue
        nr = m.group(1).strip()
        volledige_naam = m.group(2).strip()
        if volledige_naam.lower().startswith("totaal"):
            continue

        voornaam, achternaam = _split_naam(volledige_naam)
        dag_uren: dict[str, float] = {}
        notitie = ""

        # Kolommen 4..10 zijn Ma..Zo; daarna kolom 11 = totaal
        for ci, dt in enumerate(datums):
            v = ws.cell(r, 4 + ci).value
            if v is None or v == "":
                continue
            try:
                dag_uren[dt] = float(v)
            except (TypeError, ValueError):
                # Geen getal — beschouw als notitie als het een string is
                if isinstance(v, str) and v.strip():
                    notitie = v.strip()

        # Totaal-kolom (11)
        tot_cel = ws.cell(r, 11).value
        try:
            totaal = float(tot_cel)
        except (TypeError, ValueError):
            totaal = sum(dag_uren.values())

        # Optioneel notitie-kolom verderop
        if not notitie:
            for c2 in range(12, ws.max_column + 1):
                v = ws.cell(r, c2).value
                if isinstance(v, str) and v.strip():
                    notitie = v.strip()
                    break

        if totaal <= 0 and not dag_uren:
            continue

        regels.append(DoorgegevenRegel(
            nr=nr,
            naam=volledige_naam,
            voornaam=voornaam,
            achternaam=achternaam,
            week_nr=week_nr,
            jaar=jaar,
            dag_uren=dag_uren,
            totaal_uren=totaal,
            notitie=notitie,
        ))

    weektotaal = sum(r.totaal_uren for r in regels)

    return DoorgegevenWeek(
        week_nr=week_nr,
        jaar=jaar,
        uzb_code=uzb_code,
        bestandsnaam=pad.name,
        regels=regels,
        totaal=round(weektotaal, 2),
        weekdatums=datums,
    )


if __name__ == "__main__":
    # Sanity-check op week 15 L1
    pad = Path("C:/Users/dieter.KWEKERIJBAAS/Downloads/WK 15 L1.xlsx")
    week = lees_doorgegeven(pad, "L1")
    print(f"Bestand: {week.bestandsnaam}")
    print(f"Week {week.week_nr}/{week.jaar} — {week.uzb_code}")
    print(f"Datums: {week.weekdatums}")
    print(f"Aantal medewerkers: {len(week.regels)}")
    print(f"Totaal uren: {week.totaal}")
    print()
    print("Eerste 5 medewerkers:")
    for r in week.regels[:5]:
        print(f"  nr={r.nr:>4s}  {r.naam:30s}  totaal={r.totaal_uren:5.2f}  "
              f"{[f'{r.dag_uren.get(d, 0):.2f}' for d in week.weekdatums]}")
