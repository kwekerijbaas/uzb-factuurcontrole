"""Vergelijk per medewerker per dag: Nitea-werktijd vs Excel-doorgegeven uren.

Doel: in kaart brengen hoe vaak en hoeveel de arbeidsplanner handmatig corrigeert,
en in welke richting (alleen naar boven? ook naar beneden?).
"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

NITEA_PDF = Path("C:/Users/dieter.KWEKERIJBAAS/Downloads/WK 15 L1 Nitea overzicht.pdf")
EXCEL = Path("C:/Users/dieter.KWEKERIJBAAS/Downloads/WK 15 L1.xlsx")

DATA_RE = re.compile(
    r"^\s*(\d+)\s+(\d+)\s*-\s*(.+?)\s+"
    r"(\d{2}-\d{2}-\d{4})\s+"
    r"(\d{1,2}:\d{2})\s+"
    r"(\d{1,2}:\d{2})\s+"
    r"(\d{1,2}:\d{2})"
    r"(?:\s+(\d{1,2}:\d{2}))?\s*$"
)

DAGEN = ["Ma", "Di", "Wo", "Do", "Vr", "Za", "Zo"]


def hhmm_naar_uren(s: str) -> float:
    h, m = s.split(":")
    return int(h) + int(m) / 60


def parse_nitea(pdf_pad: Path) -> dict:
    """{(nr, datum_iso): werktijd_uren}"""
    resultaat: dict = {}
    with pdfplumber.open(pdf_pad) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                m = DATA_RE.match(line)
                if not m:
                    continue
                _, nr, naam, datum, _, _, werk, _ = m.groups()
                d = datetime.strptime(datum, "%d-%m-%Y").date().isoformat()
                # Som als iemand 2x per dag inklokt (bv. lunchpauze)
                resultaat[(nr, d)] = resultaat.get((nr, d), 0.0) + hhmm_naar_uren(werk)
    return resultaat


def parse_excel(xlsx_pad: Path) -> dict:
    """{(nr, datum_iso): doorgegeven_uren}.

    Gebruikt rij 4 voor data-headers (06-04, 07-04, ...) en parsed maandag t/m zondag.
    """
    wb = load_workbook(xlsx_pad, data_only=True)
    ws = wb.active

    # Datum-headers uit rij 4, kolom 4-10 (= D-J)
    datums: list[str] = []
    for c in range(4, 11):
        v = ws.cell(4, c).value
        if v is None:
            datums.append("")
            continue
        # Format "06-04" → met jaar uit sheetnaam (Week 202615)
        sheetnaam = ws.title
        m = re.search(r"(\d{4})(\d{2})", sheetnaam)
        jaar = int(m.group(1)) if m else 2026
        if isinstance(v, datetime):
            datums.append(v.date().isoformat())
        else:
            d_str = str(v).strip()
            mm = re.match(r"(\d{1,2})-(\d{1,2})", d_str)
            if mm:
                datums.append(f"{jaar}-{int(mm.group(2)):02d}-{int(mm.group(1)):02d}")
            else:
                datums.append("")

    resultaat: dict = {}
    for r in range(7, ws.max_row + 1):
        cel_naam = ws.cell(r, 3).value
        if cel_naam is None:
            continue
        m = re.match(r"\s*(\d+)\s+(.+)", str(cel_naam))
        if not m:
            continue
        nr = m.group(1).strip()
        for ci, dt in enumerate(datums):
            if not dt:
                continue
            v = ws.cell(r, 4 + ci).value
            if v is None or v == "":
                continue
            try:
                resultaat[(nr, dt)] = float(v)
            except (TypeError, ValueError):
                pass
    return resultaat


def main():
    print(f"Lezen Nitea: {NITEA_PDF.name}")
    nitea = parse_nitea(NITEA_PDF)
    print(f"  {len(nitea)} (medewerker × dag) regels\n")

    print(f"Lezen Excel: {EXCEL.name}")
    excel = parse_excel(EXCEL)
    print(f"  {len(excel)} (medewerker × dag) regels\n")

    # Vergelijk waar beide aanwezig zijn
    samen = set(nitea) & set(excel)
    alleen_nitea = set(nitea) - set(excel)
    alleen_excel = set(excel) - set(nitea)
    print(f"In beide bronnen           : {len(samen):4d}")
    print(f"Alleen in Nitea (geen Excel-uur): {len(alleen_nitea):4d}")
    print(f"Alleen in Excel (geen Nitea)   : {len(alleen_excel):4d}\n")

    correcties: list[tuple[float, str, str, float, float]] = []
    voor_counter: Counter = Counter()
    for key in samen:
        n = nitea[key]
        e = excel[key]
        delta = round(e - n, 4)
        correcties.append((delta, key[0], key[1], n, e))
        # Bucketten in 0.25u stappen
        voor_counter[round(delta * 4) / 4] += 1

    print("=== Verdeling delta (Excel - Nitea) in uren ===")
    for d in sorted(voor_counter):
        bar = "█" * min(voor_counter[d], 70)
        teken = "+" if d > 0 else ""
        print(f"  {teken}{d:+5.2f}u : {voor_counter[d]:4d}  {bar}")

    omhoog = [c for c in correcties if c[0] > 0]
    omlaag = [c for c in correcties if c[0] < 0]
    gelijk = [c for c in correcties if c[0] == 0]
    print()
    print(f"=== Samenvatting ===")
    print(f"  exact gelijk (delta = 0)  : {len(gelijk):4d}")
    print(f"  Excel hoger dan Nitea (+) : {len(omhoog):4d}")
    print(f"  Excel lager dan Nitea (-) : {len(omlaag):4d}")
    print(f"  totaal vergeleken         : {len(correcties):4d}")
    if omhoog:
        som_omhoog = sum(c[0] for c in omhoog)
        print(f"  som ophoging Excel        : +{som_omhoog:.2f}u")
    if omlaag:
        som_omlaag = sum(c[0] for c in omlaag)
        print(f"  som verlaging Excel       : {som_omlaag:.2f}u")

    if omhoog:
        print()
        print("=== Top 15 Excel-ophoging tov Nitea ===")
        for d, nr, dt, n, e in sorted(omhoog, key=lambda x: -x[0])[:15]:
            print(f"  +{d:5.2f}u : {nr:>4s} {dt}  Nitea={n:5.2f}  Excel={e:5.2f}")

    if omlaag:
        print()
        print("=== Top 15 Excel-verlaging tov Nitea ===")
        for d, nr, dt, n, e in sorted(omlaag, key=lambda x: x[0])[:15]:
            print(f"  {d:5.2f}u : {nr:>4s} {dt}  Nitea={n:5.2f}  Excel={e:5.2f}")

    if alleen_nitea:
        print()
        print(f"=== Eerste 10 in Nitea maar NIET in Excel ===")
        for nr, dt in sorted(alleen_nitea)[:10]:
            print(f"  {nr:>4s} {dt}  Nitea={nitea[(nr, dt)]:.2f}u")

    if alleen_excel:
        print()
        print(f"=== Eerste 10 in Excel maar NIET in Nitea ===")
        for nr, dt in sorted(alleen_excel)[:10]:
            print(f"  {nr:>4s} {dt}  Excel={excel[(nr, dt)]:.2f}u")


if __name__ == "__main__":
    main()
