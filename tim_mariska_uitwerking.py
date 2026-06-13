"""Werk het 'Tim/Mariska'-deel uit van Ola's bijgewerkte afwijkingenlijst.

Input : Afwijkingen_werklijst_DB (1).xlsx  (tab 'UZK zonder SNOOP', met Ola's kolommen
        J='ola', K='Tim/Mariska', L=notitie met juiste schaal).
Output: Tim-Mariska_uitwerking.xlsx — per 'Tim/Mariska'-rij de definitieve schaal, of het
        UZB-tarief daarbij klopt, wat nog ingevoerd moet worden, en bijzonderheden
        (tag-conflict / geen tag / datumwissel / HR-overleg).
"""
import re
import openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Dag-bewuste tag-occupatie hergebruiken uit de engine (één bron van waarheid).
from controle_uzb import bouw_dag_occupatie, DAGNAMEN

MAP = r'C:\Users\dieter.KWEKERIJBAAS\Kwekerij Baas\Finance - Controle 2026'
DL = r'C:\Users\dieter.KWEKERIJBAAS\Downloads'
IN = DL + r'\Afwijkingen_werklijst_DB (1).xlsx'
OUT = DL + r'\Tim-Mariska_uitwerking.xlsx'

FONT_TITLE = Font(name='Arial', size=13, bold=True, color='FFFFFF')
FONT_H = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_N = Font(name='Arial', size=10)
FILL_TITLE = PatternFill('solid', start_color='1F4E78')
FILL_H = PatternFill('solid', start_color='2E75B6')
FILL_OK = PatternFill('solid', start_color='C6EFCE')
FILL_GEEL = PatternFill('solid', start_color='FFEB9C')
FILL_ROOD = PatternFill('solid', start_color='FFC7CE')
FILL_BLAUW = PatternFill('solid', start_color='B4C7E7')
B = Border(*[Side(style='thin', color='BFBFBF')] * 4)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)
C = Alignment(horizontal='center', vertical='center')


def norm_schaal(s):
    """'B2 flex'->'B2F', 'B2 vast'->'B2V', 'B2 sw/seizoen'->'B2S', 'B2F'->'B2F', '15B2'->'15B2'."""
    if not s:
        return ''
    t = str(s).strip()
    if t.lower() in ('stond erin', 'none', 'nan', '?'):
        return ''
    m = re.match(r'^\s*(\d*[A-Za-z]\d+)\s*(.*)$', t)
    if not m:
        return t.upper()
    basis = m.group(1).upper()
    rest = m.group(2).lower()
    if basis[-1].isalpha():            # al een suffix (B2F)
        return basis
    if 'flex' in rest:
        return basis + 'F'
    if 'vast' in rest:
        return basis + 'V'
    if 'seizoen' in rest or rest.strip() in ('s', 'sw'):
        return basis + 'S'
    return basis


def parse_ola_note(note):
    """('B3F', '15 juni') uit 'is B4 flex per 15 juni word B4 vast' / 'is 17B2 per 1 mei'."""
    if not note:
        return '', ''
    t = str(note).strip()
    datum = ''
    md = re.search(r'per\s+([0-9]{1,2}[-\s][0-9A-Za-z]+(?:[-\s][0-9]{2,4})?)', t)
    if md:
        datum = md.group(1).strip()
    # eerste schaal-token
    msch = re.search(r'(\d*[A-Za-z]\d+(?:\s*(?:flex|vast|seizoen|sw))?)', t, re.I)
    sch = norm_schaal(msch.group(1)) if msch else ''
    return sch, datum


wb = openpyxl.load_workbook(IN)
ws = wb['UZK zonder SNOOP']

# DAG-bewuste tag-occupatie uit de doorgegeven (Nitea) bestanden.
# Tag-hergebruik is NORMAAL (Nitea hergebruikt tags); een tag is uniek per DAG, niet per week —
# binnen één week kan dezelfde tag aan meerdere UZK gegeven worden (de een ma-di, de ander wo-vr).
# Een ECHT conflict = zelfde tag, ZELFDE DAG, twee verschillende personen.
occ = bouw_dag_occupatie(MAP)              # (tag, jaar, week, dag) -> set(naam_norm)
# Per tag: alle namen die hem ooit gebruikten + de dagen-met-conflict
tag_personen = defaultdict(set)
tag_dagconflict = defaultdict(list)        # tag -> [(jaar, week, dag, {namen})]
for (tag, jaar, week, d), namen in occ.items():
    tag_personen[tag].update(namen)
    if len(namen) > 1:
        tag_dagconflict[tag].append((jaar, week, d, set(namen)))


def _surn(s):
    return {t for t in re.sub(r'[().,\.]', ' ', str(s).lower()).split() if len(t) >= 3 and t.isalpha()}


def tag_status(uzb, tag, naam):
    """('dagconflict'|'hergebruik'|'uniek', tekst) — dag-bewust."""
    tag = str(tag)
    eigen_surn = _surn(naam)
    alle = tag_personen.get(tag, set())
    anderen = {n for n in alle if not (_surn(n) & eigen_surn)}
    # echte conflicten op dezelfde dag waar DEZE persoon bij betrokken is
    samedag = []
    for jaar, week, d, namen in tag_dagconflict.get(tag, []):
        ik = any(_surn(n) & eigen_surn for n in namen)
        co = {n for n in namen if not (_surn(n) & eigen_surn)}
        if ik and co:
            dag = DAGNAMEN[d] if 0 <= d < 7 else f'd{d}'
            samedag.append(f'wk{week} {dag}: ook {", ".join(sorted(co))}')
    if samedag:
        return 'dagconflict', ' ; '.join(samedag[:3])
    if anderen:
        return 'hergebruik', ', '.join(sorted(anderen))
    return 'uniek', ''

rows = []
for r in range(5, ws.max_row + 1):
    if str(ws.cell(row=r, column=11).value) != 'Tim/Mariska':
        continue
    naam = ws.cell(row=r, column=1).value
    uzb = ws.cell(row=r, column=2).value
    tag = ws.cell(row=r, column=3).value
    s_snoop = ws.cell(row=r, column=4).value
    s_wb = ws.cell(row=r, column=5).value
    s_uzb = ws.cell(row=r, column=6).value
    actie = ws.cell(row=r, column=9).value
    ola = ws.cell(row=r, column=12).value

    n_snoop = norm_schaal(s_snoop)
    n_wb = norm_schaal(s_wb)
    n_uzb = norm_schaal(s_uzb)
    ola_sch, ola_datum = parse_ola_note(ola)

    # Definitieve schaal: Ola-notitie > SNOOP > UZB > Werkbestand
    if ola_sch:
        definitief, bron = ola_sch, 'Ola-notitie'
    elif n_snoop:
        definitief, bron = n_snoop, 'SNOOP'
    elif n_uzb:
        definitief, bron = n_uzb, 'UZB-tarief (afgeleid)'
    elif n_wb:
        definitief, bron = n_wb, 'Werkbestand'
    else:
        definitief, bron = '', 'ONBEKEND'

    # UZB-tarief klopt? = stemt de definitieve schaal overeen met de UZB-schaal?
    if not n_uzb:
        tarief_ok = 'geen UZB-schaal — handmatig'
        kl = FILL_BLAUW
    elif not definitief:
        tarief_ok = 'schaal onbekend — HR'
        kl = FILL_BLAUW
    elif definitief == n_uzb:
        tarief_ok = 'JA — UZB-tarief klopt bij schaal'
        kl = FILL_OK
    else:
        tarief_ok = f'NEE — UZB rekent {n_uzb}, hoort {definitief}'
        kl = FILL_ROOD

    # Nog invoeren in
    invoer = []
    if not n_snoop and str(s_snoop).strip().lower() != 'stond erin':
        invoer.append('SNOOP')
    if not n_wb:
        invoer.append('Werkbestand')
    invoer_txt = ' + '.join(invoer) if invoer else 'niets meer (al ingevoerd)'

    # Bijzonderheden (week-bewust)
    bijz = []
    hard = False  # echte blokkade
    if tag and str(tag) not in ('None', '?', ''):
        st, anderen = tag_status(uzb, tag, naam)
        if st == 'dagconflict':
            bijz.append(f'TAG-CONFLICT op dezelfde DAG: tag {tag} — {anderen} — UITZOEKEN (tag hoort per dag bij 1 persoon)')
            hard = True
        elif st == 'hergebruik':
            bijz.append(f'tag {tag} op andere dagen door {anderen} gebruikt — normaal hergebruik, geen blokkade')
    if not tag or str(tag) in ('None', '?', ''):
        bijz.append('TAGNUMMER ONTBREEKT — eerst tag opzoeken in Nitea')
        hard = True
    if ola_datum:
        bijz.append(f'INGANGSDATUM: schaalwissel per {ola_datum} — factuur t/m die datum oude tarief, daarna nieuw')
        hard = True
    if 'overleg HR' in str(actie):
        bijz.append('OVERLEG HR — geen sluitende inschaling')
        hard = True
    bijz_txt = ' | '.join(bijz)
    if hard and kl is FILL_OK:
        kl = FILL_GEEL  # geen tariefafwijking, maar wél een aandachtspunt

    rows.append({'naam': naam, 'uzb': uzb, 'tag': '' if str(tag) in ('None', '?') else tag,
                 'definitief': definitief, 'bron': bron, 'tarief_ok': tarief_ok,
                 'invoer': invoer_txt, 'datum': ola_datum, 'bijz': bijz_txt, 'kl': kl, 'hard': hard})

# ── schrijven ──
out = openpyxl.Workbook()
o = out.active
o.title = 'Tim-Mariska uitwerking'
o.merge_cells('A1:I1')
o['A1'] = 'Tim/Mariska — uitwerking van de openstaande inschalingen (62 rijen uit Ola\'s lijst)'
o['A1'].font = FONT_TITLE; o['A1'].fill = FILL_TITLE; o['A1'].alignment = C

n_ok = sum(1 for x in rows if x['tarief_ok'].startswith('JA'))
n_afw = sum(1 for x in rows if x['tarief_ok'].startswith('NEE'))
n_hr = sum(1 for x in rows if 'HR' in x['tarief_ok'] or 'handmatig' in x['tarief_ok'])
n_conf = sum(1 for x in rows if 'TAG-CONFLICT op dezelfde DAG' in x['bijz'])
n_herg = sum(1 for x in rows if 'normaal hergebruik' in x['bijz'])
n_geentag = sum(1 for x in rows if 'TAGNUMMER ONTBREEKT' in x['bijz'])
o.merge_cells('A2:I2')
o['A2'] = (f'{len(rows)} medewerkers. UZB-tarief klopt bij de (nu bekende) schaal: {n_ok}  ·  '
           f'tariefafwijking: {n_afw}  ·  HR/handmatig nodig: {n_hr}. '
           f'Tagnummers worden in Nitea HERGEBRUIKT (uniek per DAG, niet over tijd; binnen een week mag '
           f'dezelfde tag aan meerdere UZK): {n_herg} tags zijn op andere dagen door iemand anders gebruikt = '
           f'normaal, GEEN blokkade. Echte tag-conflicten (zelfde tag, ZELFDE DAG, 2 personen): {n_conf}. '
           f'Zonder tagnummer: {n_geentag}. '
           f'"UZB-tarief klopt" = de gefactureerde schaal is gelijk aan de definitieve schaal, dus geen '
           f'overfacturatie — het was alleen een ontbrekende registratie in het werkbestand.')
o['A2'].font = Font(name='Arial', size=9, italic=True); o['A2'].alignment = L
o.row_dimensions[2].height = 58

hdr = ['Naam', 'UZB', 'Tag', 'Definitieve schaal', 'Bron schaal', 'UZB-tarief klopt?',
       'Nog invoeren in', 'Ingangsdatum', 'Bijzonderheid / actie']
for i, h in enumerate(hdr, start=1):
    c = o.cell(row=4, column=i, value=h); c.font = FONT_H; c.fill = FILL_H; c.alignment = C; c.border = B

rr = 5
# sorteer: exceptions eerst (afwijking/HR/conflict), dan de rest, per UZB
def prio(x):
    if x['tarief_ok'].startswith('NEE'):
        return 0
    if 'HR' in x['tarief_ok'] or 'handmatig' in x['tarief_ok']:
        return 1
    if x['hard']:
        return 2
    return 3
for x in sorted(rows, key=lambda x: (prio(x), str(x['uzb']), str(x['naam']))):
    vals = [x['naam'], x['uzb'], x['tag'], x['definitief'], x['bron'], x['tarief_ok'],
            x['invoer'], x['datum'], x['bijz']]
    for i, v in enumerate(vals, start=1):
        c = o.cell(row=rr, column=i, value=v); c.font = FONT_N; c.border = B; c.fill = x['kl']
        c.alignment = C if i in (2, 3, 4) else L
    rr += 1

o.auto_filter.ref = f'A4:I{rr - 1}'
o.freeze_panes = 'A5'
for i, w in enumerate([26, 6, 7, 16, 18, 30, 22, 16, 56]):
    o.column_dimensions[chr(65 + i)].width = w

out.save(OUT)
print(f'Geschreven: {OUT}')
print(f'  {len(rows)} rijen | {n_ok} tarief klopt | {n_afw} afwijking | {n_hr} HR/handmatig')
print(f'  {n_conf} tag-conflicten | {n_geentag} zonder tag')
print('\nEXCEPTIES (afwijking / HR / conflict):')
for x in sorted(rows, key=lambda x: (prio(x), str(x['uzb']))):
    if prio(x) <= 2:
        print(f"  [{x['uzb']}] {x['naam']} (tag {x['tag'] or '-'}) -> {x['definitief'] or '?'} | "
              f"{x['tarief_ok']} | {x['bijz'][:70]}")
