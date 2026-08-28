"""Urenoverzicht per uitzendbureau als Excel-bestand (SPEC §3).

Tabbladen: Totaal week (opent als eerste), Per dag, Tarieven en Afwijkingen.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.calc.engine import minuten_per_bron, rond_op_kwartier
from app.services.tarief import Kaartreeks, TariefKaart
from app.services.verwerking import WeekVerwerking

# Voor wie geen tarief heeft is er geen tariefcategorie; toon dan de toeslag-
# bron zelf, zodat de uren herkenbaar blijven.
_BRON_LABEL = {
    "normaal": "100",
    "overwerk_35": "100",
    "nacht": "nachtuur",
    "avond": "150",
    "zaterdag_middag": "150",
    "dag_grens_50": "150",
    "week_grens_50": "150",
    "zondag": "200",
    "feestdag": "feestdag",
}

_KOP = Font(bold=True, color="FFFFFF")
# Magenta uit de huisstijl van Kwekerij Baas (kernwaardenblokken op de site).
_KOP_VULLING = PatternFill("solid", fgColor="D6007F")
_TITEL = Font(bold=True, size=13)
_EURO = '#,##0.00'
_UUR = "0.00"

_DAGEN = ("maandag", "dinsdag", "woensdag", "donderdag", "vrijdag", "zaterdag", "zondag")

# Wat er moet gebeuren om een afwijking op te lossen, per soort. De meldingen
# (onderaan het tabblad) dragen hun vervolgstap al in de tekst zelf.
_AFWIJKING_ACTIE = {
    "registratie_inconsistent": (
        "Controleer in Nitea of de onderbreking klopt. Zo niet: registratie "
        "laten corrigeren en de week opnieuw verwerken; klopt hij wel, dan is "
        "er niets te doen -- de uren zijn al juist geteld."
    ),
    "geen_registratie": (
        "Wel gepland, niet geklokt. Niet gewerkt: niets doen. Vergeten te "
        "klokken: Nitea laten aanvullen en de week opnieuw verwerken."
    ),
    "geen_planning": (
        "Gewerkt zonder planning. Controleer of de dienst terecht was; de uren "
        "tellen gewoon mee."
    ),
    "uren_verschil": (
        "Planning en registratie verschillen. Nitea is leidend; alleen "
        "controleren als het verschil onverwacht groot is."
    ),
    "tijd_verschil": (
        "Begin- of eindtijd wijkt af van de planning. Nitea is leidend; alleen "
        "controleren bij structurele afwijkingen."
    ),
}


def _kop(ws, rij: int, koppen: list[str], bevries: bool = True) -> None:
    for kolom, tekst in enumerate(koppen, start=1):
        cel = ws.cell(row=rij, column=kolom, value=tekst)
        cel.font = _KOP
        cel.fill = _KOP_VULLING
        cel.alignment = Alignment(horizontal="center", wrap_text=True)
    if bevries:
        ws.freeze_panes = ws.cell(row=rij + 1, column=1)


def _breedtes(ws, breedtes: list[int]) -> None:
    for i, breedte in enumerate(breedtes, start=1):
        ws.column_dimensions[get_column_letter(i)].width = breedte


def _categorieen(verwerking: WeekVerwerking, conventies=None) -> list[str]:
    gezien: set[str] = set()
    for medewerker in verwerking.medewerkers:
        gezien.update(r.categorie for r in medewerker.bedrag.regels)
        gezien.update(_uren_zonder_tarief(medewerker))
    return sorted(gezien)


def _uren_per_categorie(medewerker) -> dict[str, Decimal]:
    """Uren per tariefcategorie, opgeteld over de tariefperiodes.

    Gaat er midden in de week een nieuwe loontabel in, dan levert één categorie
    twee regels op met verschillende tarieven; in het overzicht hoort dat één
    getal te zijn."""
    per_categorie: dict[str, Decimal] = {}
    for regel in medewerker.bedrag.regels:
        per_categorie[regel.categorie] = (
            per_categorie.get(regel.categorie, Decimal("0")) + regel.uren
        )
    return per_categorie


def _uren_zonder_tarief(medewerker) -> dict[str, Decimal]:
    """Uren per categorie voor wie geen tarief heeft.

    Zonder loonschaal levert de bedragberekening geen regels op. De uren zijn
    dan wel gewerkt, dus die horen zichtbaar te blijven -- anders lijkt de
    medewerker nul uur te hebben gewerkt terwijl het weektotaal ze wel meetelt.
    """
    if medewerker.bedrag.regels:
        return {}
    per_bron = rond_op_kwartier(minuten_per_bron(medewerker.resultaat.trace))
    per_categorie: dict[str, Decimal] = {}
    for bron, minuten in per_bron.items():
        categorie = _BRON_LABEL.get(bron, bron)
        per_categorie[categorie] = per_categorie.get(categorie, Decimal("0")) + (
            Decimal(minuten) / Decimal(60)
        ).quantize(Decimal("0.01"))
    return per_categorie


def bouw_overzicht(
    verwerking: WeekVerwerking,
    uzb_naam: str,
    kaart: TariefKaart | Kaartreeks | None = None,
    controle=None,
) -> bytes:
    wb = Workbook()
    categorieen = _categorieen(verwerking)

    # --- Totaal week ---------------------------------------------------- #
    ws = wb.active
    ws.title = "Totaal week"
    ws["A1"] = f"Weektotaal per medewerker — {uzb_naam} week {verwerking.iso_week}/{verwerking.iso_jaar}"
    ws["A1"].font = _TITEL
    koppen = ["Medewerker", "Nitea-ID", "Loonschaal"]
    koppen += [f"Uren {c}" for c in categorieen] + ["Totaal uren", "Bedrag (€)"]
    _kop(ws, 3, koppen)

    rij = 4
    for medewerker in verwerking.medewerkers:
        per_categorie = _uren_per_categorie(medewerker)
        per_categorie.update(_uren_zonder_tarief(medewerker))
        ws.cell(row=rij, column=1, value=medewerker.naam)
        ws.cell(row=rij, column=2, value=medewerker.nitea_id)
        ws.cell(row=rij, column=3, value=medewerker.loonschaal)
        for i, categorie in enumerate(categorieen):
            cel = ws.cell(row=rij, column=4 + i, value=float(per_categorie.get(categorie, 0)))
            cel.number_format = _UUR
        totaal = ws.cell(row=rij, column=4 + len(categorieen), value=float(medewerker.netto_uren))
        totaal.number_format = _UUR
        bedrag = ws.cell(
            row=rij, column=5 + len(categorieen), value=float(medewerker.bedrag.totaal)
        )
        bedrag.number_format = _EURO
        rij += 1

    ws.cell(row=rij, column=1, value="TOTAAL").font = Font(bold=True)
    totaal_uren = ws.cell(row=rij, column=4 + len(categorieen), value=float(verwerking.totaal_uren))
    totaal_uren.font = Font(bold=True)
    totaal_uren.number_format = _UUR
    totaal_bedrag = ws.cell(
        row=rij, column=5 + len(categorieen), value=float(verwerking.totaal_bedrag)
    )
    totaal_bedrag.font = Font(bold=True)
    totaal_bedrag.number_format = _EURO
    _breedtes(ws, [26, 10, 14] + [11] * (len(categorieen) + 1) + [13])

    # --- Per dag --------------------------------------------------------- #
    ws2 = wb.create_sheet("Per dag")
    ws2["A1"] = f"Registratie per dag — {uzb_naam} week {verwerking.iso_week}/{verwerking.iso_jaar}"
    ws2["A1"].font = _TITEL
    _kop(ws2, 3, ["Medewerker", "Datum", "Dag", "Begin", "Eind", "Pauze (min)", "Uren"])
    rij = 4
    for medewerker in verwerking.medewerkers:
        for segment in sorted(
            {s.datum for s in medewerker.resultaat.trace}
        ):
            minuten = sum(
                s.minuut_tot - s.minuut_van
                for s in medewerker.resultaat.trace
                if s.datum == segment
            )
            ws2.cell(row=rij, column=1, value=medewerker.naam)
            ws2.cell(row=rij, column=2, value=segment).number_format = "dd-mm-yyyy"
            ws2.cell(row=rij, column=3, value=_DAGEN[segment.weekday()])
            cel = ws2.cell(row=rij, column=7, value=round(minuten / 60, 2))
            cel.number_format = _UUR
            rij += 1
    _breedtes(ws2, [26, 12, 12, 9, 9, 12, 10])

    # --- Tarieven -------------------------------------------------------- #
    ws3 = wb.create_sheet("Tarieven")
    ws3["A1"] = f"Gebruikte tarieven — {uzb_naam}"
    ws3["A1"].font = _TITEL
    reeks = kaart if isinstance(kaart, Kaartreeks) else Kaartreeks.van_kaart(kaart)
    rij = 2
    if reeks.is_leeg:
        ws3["A3"] = "Geen tariefkaart beschikbaar voor deze week."
    else:
        # Eén blok per tariefperiode: gaat er midden in de week een nieuwe
        # loontabel in, dan is per dag te zien welke tarieven zijn gebruikt.
        for vanaf, periode_kaart in reeks.periodes:
            if periode_kaart is None:
                continue
            wanneer = (
                f"Uren vanaf {vanaf:%d-%m-%Y}: " if len(reeks.periodes) > 1 else ""
            )
            ws3.cell(row=rij, column=1, value=(
                f"{wanneer}afgeleid uit de CAO-loontabel geldig vanaf "
                f"{periode_kaart.geldig_van:%d-%m-%Y}."
            ))
            _kop(ws3, rij + 2, ["Loonschaal"] + categorieen, bevries=rij == 2)
            rij += 3
            for code in sorted(periode_kaart.schalen):
                schaal = periode_kaart.schalen[code]
                ws3.cell(row=rij, column=1, value=code)
                for i, categorie in enumerate(categorieen):
                    tarief = schaal.tarief(categorie)
                    if tarief is not None:
                        cel = ws3.cell(row=rij, column=2 + i, value=float(tarief))
                        cel.number_format = _EURO
                rij += 1
            rij += 2
        _breedtes(ws3, [30] + [12] * len(categorieen))

    # --- Afwijkingen ----------------------------------------------------- #
    ws4 = wb.create_sheet("Afwijkingen")
    ws4["A1"] = "Afwijkingen en aandachtspunten"
    ws4["A1"].font = _TITEL
    _kop(ws4, 3, ["Medewerker", "Datum", "Soort", "Toelichting", "Wat te doen"])
    rij = 4
    for medewerker in verwerking.medewerkers:
        for afwijking in medewerker.afwijkingen:
            ws4.cell(row=rij, column=1, value=medewerker.naam)
            ws4.cell(row=rij, column=2, value=afwijking.datum).number_format = "dd-mm-yyyy"
            ws4.cell(row=rij, column=3, value=afwijking.soort)
            ws4.cell(row=rij, column=4, value=afwijking.detail)
            ws4.cell(row=rij, column=5, value=_AFWIJKING_ACTIE.get(afwijking.soort, ""))
            rij += 1
    for melding in verwerking.meldingen:
        ws4.cell(row=rij, column=3, value="melding")
        ws4.cell(row=rij, column=4, value=melding)
        rij += 1
    _breedtes(ws4, [26, 12, 24, 70, 60])

    if controle is not None:
        voeg_factuurcontrole_toe(wb, controle)

    wb.active = 0  # opent op 'Totaal week'
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def voeg_factuurcontrole_toe(wb, controle) -> None:
    """Extra tabblad met de vergelijking tussen overzicht en factuur."""
    ws = wb.create_sheet("Factuurcontrole")
    ws["A1"] = (
        f"Factuurcontrole — {controle.uzb_naam} week "
        f"{controle.iso_week}/{controle.iso_jaar}"
    )
    ws["A1"].font = _TITEL
    if controle.factuurnummers:
        ws["A2"] = "Factuur: " + ", ".join(controle.factuurnummers)

    _kop(ws, 4, ["", "Ons overzicht", "Factuur", "Verschil"])
    for i, (label, ons, factuur) in enumerate(
        [
            ("Uren", controle.uren_overzicht, controle.uren_factuur),
            ("Bedrag (€)", controle.bedrag_overzicht, controle.bedrag_factuur),
        ]
    ):
        rij = 5 + i
        ws.cell(row=rij, column=1, value=label)
        for kolom, waarde in enumerate([ons, factuur, factuur - ons], start=2):
            cel = ws.cell(row=rij, column=kolom, value=float(waarde))
            cel.number_format = _UUR if label == "Uren" else _EURO

    ws.cell(row=8, column=1, value=f"Bevindingen ({len(controle.bevindingen)})").font = _TITEL
    _kop(ws, 9, ["Soort", "Medewerker", "Uren ons", "Uren factuur",
                 "Bedrag ons", "Bedrag factuur", "Toelichting", "Wat te doen"])
    rij = 10
    for bevinding in controle.bevindingen:
        ws.cell(row=rij, column=1, value=bevinding.soort)
        ws.cell(row=rij, column=2, value=bevinding.naam)
        for kolom, waarde, opmaak in [
            (3, bevinding.uren_overzicht, _UUR),
            (4, bevinding.uren_factuur, _UUR),
            (5, bevinding.bedrag_overzicht, _EURO),
            (6, bevinding.bedrag_factuur, _EURO),
        ]:
            if waarde is not None:
                cel = ws.cell(row=rij, column=kolom, value=float(waarde))
                cel.number_format = opmaak
        ws.cell(row=rij, column=7, value=bevinding.melding)
        ws.cell(row=rij, column=8, value=bevinding.actie)
        rij += 1
    _breedtes(ws, [22, 26, 11, 12, 12, 13, 60, 55])


def bestandsnaam(uzb_naam: str, verwerking: WeekVerwerking) -> str:
    veilig = "".join(c if c.isalnum() else "_" for c in uzb_naam).strip("_")
    return f"UZB-overzicht_{veilig}_week_{verwerking.iso_week}_{verwerking.iso_jaar}.xlsx"


def bouw_matchingsbestand(controle, bevindingen_tekst: str = "") -> bytes:
    """Los matchingsbestand: de bewaarde week naast de ontvangen facturen."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Samenvatting"
    ws["A1"] = (
        f"Factuurcontrole — {controle.uzb_naam} week "
        f"{controle.iso_week}/{controle.iso_jaar}"
    )
    ws["A1"].font = _TITEL
    if controle.factuurnummers:
        ws["A2"] = "Facturen: " + ", ".join(controle.factuurnummers)

    _kop(ws, 4, ["", "Ons overzicht", "Factuur", "Verschil"])
    for i, (label, ons, factuur, opmaak) in enumerate(
        [
            ("Uren", controle.uren_overzicht, controle.uren_factuur, _UUR),
            ("Bedrag (€)", controle.bedrag_overzicht, controle.bedrag_factuur, _EURO),
        ]
    ):
        rij = 5 + i
        ws.cell(row=rij, column=1, value=label)
        for kolom, waarde in enumerate([ons, factuur, factuur - ons], start=2):
            cel = ws.cell(row=rij, column=kolom, value=float(waarde))
            cel.number_format = opmaak
    ws.cell(row=7, column=1, value="Medewerkers gekoppeld")
    ws.cell(row=7, column=2, value=len(controle.koppelingen))
    ws.cell(row=8, column=1, value="Bevindingen")
    ws.cell(row=8, column=2, value=len(controle.bevindingen))
    _breedtes(ws, [24, 15, 15, 13])

    # --- Bevindingen ------------------------------------------------------ #
    ws2 = wb.create_sheet("Bevindingen")
    ws2["A1"] = f"Bevindingen ({len(controle.bevindingen)})"
    ws2["A1"].font = _TITEL
    _kop(ws2, 3, ["Soort", "Medewerker", "Uren ons", "Uren factuur", "Verschil",
                  "Bedrag ons", "Bedrag factuur", "Verschil", "Toelichting",
                  "Wat te doen"])
    rij = 4
    for bevinding in controle.bevindingen:
        ws2.cell(row=rij, column=1, value=bevinding.soort)
        ws2.cell(row=rij, column=2, value=bevinding.naam)
        for kolom, waarde, opmaak in [
            (3, bevinding.uren_overzicht, _UUR),
            (4, bevinding.uren_factuur, _UUR),
            (5, bevinding.uren_verschil, _UUR),
            (6, bevinding.bedrag_overzicht, _EURO),
            (7, bevinding.bedrag_factuur, _EURO),
            (8, bevinding.bedrag_verschil, _EURO),
        ]:
            if waarde is not None:
                cel = ws2.cell(row=rij, column=kolom, value=float(waarde))
                cel.number_format = opmaak
        ws2.cell(row=rij, column=9, value=bevinding.melding)
        ws2.cell(row=rij, column=10, value=bevinding.actie)
        rij += 1
    _breedtes(ws2, [20, 26, 10, 12, 10, 12, 13, 10, 55, 50])

    # --- Alle koppelingen ------------------------------------------------- #
    ws3 = wb.create_sheet("Koppelingen")
    ws3["A1"] = "Medewerker naast factuurregel"
    ws3["A1"].font = _TITEL
    _kop(ws3, 3, ["Medewerker", "Naam op factuur", "Loonschaal", "Uren ons",
                  "Uren factuur", "Bedrag ons", "Bedrag factuur"])
    rij = 4
    for medewerker, kracht in sorted(controle.koppelingen, key=lambda p: p[0].naam):
        ws3.cell(row=rij, column=1, value=medewerker.naam)
        ws3.cell(row=rij, column=2, value=kracht.naam_ruw)
        ws3.cell(row=rij, column=3, value=medewerker.loonschaal)
        for kolom, waarde, opmaak in [
            (4, medewerker.netto_uren, _UUR),
            (5, kracht.uren, _UUR),
            (6, medewerker.bedrag.totaal, _EURO),
            (7, kracht.bedrag, _EURO),
        ]:
            cel = ws3.cell(row=rij, column=kolom, value=float(waarde))
            cel.number_format = opmaak
        rij += 1
    _breedtes(ws3, [26, 28, 14, 10, 12, 12, 13])

    # --- Bevindingenmail --------------------------------------------------- #
    if bevindingen_tekst:
        ws4 = wb.create_sheet("Bevindingenmail")
        ws4["A1"] = "Concepttekst om door te sturen"
        ws4["A1"].font = _TITEL
        for i, regel in enumerate(bevindingen_tekst.split("\n"), start=3):
            ws4.cell(row=i, column=1, value=regel)
        _breedtes(ws4, [110])

    wb.active = 0
    return _naar_bytes(wb)


def _naar_bytes(wb) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
