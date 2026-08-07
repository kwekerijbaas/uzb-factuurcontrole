"""SNOOP-export (.xlsx) inlezen: geplande inzet + loonschaal per medewerker.

Kolommen (rij 1 = header, data vanaf rij 2):
    Registratienummer | Medewerker | Datum | Starttijd | Eindtijd |
    Werkelijke starttijd | Werkelijke eindtijd | Gewerkte uren | Locatie |
    Werkgever op datum shift | Type uitzendkracht | Tarief uitzendbureau
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.calc.types import PlanningRegel

# header-namen -> interne sleutel (case-insensitief, spaties genegeerd)
_KOLOMMEN = {
    "medewerker": "medewerker",
    "datum": "datum",
    "starttijd": "start",
    "eindtijd": "eind",
    "gewerkteuren": "uren",
    "tariefuitzendbureau": "loonschaal",
    "werkgeveropdatumshift": "werkgever",
}

_VERPLICHT = {"medewerker", "datum", "start", "eind"}
# Hoeveel rijen er naar de kopregel wordt gezocht voordat het bestand als
# onbruikbaar geldt.
_MAX_KOPREGEL = 15


def _norm_naam(naam: str) -> str:
    return re.sub(r"\s+", " ", str(naam)).strip()


def _als_tijd(waarde) -> time | None:
    if waarde is None or waarde == "":
        return None
    if isinstance(waarde, time):
        return waarde
    if isinstance(waarde, datetime):
        return waarde.time()
    m = re.match(r"^(\d{1,2}):(\d{2})", str(waarde).strip())
    if not m:
        return None
    return time(int(m.group(1)), int(m.group(2)))


def _als_datum(waarde) -> date | None:
    if isinstance(waarde, datetime):
        return waarde.date()
    if isinstance(waarde, date):
        return waarde
    if waarde is None:
        return None
    s = str(waarde).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _als_minuten(uren_waarde) -> int | None:
    """'Gewerkte uren' (bv. 8.25) -> minuten (netto, pauze er al af)."""
    if uren_waarde is None or uren_waarde == "":
        return None
    try:
        return int((Decimal(str(uren_waarde)) * 60).to_integral_value())
    except (InvalidOperation, ValueError):
        return None


@dataclass
class SnoopMedewerker:
    naam: str
    loonschaal: str | None
    planning: list[PlanningRegel] = field(default_factory=list)
    # Naam van het uitzendbureau zoals SNOOP hem noteert; wordt gebruikt om te
    # controleren dat het bestand bij het gekozen bureau hoort.
    werkgever: str | None = None


def lees_snoop(bron: str | Path | bytes) -> list[SnoopMedewerker]:
    """Parse een SNOOP-export naar één SnoopMedewerker per medewerker."""
    data = BytesIO(bron) if isinstance(bron, (bytes, bytearray)) else bron
    wb = load_workbook(data, data_only=True)
    # De kopregel staat niet altijd op de eerste rij: exports beginnen soms met
    # een titel of lege regels. Daarom wordt er per tabblad naar gezocht.
    ws = None
    idx: dict[str, int] = {}
    rijen = iter(())
    gezien: list[str] = []
    for blad in wb.worksheets:
        alle = list(blad.iter_rows(max_row=_MAX_KOPREGEL, values_only=True))
        for nummer, header in enumerate(alle, start=1):
            kandidaat = {
                _KOLOMMEN[sleutel]: i
                for i, cel in enumerate(header)
                if (sleutel := re.sub(r"\s+", "", str(cel or "").lower())) in _KOLOMMEN
            }
            if _VERPLICHT <= kandidaat.keys():
                ws, idx = blad, kandidaat
                rijen = blad.iter_rows(min_row=nummer + 1, values_only=True)
                break
            gezien += [str(c).strip() for c in header if c and str(c).strip()]
        if ws is not None:
            break

    if ws is None:
        gevonden = ", ".join(dict.fromkeys(gezien)) or "geen"
        raise ValueError(
            "de kolomkoppen zijn niet herkend. Verwacht worden 'Medewerker', "
            "'Datum', 'Starttijd' en 'Eindtijd'. Gevonden koppen: " + gevonden
        )

    per_naam: dict[str, SnoopMedewerker] = {}
    schaal_stemmen: dict[str, Counter] = {}

    for rij in rijen:
        naam_ruw = rij[idx["medewerker"]] if idx["medewerker"] < len(rij) else None
        if not naam_ruw:
            continue
        naam = _norm_naam(naam_ruw)
        datum = _als_datum(rij[idx["datum"]])
        begin = _als_tijd(rij[idx["start"]])
        eind = _als_tijd(rij[idx["eind"]])
        if datum is None or begin is None or eind is None:
            continue

        minuten = _als_minuten(rij[idx["uren"]]) if "uren" in idx else None
        if minuten is None:  # val terug op begin/eind
            bruto = (eind.hour * 60 + eind.minute) - (begin.hour * 60 + begin.minute)
            if bruto <= 0:
                bruto += 24 * 60
            minuten = bruto

        mw = per_naam.setdefault(naam, SnoopMedewerker(naam=naam, loonschaal=None))
        mw.planning.append(PlanningRegel(datum, begin, eind, minuten))

        if "loonschaal" in idx and idx["loonschaal"] < len(rij):
            schaal = rij[idx["loonschaal"]]
            if schaal:
                schaal_stemmen.setdefault(naam, Counter())[str(schaal).strip()] += 1

        if "werkgever" in idx and idx["werkgever"] < len(rij):
            werkgever = rij[idx["werkgever"]]
            if werkgever and not mw.werkgever:
                mw.werkgever = str(werkgever).strip()

    for naam, teller in schaal_stemmen.items():
        per_naam[naam].loonschaal = teller.most_common(1)[0][0]

    return sorted(per_naam.values(), key=lambda m: m.naam)
