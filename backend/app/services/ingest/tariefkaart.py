"""Geconsolideerde UZB-tariefkaart (.xlsx) inlezen (SPEC §6).

Het bestand heeft één tabblad per uitzendbureau, elk met een eigen
kolomindeling. De configuratie hieronder beschrijft die verschillen; de parser
zelf is generiek.

Uit de kaart worden twee dingen gehaald:
1. de tarieven per uitzendbureau, waaruit de omrekenfactoren volgen
   (`factor = tarief / CAO-uurloon`);
2. de CAO-loontabel -- die staat als loontabel (letter x trede voor
   volwassenen, leeftijd x letter voor jeugd) in het jeugd-tabblad, en als
   directe `Loon`-kolom bij Level One.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.tarief.kaart import Loontabel
from app.services.tarief.types import (
    CAT_100,
    CAT_115,
    CAT_122,
    CAT_135,
    CAT_150,
    CAT_200,
    CAT_FEESTDAG,
    CAT_NACHT_50,
    CAT_NACHTUUR,
)


@dataclass(frozen=True)
class BladConfig:
    """Indeling van één tabblad in de tariefkaart."""

    uzb_sleutel: str
    sheet: str
    kolommen: dict[str, str]  # exacte kolomkop -> tariefcategorie
    loon_kolom: str | None = None
    # Contractueel één factor per categorie voor álle schalen van deze UZB.
    uniforme_factor: bool = False


BLADEN: tuple[BladConfig, ...] = (
    BladConfig(
        uzb_sleutel="L1",
        sheet="L1",
        kolommen={
            "100%/135%": CAT_100,
            "150%": CAT_150,
            "200%": CAT_200,
            "150% bijzonder ; 50% nachturen": CAT_FEESTDAG,
        },
        loon_kolom="Loon",
        uniforme_factor=False,  # verschilt per suffix V/F/S
    ),
    BladConfig(
        uzb_sleutel="L1_JEUGD",
        sheet="L1 jeugd-payroll",
        kolommen={
            "100%": CAT_100,
            "150%": CAT_150,
            "200%": CAT_200,
            "150% bijzonder ; 50% nachturen": CAT_FEESTDAG,
        },
        loon_kolom="Loon",
        uniforme_factor=False,
    ),
    BladConfig(
        uzb_sleutel="SW",
        sheet="Sterk Werk",
        kolommen={
            "100%/135%": CAT_100,
            "150%": CAT_150,
            "200%": CAT_200,
            "Feestdag": CAT_FEESTDAG,
            "50% nacht": CAT_NACHT_50,
            "Totaal nachtuur": CAT_NACHTUUR,
        },
        uniforme_factor=True,
    ),
    BladConfig(
        uzb_sleutel="CK",
        sheet="Cervokordaat",
        kolommen={
            "100%": CAT_100,
            "135%": CAT_135,
            "150%": CAT_150,
            "115%": CAT_115,
            "122%": CAT_122,
            "150%2": CAT_FEESTDAG,
            "200%": CAT_200,
        },
        uniforme_factor=True,
    ),
)

# Volwassen schaal: letter + trede (+ optioneel L1-suffix V/F/S), bv. "B2F".
_VOLWASSEN = re.compile(r"^([A-H])\s*(\d{1,2})\s*([VFS])?$", re.IGNORECASE)
# Jeugdschaal: leeftijd + letter + trede, bv. "15B2".
_JEUGD = re.compile(r"^(1[3-8])\s*([A-H])\s*(\d{1,2})?$", re.IGNORECASE)


def cao_schaal_code(kaartcode: str) -> str | None:
    """Zet een UZB-kaartcode om naar de CAO-schaal waarvan het loon geldt.

    "B2F"/"B2V"/"B2S" en "B2" verwijzen alle naar CAO-schaal "B2" -- Flex, Vast
    en Seizoens delen het CAO-loon maar niet het tarief. "15B2" verwijst naar de
    jeugdschaal "15B" (leeftijd + letter; de trede telt daar niet mee).
    """
    code = re.sub(r"\s+", "", str(kaartcode or "")).upper()
    if (m := _JEUGD.match(code)) :
        return f"{m.group(1)}{m.group(2)}"
    if (m := _VOLWASSEN.match(code)) :
        return f"{m.group(1)}{m.group(2)}"
    return None


def _bedrag(waarde) -> Decimal | None:
    if waarde is None or waarde == "":
        return None
    try:
        bedrag = Decimal(str(waarde).replace("€", "").replace(",", ".").strip())
    except InvalidOperation:
        return None
    return bedrag if bedrag > 0 else None


@dataclass
class KaartBlad:
    """De tarieven van één uitzendbureau zoals ingelezen."""

    uzb_sleutel: str
    uniforme_factor: bool
    tarieven: dict[str, dict[str, Decimal]] = field(default_factory=dict)
    lonen: dict[str, Decimal] = field(default_factory=dict)  # kaartcode -> loon


def _vind_header(ws, koppen: set[str]) -> tuple[int, dict[str, int]] | None:
    for i, rij in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        posities = {
            str(c).strip(): j for j, c in enumerate(rij) if c and str(c).strip() in koppen
        }
        if str(rij[0] or "").strip().lower() == "schaal" and posities:
            return i, posities
    return None


def lees_tariefkaart(
    bron: str | Path | bytes,
) -> tuple[dict[str, KaartBlad], list[str]]:
    """Lees alle UZB-tabbladen. Retourneert de bladen plus waarschuwingen."""
    data = BytesIO(bron) if isinstance(bron, (bytes, bytearray)) else bron
    wb = load_workbook(data, data_only=True)

    bladen: dict[str, KaartBlad] = {}
    waarschuwingen: list[str] = []

    for cfg in BLADEN:
        if cfg.sheet not in wb.sheetnames:
            waarschuwingen.append(f"tabblad '{cfg.sheet}' ontbreekt in de tariefkaart")
            continue
        ws = wb[cfg.sheet]
        koppen = set(cfg.kolommen) | ({cfg.loon_kolom} if cfg.loon_kolom else set())
        gevonden = _vind_header(ws, koppen)
        if gevonden is None:
            waarschuwingen.append(f"tabblad '{cfg.sheet}': kolomkoppen niet herkend")
            continue
        headerrij, posities = gevonden

        ontbrekend = set(cfg.kolommen) - set(posities)
        if ontbrekend:
            waarschuwingen.append(
                f"tabblad '{cfg.sheet}': kolom(men) {sorted(ontbrekend)} niet gevonden"
            )

        blad = KaartBlad(cfg.uzb_sleutel, cfg.uniforme_factor)
        for rij in ws.iter_rows(min_row=headerrij + 1, values_only=True):
            ruwe_code = rij[0] if rij else None
            if not ruwe_code or not str(ruwe_code).strip():
                continue
            kaartcode = re.sub(r"\s+", "", str(ruwe_code)).upper()
            if cao_schaal_code(kaartcode) is None:
                continue  # geen schaalregel (toelichting, subtotaal, ...)

            tarieven: dict[str, Decimal] = {}
            for kop, categorie in cfg.kolommen.items():
                j = posities.get(kop)
                if j is None or j >= len(rij):
                    continue
                if (bedrag := _bedrag(rij[j])) is not None:
                    tarieven[categorie] = bedrag
            if not tarieven:
                continue
            blad.tarieven[kaartcode] = tarieven

            if cfg.loon_kolom and (j := posities.get(cfg.loon_kolom)) is not None:
                if j < len(rij) and (loon := _bedrag(rij[j])) is not None:
                    blad.lonen[kaartcode] = loon

        bladen[cfg.uzb_sleutel] = blad

    return bladen, waarschuwingen


def lees_cao_loontabel(
    bron: str | Path | bytes, naam: str, ingangsdatum: date
) -> tuple[Loontabel, list[str]]:
    """Bouw de CAO-loontabel uit de tariefkaart.

    Bron 1: de loontabel rechts op het jeugd-tabblad (leeftijd x letter voor
    jeugd, trede x letter voor volwassenen). Bron 2: de `Loon`-kolom bij Level
    One, die per kaartcode het CAO-loon herhaalt -- gebruikt als aanvulling en
    als controle.
    """
    data = BytesIO(bron) if isinstance(bron, (bytes, bytearray)) else bron
    wb = load_workbook(data, data_only=True)
    lonen: dict[str, Decimal] = {}
    waarschuwingen: list[str] = []

    sheet = "L1 jeugd-payroll"
    if sheet in wb.sheetnames:
        ws = wb[sheet]
        letters: dict[int, str] = {}
        label_kolom: int | None = None
        for rij in ws.iter_rows(values_only=True):
            cellen = list(rij)
            for j, cel in enumerate(cellen):
                if cel and str(cel).strip() == "Trede/schaal":
                    label_kolom = j
                    letters = {
                        k: str(c).strip()
                        for k, c in enumerate(cellen)
                        if k > j and c and re.fullmatch(r"[A-H]", str(c).strip())
                    }
                    break
            if not letters or label_kolom is None:
                continue
            # het rijlabel staat in de kolom van 'Trede/schaal', niet links op
            # het blad -- daar staan de schaalcodes van de tarieventabel.
            if label_kolom >= len(cellen):
                continue
            label = str(cellen[label_kolom] or "").strip()
            if (m := re.fullmatch(r"(1[3-8])\s*jaar", label, re.IGNORECASE)):
                prefix, sjabloon = m.group(1), "{leeftijd}{letter}"
            elif re.fullmatch(r"\d{1,2}", label):
                prefix, sjabloon = label, "{letter}{trede}"
            else:
                continue
            for k, letter in letters.items():
                if k < len(cellen) and (bedrag := _bedrag(cellen[k])) is not None:
                    code = (
                        sjabloon.format(leeftijd=prefix, letter=letter)
                        if "leeftijd" in sjabloon
                        else sjabloon.format(letter=letter, trede=prefix)
                    )
                    lonen.setdefault(code, bedrag)

    # aanvullen/controleren met de Loon-kolom van Level One
    bladen, _ = lees_tariefkaart(bron)
    for sleutel in ("L1", "L1_JEUGD"):
        blad = bladen.get(sleutel)
        if not blad:
            continue
        for kaartcode, loon in blad.lonen.items():
            cao = cao_schaal_code(kaartcode)
            if cao is None:
                continue
            if cao in lonen and lonen[cao] != loon:
                waarschuwingen.append(
                    f"CAO-schaal {cao}: loontabel zegt {lonen[cao]}, "
                    f"Loon-kolom bij {kaartcode} zegt {loon}"
                )
            lonen.setdefault(cao, loon)

    if not lonen:
        raise ValueError("geen CAO-lonen gevonden in de tariefkaart")

    return Loontabel(naam=naam, ingangsdatum=ingangsdatum, lonen=lonen), waarschuwingen
