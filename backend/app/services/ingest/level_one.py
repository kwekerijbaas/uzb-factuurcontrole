"""Level One's eigen tariefexport (.xlsx) inlezen (SPEC §6).

Level One levert zijn tarieven in een ander formaat dan de geconsolideerde
tariefkaart: één blok per CAO-schaal, met daaronder een regel per
loonbestanddeel, en aparte kolommen voor de contractvormen Vast, Flex en
Seizoen. Elke kolom komt twee keer voor -- de oude en de nieuwe waarde -- zodat
in één bestand te zien is wat er verandert.

Voorbeeld van de indeling:

    Code    | Loon oud | Loon per 1/7/26 | Component            | Percentage | Vast oud | Vast per 1/7/26 | Flex oud | ...
    GTB B2  |    14,71 |           14,99 |                      |            |          |                 |          |
            |          |                 | Loon normale uren    |        100 |    28,03 |           28,51 |    28,94 | ...
            |          |                 | Loon overwerkuren    |        135 |    28,03 |           28,51 |    28,94 | ...

De contractvorm bepaalt het achtervoegsel van de kaartcode (`B2V`, `B2F`,
`B2S`), net als in de geconsolideerde kaart.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.tarief.types import (
    CAT_100,
    CAT_115,
    CAT_122,
    CAT_135,
    CAT_150,
    CAT_200,
    CAT_FEESTDAG,
)

# (component, percentage) -> tariefcategorie. Het percentage alleen volstaat
# niet: 150 komt twee keer voor, als overwerk en als bijzondere uren (feestdag
# en nachturen), met verschillende tarieven.
_COMPONENTEN: dict[tuple[str, str], str] = {
    ("normale uren", "100"): CAT_100,
    ("overwerkuren", "135"): CAT_135,
    ("overwerkuren", "150"): CAT_150,
    ("overwerkuren", "200"): CAT_200,
    ("onregelmatige uren", "115"): CAT_115,
    ("onregelmatige uren", "122"): CAT_122,
    ("bijzondere uren", "150"): CAT_FEESTDAG,
}

# contractvorm -> achtervoegsel van de kaartcode
_VORMEN = {"vast": "V", "flex": "F", "seizoen": "S"}

_KOLOM = re.compile(r"^(vast|flex|seizoen|loon)\s+(oud|per\b.*)$", re.IGNORECASE)

# De ingangsdatum staat in de kolomkop zelf, zodat die niet overgetypt hoeft te
# worden. Level One schrijft hem op twee manieren: "per 1/7/26" en "per 1 jul".
_MAANDEN = {
    "jan": 1, "feb": 2, "mrt": 3, "maa": 3, "apr": 4, "mei": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "okt": 10, "nov": 11, "dec": 12,
}
_CIJFERDATUM = re.compile(r"per\s+(\d{1,2})\s*[-/]\s*(\d{1,2})\s*[-/]\s*(\d{2,4})")
_NAAMDATUM = re.compile(
    r"per\s+(\d{1,2})\s+([a-z]{3,})\.?\s*(\d{2,4})?", re.IGNORECASE
)


def _samen(dag: int, maand: int, jaar: int | None, vandaag: date) -> date | None:
    """Maak er een datum van; zonder jaartal het dichtstbijzijnde jaar.

    "per 1 jul" noemt geen jaar. Van vorig, dit en volgend jaar wordt dat
    gekozen waar die dag het dichtst bij vandaag ligt -- zoals een mens hem ook
    leest. Het scherm toont de uitkomst, zodat een verkeerde gok opvalt vóór er
    weken mee worden verwerkt.
    """
    if jaar is not None:
        jaar = jaar + 2000 if jaar < 100 else jaar
        kandidaten = [jaar]
    else:
        kandidaten = [vandaag.year - 1, vandaag.year, vandaag.year + 1]
    mogelijk = []
    for kandidaat in kandidaten:
        try:
            mogelijk.append(date(kandidaat, maand, dag))
        except ValueError:
            continue
    return min(mogelijk, key=lambda d: abs((d - vandaag).days)) if mogelijk else None


def peildatum(tekst: str | None, vandaag: date | None = None) -> date | None:
    """De ingangsdatum uit een kolomkop als 'Loon per 1/7/26' of 'Loon per 1 jul'."""
    tekst = str(tekst or "")
    vandaag = vandaag or date.today()
    if (m := _CIJFERDATUM.search(tekst)) is not None:
        dag, maand, jaar = (int(g) for g in m.groups())
        return _samen(dag, maand, jaar, vandaag)
    if (m := _NAAMDATUM.search(tekst)) is not None:
        maand = _MAANDEN.get(m.group(2)[:3].lower())
        if maand is not None:
            jaar = int(m.group(3)) if m.group(3) else None
            return _samen(int(m.group(1)), maand, jaar, vandaag)
    return None


@dataclass
class LevelOneExport:
    """De tarieven uit één Level One-export, voor één peilmoment."""

    tarieven: dict[str, dict[str, Decimal]] = field(default_factory=dict)
    lonen: dict[str, Decimal] = field(default_factory=dict)  # CAO-schaal -> uurloon
    # Uit de kolomkop ("Loon per 1/7/26"); leeg bij de oude kolom, want die
    # noemt zijn eigen ingangsdatum niet.
    ingangsdatum: date | None = None


def _cel(rij, kolom: int | None):
    return rij[kolom] if kolom is not None and kolom < len(rij) else None


def _bedrag(waarde) -> Decimal | None:
    if waarde is None or waarde == "":
        return None
    try:
        bedrag = Decimal(str(waarde).replace("€", "").replace(",", ".").strip())
    except InvalidOperation:
        return None
    return bedrag if bedrag > 0 else None


def _schaal_uit_code(waarde) -> str | None:
    """"GTB B2" -> "B2". Alleen regels met een schaalcode tellen als kopregel."""
    tekst = re.sub(r"\s+", " ", str(waarde or "")).strip()
    m = re.match(r"^(?:GTB\s+)?([A-H]\s*\d{1,2})$", tekst, re.IGNORECASE)
    return re.sub(r"\s+", "", m.group(1)).upper() if m else None


def _categorie(component, percentage) -> str | None:
    tekst = str(component or "").strip().lower()
    pct = re.sub(r"[^\d]", "", str(percentage or ""))
    for (kern, verwacht_pct), categorie in _COMPONENTEN.items():
        if kern in tekst and pct == verwacht_pct:
            return categorie
    return None


def _kolommen(header) -> tuple[dict[str, dict[str, int]], list[str], dict[str, int]]:
    """Vind per contractvorm (en voor het loon) de kolom 'oud' en 'nieuw'.

    Retourneert ook de gevonden peildatum-omschrijvingen -- zodat de gebruiker
    kan controleren welke kolom is gelezen -- en de plaats van 'Code',
    'Component' en 'Percentage'. Die laatste worden opgezocht in plaats van
    vastgelegd: Level One levert de export niet altijd met evenveel lege
    tussenkolommen, en op een vaste positie rekenen levert dan een bestand op
    waarin geen enkele tariefregel wordt gevonden.
    """
    posities: dict[str, dict[str, int]] = {}
    peilmomenten: list[str] = []
    vaste: dict[str, int] = {}
    for i, cel in enumerate(header):
        tekst = re.sub(r"\s+", " ", str(cel or "")).strip()
        kern = tekst.lower().rstrip(".")
        if kern in ("code", "component", "percentage") and kern not in vaste:
            vaste[kern] = i
        m = _KOLOM.match(tekst)
        if not m:
            continue
        soort = m.group(1).lower()
        welke = "oud" if m.group(2).lower() == "oud" else "nieuw"
        posities.setdefault(soort, {})[welke] = i
        if welke == "nieuw" and m.group(2) not in peilmomenten:
            peilmomenten.append(m.group(2))
    return posities, peilmomenten, vaste


def lees_level_one_export(
    bron: str | Path | bytes, welke: str = "nieuw"
) -> tuple[LevelOneExport, list[str]]:
    """Lees een Level One-tariefexport.

    `welke` kiest tussen de kolommen "nieuw" (de aangekondigde tarieven) en
    "oud" (de tarieven die tot de ingangsdatum golden). De oude kolom is handig
    om te controleren of het bestand aansluit op wat er nu in de app staat.
    """
    if welke not in ("oud", "nieuw"):
        raise ValueError("welke moet 'oud' of 'nieuw' zijn")

    data = BytesIO(bron) if isinstance(bron, (bytes, bytearray)) else bron
    wb = load_workbook(data, data_only=True, read_only=True)

    export = LevelOneExport()
    waarschuwingen: list[str] = []
    gevonden_kolommen = False

    for ws in wb.worksheets:
        posities: dict[str, dict[str, int]] = {}
        vaste: dict[str, int] = {}
        schaal: str | None = None

        for rij in ws.iter_rows(values_only=True):
            if not posities:
                posities, peilmomenten, vaste = _kolommen(rij)
                if posities:
                    gevonden_kolommen = True
                    ontbreekt = set(_VORMEN) - set(posities)
                    if ontbreekt:
                        waarschuwingen.append(
                            f"contractvorm(en) {sorted(ontbreekt)} niet in het bestand"
                        )
                    if welke == "nieuw" and peilmomenten:
                        export.ingangsdatum = peildatum(peilmomenten[0])
                        gevonden = (
                            f" (ingangsdatum {export.ingangsdatum:%d-%m-%Y})"
                            if export.ingangsdatum
                            else " -- ingangsdatum niet uit de kop af te leiden"
                        )
                        waarschuwingen.append(
                            "nieuwe tarieven gelezen uit de kolommen "
                            f"'{peilmomenten[0]}'{gevonden}"
                        )
                continue

            if (code := _schaal_uit_code(_cel(rij, vaste.get("code", 1)))) is not None:
                schaal = code
                if "loon" in posities and (kolom := posities["loon"].get(welke)) is not None:
                    if kolom < len(rij) and (loon := _bedrag(rij[kolom])) is not None:
                        export.lonen[schaal] = loon
                continue

            if schaal is None:
                continue
            categorie = _categorie(
                _cel(rij, vaste.get("component")), _cel(rij, vaste.get("percentage"))
            )
            if categorie is None:
                continue

            for vorm, achtervoegsel in _VORMEN.items():
                kolom = posities.get(vorm, {}).get(welke)
                if kolom is None or kolom >= len(rij):
                    continue
                if (tarief := _bedrag(rij[kolom])) is not None:
                    export.tarieven.setdefault(f"{schaal}{achtervoegsel}", {})[
                        categorie
                    ] = tarief

    if not gevonden_kolommen:
        raise ValueError(
            "kolomkoppen niet herkend; verwacht kolommen als 'Loon oud', "
            "'Flex oud' en 'Flex per <datum>'"
        )
    if not export.tarieven:
        raise ValueError("geen tariefregels gevonden in de Level One-export")

    return export, waarschuwingen
