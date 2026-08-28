"""CAO-loontabel (.xlsx) inlezen (SPEC §6).

De gebruiker uploadt een loontabel met een ingangsdatum; vanaf die datum worden
de UZB-tarieven tegen deze lonen berekend. Het bestand hoeft alleen een kolom
met de schaalcode en een kolom met het uurloon te bevatten; de headers worden
soepel herkend zodat een export van de CAO-partij niet handmatig hoeft te
worden omgebouwd.
"""

from __future__ import annotations

import re
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.tarief.kaart import Loontabel

_SCHAAL_KOPPEN = ("schaal", "schaalcode", "loonschaal", "code", "functiegroep")
_LOON_KOPPEN = ("uurloon", "loon", "bedrag", "uurbedrag", "salaris")
_OMSCHRIJVING_KOPPEN = ("omschrijving", "functie", "toelichting")


def _norm(waarde) -> str:
    return re.sub(r"[^a-z0-9]", "", str(waarde or "").lower())


def _als_bedrag(waarde) -> Decimal | None:
    if waarde is None or waarde == "":
        return None
    if isinstance(waarde, (int, float, Decimal)):
        bedrag = Decimal(str(waarde))
    else:
        tekst = str(waarde).strip().replace("€", "").strip()
        # NL-notatie: 1.234,56 -> 1234.56
        if "," in tekst:
            tekst = tekst.replace(".", "").replace(",", ".")
        try:
            bedrag = Decimal(tekst)
        except InvalidOperation:
            return None
    return bedrag if bedrag > 0 else None


def _vind_kolommen(rij) -> dict[str, int] | None:
    """Herken de header-rij; None als deze rij geen header is."""
    gevonden: dict[str, int] = {}
    for i, cel in enumerate(rij):
        sleutel = _norm(cel)
        if not sleutel:
            continue
        if "schaal" not in gevonden and sleutel in _SCHAAL_KOPPEN:
            gevonden["schaal"] = i
        elif "loon" not in gevonden and sleutel in _LOON_KOPPEN:
            gevonden["loon"] = i
        elif "omschrijving" not in gevonden and sleutel in _OMSCHRIJVING_KOPPEN:
            gevonden["omschrijving"] = i
    return gevonden if {"schaal", "loon"} <= gevonden.keys() else None


def lees_loontabel(
    bron: str | Path | bytes, naam: str, ingangsdatum: date
) -> tuple[Loontabel, list[str]]:
    """Parse een CAO-loontabel. Retourneert de tabel plus waarschuwingen."""
    data = BytesIO(bron) if isinstance(bron, (bytes, bytearray)) else bron
    wb = load_workbook(data, data_only=True, read_only=True)

    waarschuwingen: list[str] = []
    lonen: dict[str, Decimal] = {}

    for ws in wb.worksheets:
        kolommen: dict[str, int] | None = None
        for rij in ws.iter_rows(values_only=True):
            if kolommen is None:
                kolommen = _vind_kolommen(rij)
                continue
            schaal = rij[kolommen["schaal"]] if kolommen["schaal"] < len(rij) else None
            loon = _als_bedrag(rij[kolommen["loon"]] if kolommen["loon"] < len(rij) else None)
            if not schaal or loon is None:
                continue
            code = re.sub(r"\s+", " ", str(schaal)).strip()
            if code in lonen and lonen[code] != loon:
                waarschuwingen.append(
                    f"schaal {code} komt meerdere keren voor met verschillende lonen "
                    f"({lonen[code]} en {loon}); de eerste is aangehouden"
                )
                continue
            lonen.setdefault(code, loon)

    if not lonen:
        raise ValueError(
            "geen loonregels gevonden; verwacht een kolom met de schaalcode "
            "(bv. 'Schaal') en een kolom met het uurloon (bv. 'Uurloon')"
        )

    return Loontabel(naam=naam, ingangsdatum=ingangsdatum, lonen=lonen), waarschuwingen
