"""Uitzendkrachtenlijst inlezen: wie werkt er, en in welke loonschaal.

De lijst is een SNOOP-export over een langere periode. Daaruit wordt per
uitzendkracht de **meest recente** loonschaal gehaald: een schaal kan in de loop
van het jaar wijzigen (in de lijst over 2026 gebeurde dat bij tien
Sterk Werk-krachten en drie van Level One), en dan is de laatste de geldende.

Deze lijst dient als terugval voor weken waarin SNOOP iemand niet bevat. De
SNOOP-regel van de week zelf gaat altijd voor.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

_KOLOMMEN = {
    "medewerker": "naam",
    "registratienummer": "code",
    "datum": "datum",
    "tariefuitzendbureau": "loonschaal",
    "typeuitzendkracht": "type",
    "werkgeveropdatumshift": "werkgever",
}


@dataclass
class UzkRegel:
    """Eén uitzendkracht uit de lijst."""

    naam: str
    externe_code: str | None = None
    loonschaal: str | None = None
    soort: str | None = None
    werkgever: str | None = None
    laatste_datum: date | None = None
    eerdere_schalen: list[str] = field(default_factory=list)

    @property
    def is_gewisseld(self) -> bool:
        return bool(self.eerdere_schalen)


def _datum(waarde) -> date | None:
    if isinstance(waarde, datetime):
        return waarde.date()
    if isinstance(waarde, date):
        return waarde
    if not waarde:
        return None
    for opmaak in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(waarde).strip(), opmaak).date()
        except ValueError:
            continue
    return None


def lees_uzk_lijst(bron: str | Path | bytes) -> tuple[list[UzkRegel], list[str]]:
    """Parse een uitzendkrachtenlijst. Retourneert één regel per medewerker."""
    data = BytesIO(bron) if isinstance(bron, (bytes, bytearray)) else bron
    wb = load_workbook(data, data_only=True, read_only=True)
    ws = wb.active

    rijen = ws.iter_rows(values_only=True)
    header = next(rijen, None)
    if header is None:
        raise ValueError("het bestand is leeg")
    idx: dict[str, int] = {}
    for i, cel in enumerate(header):
        sleutel = re.sub(r"[^a-z]", "", str(cel or "").lower())
        if sleutel in _KOLOMMEN:
            idx[_KOLOMMEN[sleutel]] = i
    if "naam" not in idx:
        raise ValueError(
            "kolom 'Medewerker' niet gevonden; verwacht een SNOOP-export"
        )

    per_naam: dict[str, UzkRegel] = {}
    schaal_op_datum: dict[str, list[tuple[date, str]]] = {}

    def haal(rij, veld):
        i = idx.get(veld)
        return rij[i] if i is not None and i < len(rij) else None

    for rij in rijen:
        naam = re.sub(r"\s+", " ", str(haal(rij, "naam") or "")).strip()
        if not naam:
            continue
        regel = per_naam.setdefault(naam, UzkRegel(naam=naam))
        if (code := haal(rij, "code")) and str(code).strip() not in ("", "0"):
            regel.externe_code = str(code).strip()
        if (soort := haal(rij, "type")):
            regel.soort = str(soort).strip()
        if (werkgever := haal(rij, "werkgever")):
            regel.werkgever = str(werkgever).strip()

        schaal = haal(rij, "loonschaal")
        if schaal and str(schaal).strip():
            dag = _datum(haal(rij, "datum")) or date.min
            schaal_op_datum.setdefault(naam, []).append((dag, str(schaal).strip()))

    waarschuwingen: list[str] = []
    for naam, paren in schaal_op_datum.items():
        paren.sort(key=lambda p: p[0])
        regel = per_naam[naam]
        regel.loonschaal = paren[-1][1]
        regel.laatste_datum = paren[-1][0] if paren[-1][0] != date.min else None
        overige = sorted({s for _, s in paren} - {regel.loonschaal})
        if overige:
            regel.eerdere_schalen = overige
            waarschuwingen.append(
                f"{naam}: loonschaal gewisseld ({', '.join(overige)} -> "
                f"{regel.loonschaal}); de laatste is aangehouden"
            )

    zonder = [r.naam for r in per_naam.values() if not r.loonschaal]
    if zonder:
        waarschuwingen.append(
            f"{len(zonder)} medewerker(s) zonder loonschaal: {', '.join(sorted(zonder)[:5])}"
            + (" ..." if len(zonder) > 5 else "")
        )

    if not per_naam:
        raise ValueError("geen medewerkers gevonden in de lijst")

    return sorted(per_naam.values(), key=lambda r: r.naam), waarschuwingen
