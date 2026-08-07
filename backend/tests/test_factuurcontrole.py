"""Tests voor de factuurcontrole (SPEC §7).

Gevalideerd tegen week 25/2026 Level One: 21 van de 21 medewerkers gekoppeld
aan 21 factuurblokken uit vier facturen, met een verschil van -10,00 uur en
-EUR 414,04 -- exact de uitkomst van de handmatige controle destijds.
"""

from decimal import Decimal

import pytest

from app.services.calc.types import WeekResultaat
from app.services.factuurcontrole import (
    SOORT_BEDRAG,
    SOORT_NIET_GEFACTUREERD,
    SOORT_NIET_IN_OVERZICHT,
    SOORT_UREN,
    bevindingenmail,
    controleer,
    koppel,
)
from app.services.ingest.factuur import Factuur, FactuurKracht, FactuurRegel
from app.services.tarief.types import BedragRegel, BedragResultaat
from app.services.verwerking import MedewerkerResultaat, WeekVerwerking


def _medewerker(naam: str, uren: str, bedrag: str) -> MedewerkerResultaat:
    minuten = int(Decimal(uren) * 60)
    return MedewerkerResultaat(
        naam=naam, nitea_id="1", loonschaal="B2 Flex", kaartcode="B2F",
        resultaat=WeekResultaat(netto_minuten=minuten, minuten_per_percentage={}),
        bedrag=BedragResultaat(
            regels=[BedragRegel("100", (), minuten, Decimal("1"), Decimal(bedrag))]
        ),
    )


def _kracht(naam: str, uren: str, bedrag: str) -> FactuurKracht:
    return FactuurKracht(
        naam_ruw=naam,
        regels=[FactuurRegel("Loon normale uren", None, Decimal(uren),
                             Decimal("28.94"), Decimal(bedrag))],
    )


def _week(medewerkers) -> WeekVerwerking:
    verwerking = WeekVerwerking("L1", 2026, 25)
    verwerking.medewerkers = list(medewerkers)
    return verwerking


def _factuur(krachten) -> Factuur:
    return Factuur(uzb_sleutel="L1", factuurnummers=["02604736"], krachten=list(krachten))


# --------------------------------------------------------------------------- #
# Koppelen van namen
# --------------------------------------------------------------------------- #
def test_initialen_op_de_factuur_koppelen_aan_volledige_namen():
    """De factuur zet "K.P. Sliwa (Kamil)", wij kennen "Kamil Sliwa"."""
    gekoppeld, zonder_factuur, zonder_overzicht = koppel(
        [_medewerker("Kamil Sliwa", "45", "1512.41")],
        [_kracht("K.P. Sliwa (Kamil)", "45", "1501.65")],
    )
    assert len(gekoppeld) == 1
    assert not zonder_factuur and not zonder_overzicht


def test_naamgenoten_worden_uit_elkaar_gehouden():
    """Twee keer Pilarz: alleen de voornaam onderscheidt ze."""
    gekoppeld, _, _ = koppel(
        [_medewerker("Malgorzata Pilarz", "47.50", "1582.21"),
         _medewerker("Kamila Pilarz", "29.75", "880.30")],
        [_kracht("M. Pilarz (Malgorzata)", "47.50", "1580.79"),
         _kracht("K.G. Pilarz (Kamila)", "29.75", "880.30")],
    )
    koppels = {m.naam: k.naam_ruw for m, k in gekoppeld}
    assert "Malgorzata" in koppels["Malgorzata Pilarz"]
    assert "Kamila" in koppels["Kamila Pilarz"]


def test_meerdere_blokken_van_een_kracht_worden_opgeteld():
    """Een nagekomen dag komt als apart blok op de factuur."""
    controle = controleer(
        _week([_medewerker("Julia Machura", "41", "1192.77")]),
        _factuur([_kracht("J.J. Machura (Julia)", "31.25", "904.38"),
                  _kracht("J.J. Machura (Julia)", "9.75", "288.39")]),
        "Level One",
    )
    assert controle.uren_factuur == Decimal("41.00")
    assert controle.bevindingen == []


# --------------------------------------------------------------------------- #
# Classificatie van verschillen
# --------------------------------------------------------------------------- #
def test_urenverschil_wordt_gemeld():
    controle = controleer(
        _week([_medewerker("Bartlomiej Janicki", "33.75", "1107.22")]),
        _factuur([_kracht("B.D. Janicki (Bartlomiej)", "27.75", "892.16")]),
        "Level One",
    )
    assert [b.soort for b in controle.bevindingen] == [SOORT_UREN]
    assert controle.bevindingen[0].uren_verschil == Decimal("-6.00")


def test_gelijke_uren_maar_afwijkend_bedrag_wijst_op_de_loonschaal():
    controle = controleer(
        _week([_medewerker("Marius Girtoi", "30.25", "875.44")]),
        _factuur([_kracht("M.P. Girtoi (Marius)", "30.25", "895.70")]),
        "Level One",
    )
    assert [b.soort for b in controle.bevindingen] == [SOORT_BEDRAG]
    assert "loonschaal" in controle.bevindingen[0].melding


def test_afronding_van_het_uurtarief_is_geen_bevinding():
    """Level One draagt drie decimalen; enkele centen verschil is ruis."""
    controle = controleer(
        _week([_medewerker("Alexandra Rebega", "38", "1099.68")]),
        _factuur([_kracht("A.R. Rebega (Alexandra)", "38", "1099.72")]),
        "Level One",
    )
    assert controle.bevindingen == []


def test_niet_gefactureerde_en_onbekende_krachten():
    controle = controleer(
        _week([_medewerker("Wel Gewerkt", "20", "578.80")]),
        _factuur([_kracht("O. Onbekend (Otto)", "10", "289.40")]),
        "Level One",
    )
    soorten = {b.soort for b in controle.bevindingen}
    assert soorten == {SOORT_NIET_GEFACTUREERD, SOORT_NIET_IN_OVERZICHT}
    assert controle.uren_overzicht == Decimal("20")
    assert controle.uren_factuur == Decimal("10")


# --------------------------------------------------------------------------- #
# Bevindingenmail
# --------------------------------------------------------------------------- #
def test_bevindingenmail_noemt_bedragen_en_bevindingen():
    controle = controleer(
        _week([_medewerker("Bartlomiej Janicki", "33.75", "1107.22")]),
        _factuur([_kracht("B.D. Janicki (Bartlomiej)", "27.75", "892.16")]),
        "Level One",
    )
    tekst = bevindingenmail([controle])
    assert "Level One — week 25/2026" in tekst
    assert "02604736" in tekst
    assert "Janicki" in tekst
    assert "minder" in tekst  # richting van het verschil


def test_bevindingenmail_bij_een_schone_controle():
    controle = controleer(
        _week([_medewerker("Alexandra Rebega", "38", "1099.72")]),
        _factuur([_kracht("A.R. Rebega (Alexandra)", "38", "1099.72")]),
        "Level One",
    )
    assert "Geen afwijkingen." in bevindingenmail([controle])


# --------------------------------------------------------------------------- #
# Matchingsbestand
# --------------------------------------------------------------------------- #
def test_matchingsbestand_bevat_de_vier_tabbladen():
    import io

    import openpyxl

    from app.services.export import bouw_matchingsbestand

    controle = controleer(
        _week([_medewerker("Bartlomiej Janicki", "33.75", "1107.22")]),
        _factuur([_kracht("B.D. Janicki (Bartlomiej)", "27.75", "892.16")]),
        "Level One",
    )
    wb = openpyxl.load_workbook(
        io.BytesIO(bouw_matchingsbestand(controle, bevindingenmail([controle])))
    )
    assert wb.sheetnames == [
        "Samenvatting", "Bevindingen", "Koppelingen", "Bevindingenmail"
    ]
    assert wb.active.title == "Samenvatting"

    koppelingen = [r for r in wb["Koppelingen"].iter_rows(min_row=4, values_only=True) if r[0]]
    assert koppelingen[0][1] == "B.D. Janicki (Bartlomiej)"  # naam zoals op de factuur
