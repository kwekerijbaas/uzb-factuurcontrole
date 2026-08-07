"""Tests voor de toegangscontrole, verwerking en export.

De databaseloze onderdelen worden hier gedekt; de volledige flow (upload ->
overzicht) is end-to-end tegen Postgres gedraaid.
"""

from __future__ import annotations

import io
from datetime import date, time
from decimal import Decimal

import openpyxl
import pytest
from fastapi import HTTPException, Response
from starlette.datastructures import Headers
from starlette.requests import Request

from app.auth import (
    SESSIE_COOKIE,
    Gebruiker,
    gebruiker_uit_cookie,
    huidige_gebruiker,
    is_toegestaan,
    zet_sessie,
)
from app.config import settings
from app.services.calc.types import PlanningRegel, RegistratieRegel
from app.services.export import bestandsnaam, bouw_overzicht
from app.services.ingest import NiteaMedewerker, SnoopMedewerker
from app.services.seed.cao_glastuinbouw import cao_toeslag_regels
from app.services.tarief import LEVEL_ONE, CAT_100, CAT_150, SchaalTarief, TariefKaart
from app.services.verwerking import normaliseer_naam, verwerk_week

MA = date(2025, 9, 1)


def _request(headers: dict[str, str] | None = None) -> Request:
    rauw = [(k.lower().encode(), v.encode()) for k, v in (headers or {}).items()]
    return Request(
        {"type": "http", "headers": Headers(raw=rauw).raw, "method": "GET", "path": "/"}
    )


def _client(monkeypatch, ingelogd: bool = False):
    from fastapi.testclient import TestClient

    from app.auth import Gebruiker, zet_sessie
    from app.main import app

    monkeypatch.setattr(settings, "auth_vereist", True)
    monkeypatch.setattr(settings, "cookie_secure", False)
    client = TestClient(app, follow_redirects=False)
    if ingelogd:
        antwoord = Response()
        zet_sessie(antwoord, Gebruiker(email="ola@kwekerijbaas.nl", id="abc"))
        client.cookies.set(SESSIE_COOKIE, antwoord.headers["set-cookie"].split("=")[1].split(";")[0])
    return client


# --------------------------------------------------------------------------- #
# Toegang
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize(
    "email,verwacht",
    [
        ("ola@kwekerijbaas.nl", True),
        ("OLA@Kwekerijbaas.NL", True),  # hoofdletterongevoelig
        ("iemand@gmail.com", False),
        ("kwekerijbaas.nl", False),  # geen adres
        ("aanvaller@kwaad.nl@gmail.com", False),
        ("", False),
    ],
)
def test_alleen_eigen_domein_mag_inloggen(email, verwacht):
    assert is_toegestaan(email) is verwacht


def test_los_toegelaten_adres(monkeypatch):
    monkeypatch.setattr(settings, "toegestane_emails", ["extern@boekhouder.nl"])
    assert is_toegestaan("extern@boekhouder.nl")
    assert not is_toegestaan("ander@boekhouder.nl")


def test_sessiecookie_rondrit():
    antwoord = Response()
    zet_sessie(antwoord, Gebruiker(email="ola@kwekerijbaas.nl", id="abc"))
    koekje = antwoord.headers["set-cookie"]
    waarde = koekje.split("=", 1)[1].split(";")[0]

    verzoek = _request({"cookie": f"{SESSIE_COOKIE}={waarde}"})
    gebruiker = gebruiker_uit_cookie(verzoek)
    assert gebruiker is not None and gebruiker.email == "ola@kwekerijbaas.nl"


def test_geknoeid_cookie_wordt_geweigerd():
    assert gebruiker_uit_cookie(_request({"cookie": f"{SESSIE_COOKIE}=verzonnen"})) is None


def test_ingetrokken_toegang_verloopt_de_sessie(monkeypatch):
    """Wie uit het toegestane domein valt, komt er met een oud cookie niet in."""
    antwoord = Response()
    zet_sessie(antwoord, Gebruiker(email="ola@kwekerijbaas.nl", id="abc"))
    waarde = antwoord.headers["set-cookie"].split("=", 1)[1].split(";")[0]

    monkeypatch.setattr(settings, "toegestane_domeinen", ["andersbedrijf.nl"])
    assert gebruiker_uit_cookie(_request({"cookie": f"{SESSIE_COOKIE}={waarde}"})) is None


def test_zonder_login_geen_toegang(monkeypatch):
    monkeypatch.setattr(settings, "auth_vereist", True)
    with pytest.raises(HTTPException) as fout:
        huidige_gebruiker(_request())
    assert fout.value.status_code == 401


def test_lokaal_zonder_login_toegestaan(monkeypatch):
    monkeypatch.setattr(settings, "auth_vereist", False)
    assert huidige_gebruiker(_request()).is_ontwikkelaar


def test_pagina_s_zijn_afgeschermd(monkeypatch):
    """De middleware sluit de hele app af, niet route voor route."""
    client = _client(monkeypatch)
    for pad in ("/", "/week", "/tarieven"):
        antwoord = client.get(pad)
        assert antwoord.status_code == 303, pad
        assert antwoord.headers["location"] == "/inloggen"


def test_gezondheid_en_inlogscherm_blijven_open(monkeypatch):
    client = _client(monkeypatch)
    assert client.get("/gezondheid").status_code == 200
    assert client.get("/inloggen").status_code == 200


# --------------------------------------------------------------------------- #
# Verwerking
# --------------------------------------------------------------------------- #
KAART = TariefKaart(
    "L1",
    date(2026, 1, 1),
    None,
    {"B2F": SchaalTarief("B2F", {CAT_100: Decimal("28.94"), CAT_150: Decimal("33.92")})},
)


def _nitea(naam: str, uren: int = 8) -> NiteaMedewerker:
    return NiteaMedewerker(
        naam=naam,
        nitea_id="1",
        registratie=[RegistratieRegel(MA, time(7, 0), time(15, 0), uren * 60, 0)],
    )


def _snoop(naam: str, schaal: str | None = "B2 Flex") -> SnoopMedewerker:
    return SnoopMedewerker(
        naam=naam,
        loonschaal=schaal,
        planning=[PlanningRegel(MA, time(7, 0), time(15, 0), 480)],
    )


@pytest.mark.parametrize(
    "links,rechts",
    [("Marius  Mic", "Marius Mic"), ("ALEXANDRA REBEGA", "Alexandra Rebega")],
)
def test_namen_uit_snoop_en_nitea_matchen(links, rechts):
    """SNOOP en Nitea verschillen in dubbele spaties en hoofdletters."""
    assert normaliseer_naam(links) == normaliseer_naam(rechts)


def test_verwerking_koppelt_uren_en_bedrag():
    verwerking = verwerk_week(
        "L1", 2026, 25, [_snoop("Marius  Mic")], [_nitea("Marius Mic")],
        cao_toeslag_regels(), KAART, LEVEL_ONE,
    )
    assert len(verwerking.medewerkers) == 1
    medewerker = verwerking.medewerkers[0]
    assert medewerker.kaartcode == "B2F"
    assert medewerker.netto_uren == Decimal("8.00")
    assert medewerker.bedrag.totaal == Decimal("231.52")  # 8 x 28,94
    assert verwerking.meldingen == []


def test_registratie_zonder_snoop_telt_gewoon_mee():
    """Nitea bepaalt wie er gewerkt heeft; ontbreken in SNOOP is op zichzelf
    geen probleem, alleen het ontbreken van een loonschaal is dat."""
    verwerking = verwerk_week(
        "L1", 2026, 25, [], [_nitea("Onbekende Kracht")],
        cao_toeslag_regels(), KAART, LEVEL_ONE,
    )
    assert verwerking.medewerkers[0].netto_uren == Decimal("8.00")
    assert any("geen loonschaal bekend" in m for m in verwerking.meldingen)


def test_planning_zonder_registratie_levert_geen_regel_op():
    """Wie gepland stond maar niet werkte, hoort niet in het overzicht."""
    verwerking = verwerk_week(
        "L1", 2026, 25, [_snoop("Niet Verschenen")], [],
        cao_toeslag_regels(), KAART, LEVEL_ONE,
    )
    assert verwerking.medewerkers == []
    assert verwerking.meldingen == []


def test_onbekende_loonschaal_levert_melding_en_geen_bedrag():
    verwerking = verwerk_week(
        "L1", 2026, 25, [_snoop("Marius Mic", "Z9 Flex")], [_nitea("Marius Mic")],
        cao_toeslag_regels(), KAART, LEVEL_ONE,
    )
    assert verwerking.medewerkers[0].bedrag.totaal == Decimal("0")
    assert any("geen tarief" in m for m in verwerking.meldingen)


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
def test_overzicht_opent_op_totaal_week_en_bevat_de_totalen():
    verwerking = verwerk_week(
        "L1", 2026, 25, [_snoop("Marius Mic")], [_nitea("Marius Mic")],
        cao_toeslag_regels(), KAART, LEVEL_ONE,
    )
    wb = openpyxl.load_workbook(io.BytesIO(bouw_overzicht(verwerking, "Level One", KAART)))

    assert wb.sheetnames == ["Totaal week", "Per dag", "Tarieven", "Afwijkingen"]
    assert wb.active.title == "Totaal week"

    ws = wb["Totaal week"]
    rijen = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0]]
    assert rijen[0][0] == "Marius Mic"
    assert rijen[-1][0] == "TOTAAL"
    assert rijen[-1][-1] == pytest.approx(231.52)


def test_meldingen_belanden_op_het_afwijkingen_tabblad():
    verwerking = verwerk_week(
        "L1", 2026, 25, [], [_nitea("Onbekende Kracht")],
        cao_toeslag_regels(), KAART, LEVEL_ONE,
    )
    wb = openpyxl.load_workbook(io.BytesIO(bouw_overzicht(verwerking, "Level One", KAART)))
    tekst = "\n".join(
        str(c.value) for rij in wb["Afwijkingen"].iter_rows() for c in rij if c.value
    )
    assert "geen loonschaal bekend" in tekst


def test_bestandsnaam_is_veilig():
    verwerking = verwerk_week("L1", 2026, 25, [], [], cao_toeslag_regels(), None, LEVEL_ONE)
    assert bestandsnaam("Level One", verwerking) == "UZB-overzicht_Level_One_week_25_2026.xlsx"


# --------------------------------------------------------------------------- #
# SNOOP-inlezer
# --------------------------------------------------------------------------- #
def _snoop_bestand(voorregels: list[list] | None = None) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for regel in voorregels or []:
        ws.append(regel)
    ws.append(
        ["Registratienummer", "Medewerker", "Datum", "Starttijd", "Eindtijd",
         "Werkelijke starttijd", "Werkelijke eindtijd", "Gewerkte uren", "Locatie",
         "Werkgever op datum shift", "Type uitzendkracht", "Tarief uitzendbureau"]
    )
    ws.append(
        ["1", "Marius Mic", "2026-06-15 00:00:00", "07:00", "15:00", None, None,
         7.5, "EW5", "Level One", "Uitzendkracht", "B2 Flex"]
    )
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_snoop_met_titelregels_boven_de_kolomkoppen():
    """Exports beginnen soms met een titel of lege regels; de kopregel staat
    dan niet op rij 1."""
    from app.services.ingest import lees_snoop

    medewerkers = lees_snoop(
        _snoop_bestand([["SNOOP export Kwekerij Baas"], [], ["periode week 25"]])
    )
    assert [m.naam for m in medewerkers] == ["Marius Mic"]
    assert medewerkers[0].loonschaal == "B2 Flex"


def test_snoop_zonder_titelregels():
    from app.services.ingest import lees_snoop

    assert len(lees_snoop(_snoop_bestand())) == 1


def test_snoop_met_onbekende_kolommen_noemt_wat_er_wel_staat():
    from io import BytesIO

    from openpyxl import Workbook

    from app.services.ingest import lees_snoop

    wb = Workbook()
    wb.active.append(["Naam", "Bedrag"])
    buffer = BytesIO()
    wb.save(buffer)
    with pytest.raises(ValueError) as fout:
        lees_snoop(buffer.getvalue())
    assert "Medewerker" in str(fout.value)  # wat verwacht wordt
    assert "Naam" in str(fout.value)  # en wat er gevonden is
