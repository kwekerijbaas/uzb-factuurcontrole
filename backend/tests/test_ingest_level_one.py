"""Tests voor Level One's eigen tariefexport (SPEC §6).

Gevalideerd tegen de export per 01-07-2026: 99 kaartcodes (33 schalen x Vast,
Flex en Seizoen) en 33 CAO-lonen. De 'oud'-kolom reproduceerde 391 van de 396
waarden uit de geconsolideerde kaart per 01-01-2026.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.ingest.level_one import lees_level_one_export
from app.services.tarief import CAT_100, CAT_135, CAT_150, CAT_200, CAT_FEESTDAG

_KOP = [
    "Relatie", "Code", "Nr.", None,
    "Loon oud", "Loon per 1/7/26", "EF", None,
    "Component", "Percentage",
    "Vast oud", "Vast per 1/7/26", None, None,
    "Flex oud", "Flex per 1/7/26", None, None,
    "Seizoen oud", "Seizoen per 1/7/26",
]


def _schaal(code, loon_oud, loon_nieuw):
    return ["A. Baas", code, "291902", None, loon_oud, loon_nieuw, None, None, None, None]


def _component(oms, pct, vast, flex, seiz):
    """(oud, nieuw) per contractvorm."""
    return [
        None, None, None, None, None, None, None, "Verwijderen", oms, pct,
        vast[0], vast[1], None, None,
        flex[0], flex[1], None, None,
        seiz[0], seiz[1],
    ]


def _bestand(rijen) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(_KOP)
    for rij in rijen:
        ws.append(rij)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


BASIS = _bestand(
    [
        _schaal("GTB B2", 14.71, 14.99),
        _component("Loon normale uren", 100, (28.03, 28.51), (28.94, 29.44), (29.25, 29.77)),
        _component("Loon overwerkuren", 135, (28.03, 28.51), (28.94, 29.44), (29.25, 29.77)),
        _component("Loon overwerkuren", 150, (32.26, 32.86), (33.92, 34.54), (34.12, 34.76)),
        _component("Loon overwerkuren", 200, (43.34, 44.18), (44.97, 45.84), (46.51, 47.40)),
        _component("Loon bijzondere uren", 150, (38.71, 39.43), (40.01, 40.76), (40.57, 41.33)),
        _schaal("GTB B4", 15.07, 15.07),
        _component("Loon normale uren", 100, (28.66, 28.66), (29.59, 29.59), (29.92, 29.92)),
    ]
)


def test_contractvormen_worden_eigen_kaartcodes():
    """Vast, Flex en Seizoen delen het CAO-loon maar niet het tarief."""
    export, _ = lees_level_one_export(BASIS, "nieuw")

    assert export.tarieven["B2V"][CAT_100] == Decimal("28.51")
    assert export.tarieven["B2F"][CAT_100] == Decimal("29.44")
    assert export.tarieven["B2S"][CAT_100] == Decimal("29.77")
    assert export.lonen["B2"] == Decimal("14.99")  # één loon voor alle drie


def test_oude_kolom_leest_de_vorige_tarieven():
    export, _ = lees_level_one_export(BASIS, "oud")
    assert export.tarieven["B2F"][CAT_100] == Decimal("28.94")
    assert export.lonen["B2"] == Decimal("14.71")


def test_beide_150_regels_worden_onderscheiden():
    """150 komt twee keer voor: als overwerk en als bijzondere uren. Het
    percentage alleen volstaat dus niet om de categorie te bepalen."""
    export, _ = lees_level_one_export(BASIS, "nieuw")
    tarieven = export.tarieven["B2F"]

    assert tarieven[CAT_150] == Decimal("34.54")  # overwerk
    assert tarieven[CAT_FEESTDAG] == Decimal("40.76")  # bijzondere uren
    assert tarieven[CAT_135] == Decimal("29.44")
    assert tarieven[CAT_200] == Decimal("45.84")


def test_schaal_zonder_loonsverhoging_houdt_zijn_tarief():
    export_oud, _ = lees_level_one_export(BASIS, "oud")
    export_nieuw, _ = lees_level_one_export(BASIS, "nieuw")
    assert export_oud.tarieven["B4F"] == export_nieuw.tarieven["B4F"]
    assert export_oud.lonen["B4"] == export_nieuw.lonen["B4"]


def test_peilmoment_wordt_teruggemeld():
    """Zodat de gebruiker ziet welke kolom is ingelezen."""
    _, waarschuwingen = lees_level_one_export(BASIS, "nieuw")
    assert any("1/7/26" in w for w in waarschuwingen)


def test_onbekende_kolomkoppen_worden_geweigerd():
    wb = Workbook()
    wb.active.append(["Naam", "Waarde"])
    wb.active.append(["B2", 1])
    buffer = BytesIO()
    wb.save(buffer)
    with pytest.raises(ValueError, match="kolomkoppen niet herkend"):
        lees_level_one_export(buffer.getvalue())


def test_ongeldige_kolomkeuze():
    with pytest.raises(ValueError, match="oud"):
        lees_level_one_export(BASIS, "allebei")


@pytest.mark.parametrize(
    "kop,verwacht",
    [
        ("Loon per 1/7/26", date(2026, 7, 1)),
        ("Vast per 01-07-2026", date(2026, 7, 1)),
        ("Flex per 1-8-26", date(2026, 8, 1)),
        ("Loon oud", None),
        ("Vast per 31/2/26", None),  # bestaat niet
    ],
)
def test_ingangsdatum_uit_de_kolomkop(kop, verwacht):
    """De ingangsdatum staat in het bestand; die hoeft niemand over te typen."""
    from app.services.ingest.level_one import peildatum

    assert peildatum(kop) == verwacht


def test_export_kent_zijn_eigen_ingangsdatum():
    export, _ = lees_level_one_export(BASIS, "nieuw")
    assert export.ingangsdatum == date(2026, 7, 1)
    oud, _ = lees_level_one_export(BASIS, "oud")
    assert oud.ingangsdatum is None  # de oude kolom noemt geen datum


def test_export_zonder_lege_tussenkolom():
    """Level One levert de export niet altijd met evenveel lege tussenkolommen.
    Op een vaste positie rekenen leverde dan een bestand op waarin geen enkele
    tariefregel werd gevonden (BAAS_01072026.xlsx, juli 2026)."""
    kop = [
        "Relatie", "Code", "Nr.", None,
        "Loon oud", "Loon per 1 jul", "EF",
        "Component", "Percentage",
        "Vast oud", "Vast per 1 jul", None, None,
        "Flex oud", "Flex per 1 jul", None, None,
        "Seizoen oud", "Seizoen per 1 jul",
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(kop)
    ws.append(["A. Baas", "GTB B2", "291902", None, 14.71, 14.99, None, "Toevoegen", None])
    ws.append(
        [None, None, None, None, None, None, None, "Loon normale uren", 100,
         28.03, 28.51, None, None, 28.94, 29.44, None, None, 29.25, 29.77]
    )
    buffer = BytesIO()
    wb.save(buffer)

    export, _ = lees_level_one_export(buffer.getvalue(), "nieuw")
    assert export.lonen["B2"] == Decimal("14.99")
    assert export.tarieven["B2F"][CAT_100] == Decimal("29.44")


def test_kolomkop_zonder_jaartal():
    """'per 1 jul' noemt geen jaar; dan wordt het dichtstbijzijnde jaar gekozen
    en op het scherm getoond, zodat een verkeerde gok opvalt."""
    from app.services.ingest.level_one import peildatum

    assert peildatum("Loon per 1 jul", date(2026, 8, 18)) == date(2026, 7, 1)
    assert peildatum("Loon per 1 jul", date(2027, 2, 1)) == date(2027, 7, 1)
    assert peildatum("Loon per 1 juli 2026", date(2030, 1, 1)) == date(2026, 7, 1)
