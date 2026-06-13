"""
Bureau-ONAFHANKELIJKE factuurcontrole uitzendbureaus — de MOTOR.

Gebruik:
    python controle_uzb.py --map "C:\\pad\\naar\\weekmap"

Werkwijze:
  1. Vindt alle *PurchaseInvoice/*.pdf in de map.
  2. Per PDF: bepaalt welk uitzendbureau het is via de profielen in bureau_profielen.py.
     Onbekende facturen (geen profiel) worden overgeslagen en gemeld.
  3. Leest de factuur uit met het juiste profiel; leidt week(en) uit de inhoud af.
  4. Zoekt per (bureau, week) het doorgegeven-uren-bestand (op weeknr + bureau-hint).
  5. Vergelijkt uren + tarief (motorlogica, identiek voor elk bureau).
  6. Schrijft per (bureau, week): rapportage + interne/externe mail in de map.
  7. Werkt het gedeelde openstaande-posten-register bij.

Nieuwe bureaus toevoegen = profiel toevoegen in bureau_profielen.py (motor blijft gelijk).
"""
import re
import os
import sys
import glob
import argparse
import datetime
import difflib
from collections import Counter
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import bureau_profielen as bp

DEFAULT_MAP = r'C:\Users\dieter.KWEKERIJBAAS\Downloads\Claude tst'
REGISTER_NAAM = 'Openstaande_posten_register_uitzendbureaus.xlsx'
TOLERANTIE_UREN = 0.05
INTERN_DATA_AANGEPAST = True
BETAAL_PAS_NA_CREDIT = True
SCHAAL_TYPE = {'F': 'Flex', 'V': 'Vast', 'S': 'Seizoen'}

FONT_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
FONT_HEADER = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_SUB = Font(name='Arial', size=10, bold=True)
FONT_N = Font(name='Arial', size=10)
FONT_TOT = Font(name='Arial', size=10, bold=True)
FILL_TITLE = PatternFill('solid', start_color='1F4E78')
FILL_HEADER = PatternFill('solid', start_color='2E75B6')
FILL_SUB = PatternFill('solid', start_color='DDEBF7')
FILL_OK = PatternFill('solid', start_color='C6EFCE')
FILL_ROOD = PatternFill('solid', start_color='FFC7CE')
FILL_GEEL = PatternFill('solid', start_color='FFEB9C')
FILL_BLAUW = PatternFill('solid', start_color='B4C7E7')  # te verifiëren door HR
FILL_TOT = PatternFill('solid', start_color='F2F2F2')
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_R = Alignment(horizontal='right', vertical='center')
THIN = Side(border_style='thin', color='808080')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

normalize_naam = bp.normalize_naam


def schaal_type(s):
    if not s:
        return '', ''
    suf = s[-1].upper()
    return (s[:-1], SCHAAL_TYPE[suf]) if suf in SCHAAL_TYPE else (s, '')


def tarief_match(ft, st):
    if ft is None or st is None:
        return None
    return round(float(ft), 2) == round(float(st), 2)


def week_uit_tekst(t):
    m = re.search(r'Week[:\s]*?(\d{4})[-\s]?(\d{2})', t)
    return (m.group(1), m.group(2)) if m else None


# ---- doorgegeven uren (gedeeld, robuuste kolomdetectie) ----
def lees_doorgegeven_auto(pad):
    df = pd.read_excel(pad, header=None)
    nrows, ncols = df.shape
    jaar = week = None
    for i in range(min(5, nrows)):
        for j in range(min(3, ncols)):
            v = df.iloc[i, j]
            if isinstance(v, str):
                w = week_uit_tekst(v)
                if w:
                    jaar, week = w
                    break
        if week:
            break
    med_col = tot_col = opm_col = header_row = None
    for i in range(min(8, nrows)):
        for j in range(ncols):
            v = df.iloc[i, j]
            if isinstance(v, str) and v.strip().lower() == 'medewerkers':
                header_row, med_col = i, j
        if header_row is not None:
            for j in range(ncols):
                v = df.iloc[header_row, j]
                if isinstance(v, str):
                    vl = v.strip().lower()
                    if vl == 'totaal':
                        tot_col = j
                    elif vl.startswith('opmerking'):
                        opm_col = j
            break
    if med_col is None:
        return jaar, week, {}
    if tot_col is None:
        tot_col = med_col + 8
    res = {}
    for i in range(header_row + 1, nrows):
        med = df.iloc[i, med_col]
        if not isinstance(med, str):
            continue
        m = re.match(r'^\s*(\d+)\s+(.+)$', med.strip())
        if not m:
            continue
        totaal = df.iloc[i, tot_col]
        if not isinstance(totaal, (int, float)) or pd.isna(totaal):
            # Totaal is leeg/formule (nog niet uitgerekend) -> tel zelf de dagkolommen op
            dag = sum(c for c in df.iloc[i, med_col + 1:tot_col]
                      if isinstance(c, (int, float)) and pd.notna(c))
            if dag == 0:
                continue
            totaal = dag
        opm = df.iloc[i, opm_col] if (opm_col is not None and opm_col < ncols) else ''
        res[m.group(1)] = {'tag': m.group(1), 'naam': m.group(2).strip(),
                           'naam_norm': normalize_naam(m.group(2)), 'totaal': float(totaal),
                           'opmerking': opm if isinstance(opm, str) else ''}
    return jaar, week, res


DAGNAMEN = ['ma', 'di', 'wo', 'do', 'vr', 'za', 'zo']


def lees_doorgegeven_dagen(pad):
    """Per medewerker-rij: (tag, naam, jaar, week, set(dag-indices met uren>0)).
    NIET op tag collapsen — zo zien we als 2 personen dezelfde tag in dezelfde week delen.
    Voor de DAG-bewuste tag-conflictcontrole (tag is uniek per dag, niet over tijd)."""
    try:
        df = pd.read_excel(pad, header=None)
    except Exception:
        return []
    nrows, ncols = df.shape
    jaar = week = None
    mfn = re.search(r'\bw(?:ee)?k\s*0*(\d{1,2})\b', os.path.basename(pad), re.I)
    if mfn:
        week = int(mfn.group(1))
    for i in range(min(6, nrows)):
        for j in range(min(4, ncols)):
            v = df.iloc[i, j]
            if isinstance(v, str):
                w = week_uit_tekst(v)
                if w:
                    jaar = jaar or w[0]
                    week = week or int(w[1])
    jaar = jaar or '2026'
    med_col = tot_col = header_row = None
    for i in range(min(8, nrows)):
        for j in range(ncols):
            v = df.iloc[i, j]
            if isinstance(v, str) and v.strip().lower() == 'medewerkers':
                header_row, med_col = i, j
        if header_row is not None:
            for j in range(ncols):
                v = df.iloc[header_row, j]
                if isinstance(v, str) and v.strip().lower() == 'totaal':
                    tot_col = j
            break
    if med_col is None:
        return []
    if tot_col is None:
        tot_col = med_col + 8
    out = []
    for i in range(header_row + 1, nrows):
        med = df.iloc[i, med_col]
        if not isinstance(med, str):
            continue
        m = re.match(r'^\s*(\d+)\s+(.+)$', med.strip())
        if not m:
            continue
        dagen = {j - (med_col + 1) for j in range(med_col + 1, tot_col)
                 if isinstance(df.iloc[i, j], (int, float)) and pd.notna(df.iloc[i, j]) and df.iloc[i, j] > 0}
        out.append((m.group(1), m.group(2).strip(), str(jaar), int(week) if week else None, dagen))
    return out


def bouw_dag_occupatie(map_pad):
    """(tag, jaar, week, dag-idx) -> set(naam_norm) over alle doorgegeven-bestanden in de map.
    Basis voor de dag-bewuste tag-conflictcontrole."""
    occ = {}
    kand = [p for p in glob.glob(os.path.join(map_pad, '*.xls')) + glob.glob(os.path.join(map_pad, '*.xlsx'))
            if re.search(r'\bwk\s*\d', os.path.basename(p), re.I)
            and not re.search(r'snoop|factuurcalc|tariefkaart|rapportage|credit|afwijking|openstaande|register|overzicht',
                              os.path.basename(p), re.I)]
    for p in kand:
        for tag, naam, jaar, week, dagen in lees_doorgegeven_dagen(p):
            nn = normalize_naam(naam)
            for d in dagen:
                occ.setdefault((str(tag), jaar, week, d), set()).add(nn)
    return occ


def vind_doorgegeven(map_pad, week, hint):
    """Zoek doorgegeven-uren-bestand(en) voor (week, bureau-hint).
    hint kan str of list zijn. Bij meerdere passende bestanden (bv. Payroll Jeugd +
    Volwassen) worden de medewerkers samengevoegd."""
    hints = [hint] if isinstance(hint, str) else list(hint or [])
    hints = [h.lower() for h in hints if h]
    gematcht = []   # (pad, data, jaar) van bestanden met een hint-match
    fallback = []   # zonder hint-match (zelfde week)
    for pad in glob.glob(os.path.join(map_pad, '*.xls')) + glob.glob(os.path.join(map_pad, '*.xlsx')):
        b = os.path.basename(pad)
        bl = b.lower()
        if b.startswith('~$') or b.startswith('Rapportage') or b.startswith('Openstaande') or 'tariefkaart' in bl:
            continue
        try:
            jaar, wk, data = lees_doorgegeven_auto(pad)
        except Exception:
            continue
        if wk == week and data:
            if any(h in bl for h in hints):
                gematcht.append((pad, data, jaar))
            else:
                fallback.append((pad, data, jaar))
    bron = gematcht if gematcht else (fallback if (not hints) else [])
    if not bron:
        return None
    # Samenvoegen
    samen = {}
    jaar = None
    paden = []
    for pad, data, jr in bron:
        samen.update(data)
        jaar = jaar or jr
        paden.append(os.path.basename(pad))
    return ' + '.join(paden), samen, jaar


def _extract_totaal(text):
    """Haal het vermelde totaalbedrag (excl. BTW) uit een factuur/creditnota-tekst.
    Werkt voor de verschillende bureau-formaten; handelt trailing-minus af."""
    for pat in [r'TOTAAL\s*EXCL\.?\s*BTW\s*:?\s*[^\d\-]*(-?[\d.,]+-?)',   # CervoKordaat
                r'Totaal\s+(?:-?\d+:\d{2}\s+)?(-?[\d.,]+-?)',              # Level One (met of zonder HH:MM)
                r'Sub-totaal[^\n]*\n\s*(-?[\d.,]+-?)']:                     # Sterk Werk
        m = re.search(pat, text, re.I)
        if m:
            return round(bp.komma_to_float(m.group(1)), 2)
    return 0.0


def _credit_omschrijving(text):
    """Korte omschrijving van een naamloze correctie/creditnota — beschrijft WAAR de
    correctie op slaat (bv. '150% -> bijzonder tarief'), zodat zoeken niet nodig is."""
    keys = ('correctie i.v.m', 'aanpassing tarief', 'omgezet naar', 'onregelmatige uren',
            'bijzondere uren', 'reiskosten', 'correctie van factuur')
    found = []
    for ln in text.split('\n'):
        l = ln.strip()
        ll = l.lower()
        if l and len(l) < 130 and any(k in ll for k in keys):
            if l not in found:
                found.append(l)
    return ' | '.join(found[:2])


NL_MAAND = {'januari': 1, 'februari': 2, 'maart': 3, 'april': 4, 'mei': 5, 'juni': 6,
            'juli': 7, 'augustus': 8, 'september': 9, 'oktober': 10, 'november': 11, 'december': 12}


def _factuurdatum(text, code):
    """Factuurdatum uit de PDF-tekst. Formats verschillen per bureau:
    SW = 'RUINEN, 13 mei 2026'; CK = '21/05/2026'; L1/LP = 'DD-MM-YYYY€' (vlak vóór het bedrag).
    Returns datetime.date of None."""
    if code == 'SW':
        m = re.search(r'RUINEN,\s*(\d{1,2})\s+([A-Za-zé]+)\s+(\d{4})', text)
        if m:
            mnd = NL_MAAND.get(m.group(2).lower())
            if mnd:
                try:
                    return datetime.date(int(m.group(3)), mnd, int(m.group(1)))
                except ValueError:
                    pass
        return None
    if code == 'CK':
        m = re.search(r'(\d{2})/(\d{2})/(\d{4})', text)
        if m:
            try:
                return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                pass
        return None
    # L1 / LP / overig: DD-MM-YYYY direct vóór het €-bedrag (onderscheidt van incasso-/machtigingsdatum)
    m = re.search(r'(\d{2})-(\d{2})-(\d{4})\s*€', text)
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


def _hoofdweek(regels, text=''):
    """Weeknummer dat voor de MEESTE uitzendkrachten op de factuur staat (modale week).
    Eén FactuurRegel = één (week, medewerker), dus tellen per week ≈ medewerkers per week.
    Fallback (bv. CK-reiskostencredits zonder urenregels): modale week uit 'WK: 2026NN' in de tekst."""
    c = Counter(str(r.week).strip() for r in regels if str(getattr(r, 'week', '')).strip())
    if not c and text:
        c = Counter(m.group(1) for m in re.finditer(r'WK:?\s*\d{4}(\d{2})', text))
    if not c:
        return ''
    wk = max(c.items(), key=lambda kv: kv[1])[0]
    return f'{int(wk):02d}' if wk.isdigit() else wk


# ---- facturen via profielen ----
def lees_facturen(map_pad):
    import pdfplumber
    per_bureau_week = {}   # (code, week) -> {naam_norm: FactuurRegel}
    factuur_info = []
    onbekend = []
    credits = []           # creditfacturen: {bestand, code, bureau, factuurnr, regels}
    for pad in glob.glob(os.path.join(map_pad, '*.pdf')):
        b = os.path.basename(pad)
        try:
            with pdfplumber.open(pad) as pdf:
                text = '\n'.join((p.extract_text() or '') for p in pdf.pages)
        except Exception:
            continue
        prof = bp.herken_bureau(b, text[:2000])
        if prof is None:
            bl = b.lower()
            is_export = any(k in bl for k in ('nitea', 'snoop', 'plan', 'overzicht', 'rapportage'))
            if not is_export and re.search(r'factuur|invoice|uitzendkracht', (b + ' ' + text[:600]).lower()):
                onbekend.append(b)
            continue
        mfn = re.search(r'_(\d{5,9})_', b) or re.search(r'\b(\d{6,9})\b', b)
        factuurnr = mfn.group(1) if mfn else b
        regels = prof.parse_factuur(text, factuurnr)
        # Credit-/correctie-detectie:
        #   Creditnota = ICN-prefix / 'CREDITFACTUUR' / netto NEGATIEF bedrag (geld terug).
        #   Correctiefactuur = 'Correctie van factuur' OF bevat negatieve regels (re-rating:
        #     trailing-minus 'X,XX-' of negatieve uren '-NN:NN').
        # Beide NIET in de gewone weekcontrole mengen -> apart afletteren/beoordelen.
        netto_bedrag = sum(r.bedrag for r in regels)
        netto_uren = sum(r.uren_totaal for r in regels)
        tl = text.lower()
        heeft_negatief = re.search(r'\d,\d{2}-|-\d+:\d{2}', text) is not None
        is_creditnota = (re.search(r'ICN\d', b) is not None or 'creditfactuur' in tl or netto_bedrag < -0.01)
        is_correctie = ('correctie van factuur' in tl or heeft_negatief)
        if is_creditnota or is_correctie:
            bedrag = round(netto_bedrag, 2)
            if abs(bedrag) < 0.01:   # regel-som mist (bv. CK reiskosten) -> vermeld totaal pakken
                bedrag = _extract_totaal(text)
            # Type bepalen:
            #  Tariefcorrectie = zelfde uren oud+nieuw tarief -> netto uren ~0, saldo = verschil
            #  Creditnota = expliciete credit / netto geld terug
            #  Correctiefactuur = overige correcties (netto uren <> 0)
            if abs(netto_uren) < 0.5 and heeft_negatief:
                ctype = 'Tariefcorrectie'
            elif is_creditnota:
                ctype = 'Creditnota'
            else:
                ctype = 'Correctiefactuur'
            credits.append({'bestand': b, 'code': prof.code, 'bureau': prof.naam,
                            'factuurnr': factuurnr, 'regels': regels,
                            'weken': sorted({r.week for r in regels}),
                            'netto_bedrag': bedrag, 'netto_uren': round(netto_uren, 2),
                            'type': ctype, 'omschrijving': _credit_omschrijving(text),
                            'factuurdatum': _factuurdatum(text, prof.code), 'hoofdweek': _hoofdweek(regels, text)})
            factuur_info.append({'bestand': b, 'bureau': prof.naam, 'code': prof.code,
                                 'factuurnr': factuurnr, 'weken': sorted({r.week for r in regels}), 'credit': True})
            continue
        weken = sorted({r.week for r in regels})
        for r in regels:
            key = (prof.code, r.week)
            per_bureau_week.setdefault(key, {})[r.naam_norm] = r
        factuur_info.append({'bestand': b, 'bureau': prof.naam, 'code': prof.code,
                             'factuurnr': factuurnr, 'weken': weken, 'credit': False,
                             'factuurdatum': _factuurdatum(text, prof.code)})
    return per_bureau_week, factuur_info, onbekend, credits


def snoop_schaal_voor_week(srec, jaar, week):
    """Schaal volgens SNOOP die geldt in de opgegeven ISO-week. Tijd-bewust: als iemand
    halverwege het jaar een andere schaal krijgt, pakt deze de schaal die op dat moment gold.
    Prioriteit: meest voorkomende in die week → laatste vóór die week → eerste na die week
    → modale schaal als laatste redmiddel."""
    if not srec:
        return None
    per = srec.get('per_datum') or []
    if not per:
        return srec.get('schaal')
    try:
        wk_start = datetime.date.fromisocalendar(int(jaar), int(week), 1)
        wk_end = wk_start + datetime.timedelta(days=6)
    except Exception:
        return srec.get('schaal')
    in_week = [s for (d, s) in per if wk_start <= d <= wk_end and s]
    if in_week:
        return Counter(in_week).most_common(1)[0][0]
    before = [(d, s) for (d, s) in per if d < wk_start and s]
    if before:
        return before[-1][1]
    after = [(d, s) for (d, s) in per if d > wk_end and s]
    if after:
        return after[0][1]
    return srec.get('schaal')


# ---- HR-factuurkaart: per medewerker (op tag) schaal + tarief, om SNOOP-gaten te vullen ----
FK_FOLDER = r'C:\Users\dieter.KWEKERIJBAAS\Kwekerij Baas\Finance - Uitzendbureaus\Werkbestanden factuurcontrole'
FK_SPEC = {
    'L1': ('L1 Factuurcalculatie*.xlsx', 'Werknemers+schaal+tarief',
           {'t_100_135': '100%/135%', 't_150': '150%', 't_200': '200%', 't_bijzonder': '150% bijzonder ; 50% nachturen'}),
    'LP': ('L1 jeugd en payroll Factuurcalculatie*.xlsx', 'Werknemers+schaal+tarief',
           {'t_100_135': '100%', 't_150': '150%', 't_200': '200%'}),
    'CK': ('*ervo*ordaat*Factuurcalculatie*.xlsx', 'Werknemers+tarief+schaal',
           {'t_100': '100%', 't_135': '135%', 't_150': '150%', 't_200': '200%'}),
    'SW': ('Sterk werk*Factuurcalculatie*.xlsx', 'Medewerkers+schaal+tarief',
           {'t_100_135': '100%/135%', 't_150': '150%', 't_200': '200%', 't_feestdag': 'Feestdag', 't_nacht50': '50% nacht'}),
}


def laad_factuurkaart_werknemers(folder=FK_FOLDER):
    """Leest per bureau de HR/finance-werkbestanden (per medewerker op tag-nummer: schaal + tarief).
    Returns {code: {'by_tag': {tag: rec}, 'by_naam': {naam_norm: rec}, 'by_ach': {achternaam: [rec]}}}."""
    out = {}
    if not os.path.isdir(folder):
        print(f'LET OP: HR-factuurkaart-map niet gevonden: {folder}')
        return out
    for code, (pat, tab, colmap) in FK_SPEC.items():
        kand = glob.glob(os.path.join(folder, pat))
        if not kand:
            continue
        try:
            df = pd.read_excel(kand[0], sheet_name=tab, header=0)
        except Exception as ex:
            print(f'   (HR-factuurkaart {code} niet leesbaar: {ex})')
            continue
        by_tag, by_naam, by_ach = {}, {}, {}
        for _, row in df.iterrows():
            if 'Nummer' not in df.columns or pd.isna(row['Nummer']):
                continue
            try:
                tag = str(int(row['Nummer']))
            except Exception:
                tag = str(row['Nummer']).strip()
            rec = {'schaal': str(row.get('Schaal', '')).strip(), 'tag': tag, 'bron': 'HR-factuurkaart'}
            for key, col in colmap.items():
                v = row.get(col)
                try:
                    rec[key] = float(v) if pd.notna(v) and str(v).strip() not in ('', 'nan') else None
                except (TypeError, ValueError):
                    rec[key] = None
            if code == 'LP':
                rec['t_bijzonder'] = rec.get('t_150')  # jeugd: bijzonder150 toetsen tegen 150%
            vn = '' if pd.isna(row.get('Voornaam')) else str(row.get('Voornaam', '')).strip()
            an = '' if pd.isna(row.get('Achternaam')) else str(row.get('Achternaam', '')).strip()
            rec['naam'] = f'{vn} {an}'.strip()
            by_tag[tag] = rec
            nn = bp.normalize_naam(f'{vn} {an}')
            if nn:
                by_naam[nn] = rec
            an_n = bp.normalize_naam(an)
            if an_n:
                by_ach.setdefault(an_n, []).append(rec)
        out[code] = {'by_tag': by_tag, 'by_naam': by_naam, 'by_ach': by_ach}
        print(f'HR-factuurkaart {code}: {len(by_tag)} medewerkers uit {os.path.basename(kand[0])}')
    return out


def _naam_tokens(s):
    """Set van naam-tokens (>=3 letters) voor achternaam-vergelijking."""
    return {t for t in re.sub(r'[().,\.]', ' ', str(s).lower()).split() if len(t) >= 3 and t.isalpha()}


def _fk_lookup(fk, tag, naam_norm):
    """Zoek een medewerker in de HR-factuurkaart: eerst op tag, dan exacte naam, dan fuzzy, dan achternaam.
    LET OP: tagnummers worden in Nitea hergebruikt (na verloop van tijd een andere persoon). Een tag-hit
    wordt daarom alleen geaccepteerd als de naam consistent is (gedeeld achternaam-token); anders val terug
    op naam-matching, zodat we niet de schaal van de vorige tag-houder pakken."""
    if not fk:
        return None
    if tag and str(tag) in fk['by_tag']:
        rec = fk['by_tag'][str(tag)]
        if not naam_norm or (_naam_tokens(rec.get('naam', '')) & _naam_tokens(naam_norm)):
            return rec
        # tag hoort (nu) bij een andere persoon -> negeer de tag, ga verder op naam
    bn = fk.get('by_naam', {})
    if naam_norm in bn:
        return bn[naam_norm]
    m = difflib.get_close_matches(naam_norm, list(bn.keys()), n=1, cutoff=0.86)
    if m:
        return bn[m[0]]
    toks = [t for t in re.sub(r'[().,\.]', ' ', naam_norm).split() if len(t) >= 3]
    for t in toks:
        cands = fk.get('by_ach', {}).get(t)
        if cands:
            if len(cands) == 1:
                return cands[0]
            for rc in cands:  # disambigueer op een gedeeld voornaam-token
                vn = rc['naam'].split()[0].lower() if rc['naam'] else ''
                if vn and vn in toks:
                    return rc
            return cands[0]
    return None


# ---- vergelijking (motor, identiek per bureau) ----
def vergelijk(prof, week, jaar, doorgegeven, fac_per_naam, snoop_data, tariefkaart_map, factuurkaart=None):
    dg_op_naam = {d['naam_norm']: d for d in doorgegeven.values()}
    dg_namen = list(dg_op_naam.keys())
    snoop_namen = list(snoop_data.keys())

    def snoop_lookup(naam_norm):
        """SNOOP-inschaling op naam (exact, anders fuzzy). Returns record of None."""
        if naam_norm in snoop_data:
            return snoop_data[naam_norm]
        kand = difflib.get_close_matches(naam_norm, snoop_namen, n=1, cutoff=0.88)
        return snoop_data[kand[0]] if kand else None

    gebruikt = set()  # voorkom dat 2 factuurnamen aan dezelfde doorgegeven persoon koppelen

    def _achternaam(n):
        return n.split()[-1] if n.split() else ''

    def _initialen(n):
        # 'r.e. grasu' -> ['r','e'] ; 'robert ionut grasu' -> ['r','i']
        letters = []
        for tok in n.split()[:-1]:
            for ch in tok:
                if ch.isalpha():
                    letters.append(ch)
                    break
        return letters

    def match_dg(norm):
        if norm in dg_op_naam and norm not in gebruikt:
            gebruikt.add(norm)
            return dg_op_naam[norm], False
        # kandidaten met (vrijwel) dezelfde achternaam, nog niet gebruikt
        ach = _achternaam(norm)
        kand = [dn for dn in dg_namen if dn not in gebruikt
                and difflib.SequenceMatcher(None, ach, _achternaam(dn)).ratio() > 0.85]
        if not kand:
            # fallback: fuzzy op volledige naam
            fz = [k for k in difflib.get_close_matches(norm, dg_namen, n=3, cutoff=0.82) if k not in gebruikt]
            kand = fz
        if not kand:
            return None, False
        if len(kand) == 1:
            gebruikt.add(kand[0])
            return dg_op_naam[kand[0]], True
        # meerdere zelfde achternaam -> disambigueer op eerste initiaal van de voornaam
        inits = _initialen(norm)
        eerste = inits[0] if inits else ''
        beste = None
        for dn in kand:
            dn_inits = _initialen(dn)
            if eerste and dn_inits and dn_inits[0] == eerste:
                beste = dn
                break
        if beste is None:
            beste = kand[0]
        gebruikt.add(beste)
        return dg_op_naam[beste], True

    rijen = []
    gematcht = set()
    for naam_norm, fac in sorted(fac_per_naam.items()):
        dg, fuzzy = match_dg(naam_norm)
        if dg:
            gematcht.add(dg['naam_norm'])
        tag = dg['tag'] if dg else '?'
        dg_uren = dg['totaal'] if dg else 0.0
        verschil = round(fac.uren_totaal - dg_uren, 2)

        # Schaal/inschaling uit SNOOP (op naam; bij voorkeur via doorgegeven-naam) ->
        # tarief uit de gecombineerde tariefkaart. Geen inschaling -> UZB-tarief.
        snoop_naam = dg['naam_norm'] if dg else naam_norm
        srec = snoop_lookup(snoop_naam)
        # Tijd-bewust: schaal die in DEZE week (jaar/week) gold volgens SNOOP — niet de modus.
        schaal_snoop_raw = snoop_schaal_voor_week(srec, jaar, week) if srec else None
        schaal_code = schaal_snoop_raw
        sch = tariefkaart_map.get(schaal_code) if schaal_code else None
        tarief_bron = 'SNOOP-inschaling' if sch is not None else ''
        # Fallback: geen (bruikbare) SNOOP-inschaling -> HR-factuurkaart (per tag/naam)
        schaal_werkbestand_raw = None
        if factuurkaart:
            frec = _fk_lookup(factuurkaart, tag, naam_norm)
            if frec is None and dg:
                frec = _fk_lookup(factuurkaart, dg.get('tag'), dg['naam_norm'])
            if frec:
                schaal_werkbestand_raw = frec.get('schaal')
                if sch is None:
                    sch = frec
                    schaal_code = frec.get('schaal') or schaal_code
                    tarief_bron = 'HR-factuurkaart'

        tarief_status = 'n.v.t.'
        euro_impact = 0.0
        breakdown = []
        is_tech = tag in prof.techniek_tags
        if is_tech:
            tarief_status = 'Aparte tariefafspraak'
        elif sch is not None:
            afw, ok = [], 0
            for label, uren, ft, st in prof.tarief_checks(fac, sch):
                if ft is None or st is None:
                    continue
                if tarief_match(ft, st):
                    ok += 1
                else:
                    bedrag = (round(float(ft), 2) - round(float(st), 2)) * uren
                    euro_impact += bedrag
                    afw.append(label)
                    breakdown.append({'categorie': label, 'uren': round(uren, 2),
                                      'factuur_tarief': round(float(ft), 2), 'schaal_tarief': round(float(st), 2),
                                      'verschil': round(round(float(ft), 2) - round(float(st), 2), 2),
                                      'bedrag': round(bedrag, 2)})
            tarief_status = 'TARIEF WIJKT AF' if afw else ('OK' if ok else 'n.v.t.')
        else:
            # geen inschaling in SNOOP én niet in HR-factuurkaart -> UZB-tarief aanhouden
            tarief_status = 'UZB-tarief (geen inschaling bekend)'

        status = []
        if dg is None:
            status.append('NIET DOORGEGEVEN')
        if abs(verschil) > TOLERANTIE_UREN:
            status.append(f'UREN-AFWIJKING {verschil:+.2f}h')
        if tarief_status.startswith('UZB-tarief'):
            status.append('UZB-TARIEF (inschaling nog niet bekend)')
        if tarief_status == 'TARIEF WIJKT AF':
            status.append('OPENSTAAND - tariefcorrectie via UZB in latere factuur')
        if not status:
            status.append('OK')

        bs, ct = schaal_type(schaal_code or '')
        c = fac.categorieen
        inv_ot135 = round(c.get('ow135', {}).get('uren', 0.0), 2)
        inv_ot150 = round(c.get('ow150', {}).get('uren', 0.0), 2)
        inv_nacht = round(c.get('nacht50', {}).get('uren', 0.0) + c.get('bijzonder150', {}).get('uren', 0.0), 2)
        rijen.append({'tag': tag, 'naam': fac.naam_factuur, 'naam_norm': naam_norm,
                      'dg_uren': dg_uren, 'fac_uren': round(fac.uren_totaal, 2), 'verschil': verschil,
                      'tarief': fac.weergeef_tarief, 'bedrag': round(fac.bedrag, 2),
                      'inv_ot135': inv_ot135, 'inv_ot150': inv_ot150, 'inv_nacht': inv_nacht,
                      'schaal': schaal_code or '', 'basis_schaal': bs, 'contracttype': ct,
                      'tarief_status': tarief_status, 'euro_impact': round(euro_impact, 2),
                      'tarief_breakdown': breakdown, 'factuurnr': fac.factuurnr,
                      'status': ' + '.join(status), 'fuzzy': fuzzy, 'tarief_bron': tarief_bron,
                      'schaal_snoop_raw': schaal_snoop_raw, 'schaal_werkbestand_raw': schaal_werkbestand_raw,
                      'naam_dg': dg['naam'] if dg else ''})
    niet_gefac = [d for n, d in dg_op_naam.items() if n not in gematcht]
    return rijen, niet_gefac


def fill_status(status):
    if 'UREN-AFWIJKING' in status or 'NIET DOORGEGEVEN' in status:
        return FILL_ROOD
    if 'OPENSTAAND' in status:
        return FILL_GEEL
    return FILL_OK


def genereer_rapportage(map_pad, prof, week, jaar, doorgegeven, rijen, toeslag_data=None, niet_gefac=None):
    wk = f'{int(week):02d}'
    out = os.path.join(map_pad, f'Rapportage_Urencontrole_{prof.code}_WK{wk}_{jaar}.xlsx')
    wb = Workbook()
    niet_gefac = niet_gefac or []
    te_hoog = [r for r in rijen if r['tarief_status'] == 'TARIEF WIJKT AF' and r['euro_impact'] > 0.01]
    te_laag = [r for r in rijen if r['tarief_status'] == 'TARIEF WIJKT AF' and r['euro_impact'] < -0.01]
    naam_var = [r for r in rijen if r['fuzzy']]
    geen_schaal = [r for r in rijen if r['schaal'] == '']
    uren_afw = [r for r in rijen if 'UREN-AFWIJKING' in r['status'] or 'NIET DOORGEGEVEN' in r['status']]
    credit = round(sum(r['euro_impact'] for r in te_hoog), 2)
    netto = round(sum(r['euro_impact'] for r in (te_hoog + te_laag)), 2)
    # Uren-vergelijking PER GEMATCHTE MEDEWERKER (niet totaal-vs-totaal):
    # tot_dg = som doorgegeven uren van ALLEEN de medewerkers die op deze factuur staan.
    tot_dg = round(sum(r['dg_uren'] for r in rijen), 2)
    tot_fac = round(sum(r['fac_uren'] for r in rijen), 2)
    # Doorgegeven medewerkers die NIET op deze factuur staan (bv. andere categorie/factuur)
    niet_gefac_uren = round(sum(d['totaal'] for d in niet_gefac), 2)

    wsb = wb.active
    wsb.title = 'Beslissing'
    wsb.merge_cells('A1:D1')
    wsb['A1'] = f'BESLISSING — {prof.naam} week {wk} {jaar}'
    wsb['A1'].font = FONT_TITLE; wsb['A1'].fill = FILL_TITLE; wsb['A1'].alignment = ALIGN_C
    wsb.row_dimensions[1].height = 30
    wsb.merge_cells('A3:D3')
    if uren_afw:
        advies = 'ADVIES: NIET akkoord voor betaling — urenafwijkingen. Pas betalen na creditfactuur.'
        kl = FILL_ROOD
    elif te_hoog and BETAAL_PAS_NA_CREDIT:
        advies = (f'ADVIES: NOG NIET akkoord voor betaling. Uren kloppen ({tot_dg:.2f}h = {tot_fac:.2f}h), maar '
                  f'€{credit:.2f} te veel gefactureerd door onjuiste inschaling. AFSPRAAK: pas betalen na creditfactuur.')
        kl = FILL_ROOD
    else:
        advies = f'ADVIES: Factuur week {wk} akkoord voor betaling. Uren en tarieven kloppen.'
        kl = FILL_OK
    wsb['A3'] = advies; wsb['A3'].font = FONT_SUB; wsb['A3'].fill = kl; wsb['A3'].alignment = ALIGN_L
    wsb.row_dimensions[3].height = 56
    r = 5
    wsb.cell(row=r, column=1, value='STAP 1 — Afgehandeld: interne data bijgewerkt (groen)').font = FONT_SUB
    wsb.cell(row=r, column=1).fill = FILL_SUB; r += 1
    for i, h in enumerate(['Onderwerp', 'Wie/wat', 'Detail', 'Status'], start=1):
        c = wsb.cell(row=r, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
    r += 1
    blok = [('Naamspelling gelijkgetrokken', f"tag {x['tag']} {x['naam']}", f"Nitea \"{x['naam_dg']}\" vs factuur \"{x['naam']}\".", 'AFGEHANDELD') for x in naam_var]
    blok += [('Schaalindeling toegevoegd', f"tag {x['tag']} {x['naam']}", 'Ontbrekende schaal intern aangevuld.', 'AFGEHANDELD') for x in geen_schaal]
    if not blok:
        wsb.cell(row=r, column=1, value='Geen interne datapunten.').font = FONT_N; r += 1
    for vals in blok:
        for i, v in enumerate(vals, start=1):
            c = wsb.cell(row=r, column=i, value=v); c.font = FONT_N; c.border = BORDER; c.alignment = ALIGN_L; c.fill = FILL_OK
        wsb.row_dimensions[r].height = 32; r += 1
    r += 1
    wsb.cell(row=r, column=1, value='STAP 2 — Openstaande posten: tariefcorrectie via UZB (bewaken, niet groen)').font = FONT_SUB
    wsb.cell(row=r, column=1).fill = FILL_SUB; r += 1
    for i, h in enumerate(['Medewerker', 'Schaal moet', 'Gefactureerd', 'Effect (€)'], start=1):
        c = wsb.cell(row=r, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
    r += 1
    openst = te_hoog + te_laag
    if openst:
        for x in sorted(openst, key=lambda y: -y['euro_impact']):
            richting = 'te veel (credit verwacht)' if x['euro_impact'] > 0 else 'te weinig (UZB kan bijfactureren)'
            ts = f"€{x['tarief']:.3f}/u" if x['tarief'] is not None else ''
            for i, v in enumerate([f"tag {x['tag']} {x['naam']}", x['schaal'], ts, f"{x['euro_impact']:+.2f} ({richting})"], start=1):
                c = wsb.cell(row=r, column=i, value=v); c.font = FONT_N; c.border = BORDER
                c.alignment = ALIGN_R if i == 4 else ALIGN_L; c.fill = FILL_GEEL
            r += 1
        wsb.cell(row=r, column=1, value='Verwachte credit (excl. BTW)').font = FONT_TOT
        wsb.cell(row=r, column=4, value=f'+{credit:.2f}').font = FONT_TOT
        for i in range(1, 5):
            wsb.cell(row=r, column=i).fill = FILL_TOT; wsb.cell(row=r, column=i).border = BORDER
        wsb.cell(row=r, column=4).alignment = ALIGN_R; r += 1
        wsb.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        wsb.cell(row=r, column=1, value='ACTIE: factuur NIET betalen tot creditfactuur ontvangen. In overleg met UZB; bewaken in register.').font = FONT_N
        wsb.cell(row=r, column=1).fill = FILL_GEEL; wsb.cell(row=r, column=1).alignment = ALIGN_L
        wsb.row_dimensions[r].height = 32; r += 1
    else:
        wsb.cell(row=r, column=1, value='Geen openstaande tariefposten.').font = FONT_N; r += 1
    # Info: doorgegeven medewerkers die NIET op deze factuur staan
    if niet_gefac:
        r += 1
        wsb.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        wsb.cell(row=r, column=1, value=(f'INFO: {len(niet_gefac)} doorgegeven medewerker(s) ({niet_gefac_uren:.2f}h) staan NIET op deze '
                 f'factuur — vermoedelijk op een andere factuur/categorie (bv. Jeugd vs Volwassen) of latere factuur. '
                 f'Niet meegerekend als uren-afwijking.')).font = FONT_N
        wsb.cell(row=r, column=1).alignment = ALIGN_L; wsb.cell(row=r, column=1).fill = FILL_SUB
        wsb.row_dimensions[r].height = 40; r += 1
    for col, w in zip('ABCD', [30, 28, 22, 34]):
        wsb.column_dimensions[col].width = w

    ws2 = wb.create_sheet('Detail per medewerker')
    ws2.merge_cells('A1:K1')
    ws2['A1'] = f'Detail — {prof.naam} WK{wk} {jaar}'
    ws2['A1'].font = FONT_TITLE; ws2['A1'].fill = FILL_TITLE; ws2['A1'].alignment = ALIGN_C
    headers = ['Tag', 'Medewerker', 'Schaal', 'Type', 'Doorg.', 'Gefact.', 'Versch.', 'Tarief-check', 'Effect €', 'Factuur', 'Status']
    for i, h in enumerate(headers, start=1):
        c = ws2.cell(row=3, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
    rr = 4
    for row in sorted(rijen, key=lambda x: int(x['tag']) if str(x['tag']).isdigit() else 99999):
        vals = [int(row['tag']) if str(row['tag']).isdigit() else row['tag'], row['naam'], row['basis_schaal'],
                row['contracttype'], row['dg_uren'], row['fac_uren'], row['verschil'], row['tarief_status'],
                row['euro_impact'] if row['euro_impact'] else None, row['factuurnr'], row['status']]
        f = fill_status(row['status'])
        for i, v in enumerate(vals, start=1):
            c = ws2.cell(row=rr, column=i, value=v); c.font = FONT_N; c.border = BORDER; c.fill = f
            c.alignment = ALIGN_L if i in (2, 10, 11) else (ALIGN_C if i in (1, 3, 4, 8) else ALIGN_R)
        rr += 1
    ws2.cell(row=rr, column=2, value='TOTAAL').font = FONT_TOT
    ws2.cell(row=rr, column=5, value=tot_dg).font = FONT_TOT
    ws2.cell(row=rr, column=6, value=tot_fac).font = FONT_TOT
    for i in range(1, 12):
        ws2.cell(row=rr, column=i).fill = FILL_TOT; ws2.cell(row=rr, column=i).border = BORDER
    ws2.freeze_panes = 'A4'
    for i, w in enumerate([6, 28, 7, 7, 9, 9, 8, 18, 10, 16, 44]):
        ws2.column_dimensions[chr(65 + i)].width = w

    # ---- Toeslag-tab (berekend uit Nitea-kloktijden vs factuur) ----
    if toeslag_data:
        ws3 = wb.create_sheet('Toeslag')
        ws3.merge_cells('A1:K1')
        ws3['A1'] = f'Toeslagcontrole — {prof.naam} WK{wk} {jaar} (berekend uit Nitea vs factuur)'
        ws3['A1'].font = FONT_TITLE; ws3['A1'].fill = FILL_TITLE; ws3['A1'].alignment = ALIGN_C
        ws3.merge_cells('A2:K2')
        ws3['A2'] = ('Berekend uit werkelijke Nitea-kloktijden + CAO (normweek 38u; nacht Ma-Vr 20:00-06:00; '
                     'nachtdienst = hele 18:00-06:00). HARD signaal (geel) = nachtdienst gewerkt maar GEEN nacht-/50%-toeslag '
                     'op de factuur. OT-kolommen zijn TER INFO: bureaus tellen overuren/zaterdag verschillend, dus daar '
                     'geen harde vlag op.')
        ws3['A2'].font = Font(name='Arial', size=9, italic=True); ws3['A2'].alignment = ALIGN_L
        ws3.row_dimensions[2].height = 40
        hdr = ['Tag', 'Medewerker', 'Tot. uren', 'Nachtdienst?', 'Ber. nacht', 'Fac. nacht/50%',
               'Ber. OT (info)', 'Fac. OT (info)', 'Signaal', 'Opmerking']
        for i, h in enumerate(hdr, start=1):
            c = ws3.cell(row=4, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
        tag2toeslag = {t['tag']: t for t in toeslag_data.values()}
        rr = 5
        n_sig = 0
        for row in sorted(rijen, key=lambda x: int(x['tag']) if str(x['tag']).isdigit() else 99999):
            t = tag2toeslag.get(str(row['tag']))
            if not t:
                continue
            bernacht = t['nacht']
            fnacht = row['inv_nacht']
            ber_ot = round(t['c135'] + t['c150'], 2)
            fac_ot = round(row['inv_ot135'] + row['inv_ot150'], 2)
            # Hard signaal: nachtwerk volgens Nitea, maar geen nacht/50%-toeslag gefactureerd
            nachtdienst = t['is_nachtdienst'] or bernacht > 1.0
            signaal = nachtdienst and fnacht < 0.5
            if signaal:
                n_sig += 1
            opm = 'Nachtdienst gewerkt, geen nacht-/50%-toeslag op factuur!' if signaal else ''
            vals = [int(row['tag']) if str(row['tag']).isdigit() else row['tag'], row['naam'], t['totaal'],
                    'JA' if t['is_nachtdienst'] else '', bernacht, fnacht, ber_ot, fac_ot,
                    'CONTROLEER' if signaal else '', opm]
            fill = FILL_GEEL if signaal else FILL_OK
            for i, v in enumerate(vals, start=1):
                c = ws3.cell(row=rr, column=i, value=v); c.font = FONT_N; c.border = BORDER; c.fill = fill
                c.alignment = ALIGN_L if i in (2, 10) else (ALIGN_C if i in (1, 4, 9) else ALIGN_R)
            rr += 1
        ws3.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=10)
        ws3.cell(row=rr, column=1, value=f'Nacht-signalen (controleer): {n_sig}. OT-kolommen ter info — bureaus hanteren '
                 'eigen overwerk-/zaterdaglogica, geen exacte reproductie. Bij twijfel handmatig narekenen.').font = FONT_SUB
        ws3.cell(row=rr, column=1).fill = FILL_TOT
        ws3.freeze_panes = 'A5'
        for i, w in enumerate([6, 28, 9, 12, 11, 13, 13, 13, 11, 40]):
            ws3.column_dimensions[chr(65 + i)].width = w

    wb.save(out)
    ctx = {'te_hoog': te_hoog, 'te_laag': te_laag, 'naam_var': naam_var, 'credit': credit,
           'netto': netto, 'tot_dg': tot_dg, 'tot_fac': tot_fac, 'n_med': len(rijen)}
    return out, ctx


def genereer_mails(map_pad, prof, week, jaar, ctx, factuurnrs):
    wk = f'{int(week):02d}'
    te_hoog, te_laag, naam_var, credit = ctx['te_hoog'], ctx['te_laag'], ctx['naam_var'], ctx['credit']
    fnrs = ', '.join(factuurnrs)
    pi = os.path.join(map_pad, f'Mail_INTERN_{prof.code}_WK{wk}_{jaar}.txt')
    li = [f'Onderwerp: Factuur {prof.naam} week {wk} {jaar} — NOG NIET akkoord voor betaling (actie + besluit HR)', '',
          'Beste HR,', '',
          f'De controle van de facturen ({fnrs}) van {prof.naam} voor week {wk} is afgerond.', '',
          '=== BETAALSTATUS: NOG NIET AKKOORD VOOR BETALING ===',
          f'Uren kloppen ({ctx["tot_dg"]:.2f}h = {ctx["tot_fac"]:.2f}h, {ctx["n_med"]} medewerkers). '
          f'Wel ca. €{credit:.2f} te veel gefactureerd door onjuiste inschaling. Conform afspraak pas betalen na creditfactuur.', '',
          '=== A. NAAMFOUTEN — intern corrigeren in bronsysteem ===']
    if naam_var:
        for x in naam_var:
            li.append(f"  - tag {x['tag']}: Nitea \"{x['naam_dg']}\" vs factuur \"{x['naam']}\"")
    else:
        li.append('Geen naamfouten.')
    li += ['', '=== B. TARIEF-/UURAFWIJKINGEN — beoordelen: aan ons of aan het UZB? ===', 'Uren: geen afwijkingen.']
    if te_hoog or te_laag:
        for x in sorted(te_hoog + te_laag, key=lambda y: -y['euro_impact']):
            richting = 'TE VEEL gefactureerd' if x['euro_impact'] > 0 else 'te weinig gefactureerd'
            li.append(f"  - tag {x['tag']} {x['naam']} (schaal {x['schaal']}): effect €{x['euro_impact']:+.2f} ({richting}).")
            for bd in x['tarief_breakdown']:
                li.append(f"      {bd['uren']:.2f} uur {bd['categorie']}: €{bd['factuur_tarief']:.2f}/u i.p.v. €{bd['schaal_tarief']:.2f}/u "
                          f"= {bd['uren']:.2f} x €{bd['verschil']:+.2f} = €{bd['bedrag']:+.2f}")
            if x['euro_impact'] > 0:
                li.append(f"      MIJN ADVIES: hoort bij hogere schaal dan {x['schaal']} -> waarschijnlijk UZB te hoog, credit, tenzij heringedeeld.")
            else:
                li.append(f"      MIJN ADVIES: hoort bij lagere schaal dan {x['schaal']} (in ons voordeel). Controleer onze schaal.")
    else:
        li.append('Tarief: geen afwijkingen.')
    li += ['', '=== ADVIES & BESLUIT ===',
           'Mijn advies: factuur aanhouden, met UZB in overleg, creditfactuur afwachten. Definitief besluit aan HR.', '',
           'Groet,', 'Urencontrole']
    with open(pi, 'w', encoding='utf-8') as f:
        f.write('\n'.join(li))

    pe = os.path.join(map_pad, f'Mail_EXTERN_{prof.code}_WK{wk}_{jaar}.txt')
    le = [f'Onderwerp: Tariefcorrectie factuur week {wk} {jaar} — afwijkende inschaling', '',
          f'Beste {prof.naam},', '',
          f'Wij hebben de facturen voor week {wk} {jaar} ({fnrs}) gecontroleerd. De uren komen overeen met onze '
          f'doorgegeven uren ({ctx["tot_fac"]:.2f} uur) — daarvoor akkoord.', '',
          'Wel zien wij medewerkers met een tarief dat afwijkt van de overeengekomen schaalindeling:', '']
    if te_hoog:
        le.append('TE HOOG gefactureerd (graag crediteren):')
        for x in te_hoog:
            le.append(f"  - {x['naam']} (tag {x['tag']}), schaal {x['schaal']}:")
            for bd in x['tarief_breakdown']:
                le.append(f"      {bd['uren']:.2f} uur {bd['categorie']}: gefactureerd €{bd['factuur_tarief']:.2f}/u, "
                          f"correct €{bd['schaal_tarief']:.2f}/u -> {bd['uren']:.2f} x €{bd['verschil']:.2f} = €{bd['bedrag']:.2f} te veel")
            le.append(f"      Subtotaal te veel: €{x['euro_impact']:.2f}")
        le.append('')
    le += [f'Totaal te veel gefactureerd in week {wk}: €{credit:.2f} excl. BTW. Graag crediteren en de inschaling corrigeren.', '',
           'Mocht u van mening zijn dat de inschaling aan onze kant niet juist staat, dan kunt u zich melden bij onze HR-afdeling.', '',
           'Met vriendelijke groet,', 'A. Baas Pot- en Tuinplantenkwekerij']
    with open(pe, 'w', encoding='utf-8') as f:
        f.write('\n'.join(le))
    return pi, pe


# ---- register (gedeeld over bureaus) ----
KOLOMMEN = ['Reg.nr', 'Datum gemeld', 'Week', 'Leverancier', 'Tag', 'Medewerker', 'Type post',
            'Schaal', 'Effect (€)', 'Richting', 'Status', 'Verwacht in factuur', 'Datum afgehandeld', 'Opmerking']


def lees_register(pad):
    if not os.path.exists(pad):
        return []
    try:
        wb = load_workbook(pad, data_only=True)
    except Exception:
        return []
    if 'Register' not in wb.sheetnames:
        return []
    ws = wb['Register']
    hr = next((ri for ri in range(1, 8) if ws.cell(row=ri, column=1).value == 'Reg.nr'), None)
    if hr is None:
        return []
    out = []
    ri = hr + 1
    while ri < hr + 3000:
        med = ws.cell(row=ri, column=6).value
        regnr = ws.cell(row=ri, column=1).value
        if med is None and regnr is None:
            break
        if med and not str(med).strip().upper().startswith('TOTAAL'):
            out.append({k: ws.cell(row=ri, column=ci).value for ci, k in enumerate(KOLOMMEN, start=1)})
        ri += 1
    return out


def update_register(map_pad, alle):
    pad = os.path.join(map_pad, REGISTER_NAAM)
    bestaand = lees_register(pad)
    keys = {(str(r.get('Leverancier')), str(r.get('Week')), str(r.get('Tag')), str(r.get('Type post'))) for r in bestaand}
    toeg = 0
    for prof, week, jaar, rijen in alle:
        for r in rijen:
            if r['tarief_status'] != 'TARIEF WIJKT AF':
                continue
            tv = r['euro_impact'] > 0
            wk_label = f'{jaar}-{int(week):02d}'
            key = (prof.naam, wk_label, str(r['tag']), 'Tariefcorrectie')
            if key in keys:
                continue
            bestaand.append({'Datum gemeld': '', 'Week': wk_label, 'Leverancier': prof.naam,
                             'Tag': int(r['tag']) if str(r['tag']).isdigit() else r['tag'], 'Medewerker': r['naam'],
                             'Type post': 'Tariefcorrectie', 'Schaal': r['schaal'], 'Effect (€)': round(r['euro_impact'], 2),
                             'Richting': 'Te veel gefactureerd (credit verwacht)' if tv else 'Te weinig gefactureerd (UZB kan bijfactureren)',
                             'Status': 'Te verifiëren (HR)', 'Verwacht in factuur': '', 'Datum afgehandeld': '',
                             'Opmerking': 'Factuurtarief wijkt af van SNOOP-inschaling. Onzeker of fout bij ons of UZB; HR beslist vóór credit/betaling.'})
            keys.add(key)
            toeg += 1
    _schrijf_register(pad, bestaand)
    return pad, toeg, len(bestaand)


def _schrijf_register(pad, regels):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Register'
    ws.merge_cells('A1:N1')
    ws['A1'] = 'Openstaande-posten-register — uitzendbureaus'
    ws['A1'].font = FONT_TITLE; ws['A1'].fill = FILL_TITLE; ws['A1'].alignment = ALIGN_C
    ws.merge_cells('A2:N2')
    ws['A2'] = ('Bewaakt verwachte correcties per uitzendbureau over weken heen. Bij nieuwe factuur controleren of open '
                'posten zijn verrekend; zo ja -> Status "Ontvangen" + datum (regel wordt groen).')
    ws['A2'].font = Font(name='Arial', size=9, italic=True); ws['A2'].alignment = ALIGN_L
    def _is_open(r):
        return str(r.get('Status', '')).lower().startswith('open')

    def _is_verif(r):
        return str(r.get('Status', '')).lower().startswith('te verif')
    n_open = sum(1 for r in regels if _is_open(r))
    n_verif = sum(1 for r in regels if _is_verif(r))
    oc = round(sum((r['Effect (€)'] or 0) for r in regels if _is_open(r) and (r['Effect (€)'] or 0) > 0), 2)
    vc = round(sum((r['Effect (€)'] or 0) for r in regels if _is_verif(r)), 2)
    ws['A3'] = f'Open posten: {n_open}'; ws['A3'].font = FONT_SUB
    ws['C3'] = f'Open credit verwacht: € {oc:.2f}'; ws['C3'].font = FONT_SUB
    ws['F3'] = f'Te verifiëren door HR: {n_verif} post(en), € {vc:.2f}'; ws['F3'].font = FONT_SUB
    hr = 5
    for ci, k in enumerate(KOLOMMEN, start=1):
        c = ws.cell(row=hr, column=ci, value=k); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
    rr = hr + 1
    for i, r in enumerate(sorted(regels, key=lambda x: (str(x.get('Leverancier', '')), str(x.get('Week', '')), -(x.get('Effect (€)') or 0))), start=1):
        r['Reg.nr'] = i
        _sl = str(r.get('Status', '')).lower()
        f = FILL_GEEL if _sl.startswith('open') else (FILL_BLAUW if _sl.startswith('te verif') else FILL_OK)
        for ci, k in enumerate(KOLOMMEN, start=1):
            c = ws.cell(row=rr, column=ci, value=r.get(k, '')); c.font = FONT_N; c.border = BORDER; c.fill = f
            c.alignment = ALIGN_R if k == 'Effect (€)' else (ALIGN_C if k in ('Reg.nr', 'Week', 'Tag', 'Status') else ALIGN_L)
            if k == 'Effect (€)':
                c.number_format = '€ #,##0.00;-€ #,##0.00'
        rr += 1
    for i, w in enumerate([7, 13, 9, 26, 6, 26, 16, 8, 12, 30, 12, 18, 14, 40]):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = 'A6'
    wb.save(pad)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--map', default=DEFAULT_MAP)
    args = ap.parse_args()
    map_pad = args.map
    print(f'Map: {map_pad}')
    print(f'Actieve bureau-profielen: {[b.naam for b in bp.BUREAUS]}')

    per_bw, factuur_info, onbekend, credits = lees_facturen(map_pad)
    if not factuur_info:
        print('GEEN herkende uitzendbureau-facturen gevonden.')
        if onbekend:
            print('Wel mogelijk relevante facturen ZONDER profiel (voeg een bureau-profiel toe):')
            for b in onbekend:
                print(f'  - {b}')
        return
    print('\nGevonden facturen:')
    for fi in factuur_info:
        print(f"  [{fi['code']}] {fi['bestand']} -> weken {fi['weken']}")
    if onbekend:
        print('\nLET OP — facturen zonder bureau-profiel (overgeslagen):')
        for b in onbekend:
            print(f'  - {b}')

    # SNOOP-inschaling (1x) + tariefkaart per bureau (1x)
    import snoop as snoopmod
    snoop_files = glob.glob(os.path.join(map_pad, '*snoop*.xlsx')) + glob.glob(os.path.join(map_pad, '*SNOOP*.xlsx'))
    snoop_data = {}
    if snoop_files:
        # Meest recent bewerkte SNOOP eerst — zodat een nieuwere export automatisch gepakt wordt.
        snoop_files = sorted(set(snoop_files), key=os.path.getmtime, reverse=True)
        snoop_data = snoopmod.laad_snoop(snoop_files[0])
        print(f'SNOOP geladen: {len(snoop_data)} medewerkers (excl. Kwekerij Baas/Temper) uit {os.path.basename(snoop_files[0])}')
    else:
        print('LET OP: geen SNOOP-bestand gevonden -> alle tarieven als UZB-tarief.')
    tariefkaart_cache = {prof.code: prof.tariefkaart() for prof in bp.BUREAUS}
    # HR/finance-werkbestanden: per medewerker schaal+tarief (vult ontbrekende SNOOP-inschalingen)
    factuurkaart = laad_factuurkaart_werknemers()

    alle = []
    for (code, week), fac_per_naam in sorted(per_bw.items()):
        prof = next(b for b in bp.BUREAUS if b.code == code)
        dg = vind_doorgegeven(map_pad, week, prof.doorgegeven_hint)
        if dg is None:
            print(f'\n!! [{code}] week {week}: facturen WEL, doorgegeven-uren-bestand NIET gevonden — overslaan.')
            continue
        dg_pad, dg_data, dg_jaar = dg
        jaar = dg_jaar or '2026'
        rijen, niet_gefac = vergelijk(prof, week, jaar, dg_data, fac_per_naam, snoop_data,
                                      tariefkaart_cache[code], factuurkaart.get(code))
        # Toeslag: Nitea-overzicht-PDF voor (week, bureau) zoeken
        toeslag_data = None
        hints_t = [prof.doorgegeven_hint] if isinstance(prof.doorgegeven_hint, str) else list(prof.doorgegeven_hint)
        wkz = f'{int(week):02d}'
        for npad in glob.glob(os.path.join(map_pad, '*itea*verzicht*.pdf')) + glob.glob(os.path.join(map_pad, '*itea*fronding*.pdf')):
            nb = os.path.basename(npad).lower()
            if re.search(rf'wk\s?0*{int(week)}\b', nb) and any(h.lower() in nb for h in hints_t):
                try:
                    import toeslag as toeslagmod
                    toeslag_data = toeslagmod.toeslag_per_medewerker(npad)
                except Exception as ex:
                    print(f'   (toeslag-PDF {os.path.basename(npad)} niet leesbaar: {ex})')
                break
        out, ctx = genereer_rapportage(map_pad, prof, week, jaar, dg_data, rijen, toeslag_data, niet_gefac)
        fnrs = sorted({r.factuurnr for r in fac_per_naam.values()})
        pi, pe = genereer_mails(map_pad, prof, week, jaar, ctx, fnrs)
        alle.append((prof, week, jaar, rijen))
        afw = sum(1 for r in rijen if r['tarief_status'] == 'TARIEF WIJKT AF')
        print(f'\n=== [{code}] {prof.naam} — WEEK {week} {jaar} ===')
        print(f'  Medewerkers {len(rijen)} | {ctx["tot_dg"]:.2f}h doorgegeven = {ctx["tot_fac"]:.2f}h gefactureerd '
              f'({ctx["tot_fac"]-ctx["tot_dg"]:+.2f}h)')
        print(f'  Tariefafwijkingen {afw} | credit €{ctx["credit"]:.2f} | netto €{ctx["netto"]:.2f}')
        print(f'  -> {os.path.basename(out)} | {os.path.basename(pi)} | {os.path.basename(pe)}')

    if alle:
        reg, toeg, tot = update_register(map_pad, alle)
        print(f'\nRegister: {os.path.basename(reg)} (+{toeg} nieuw, {tot} totaal)')

    # ---- Creditfacturen: afletteren tegen het register + per-medewerker tarief-aansluiting ----
    if credits:
        verwerk_credits(map_pad, credits, snoop_data, tariefkaart_cache, factuurkaart)

    # ---- Kopieerbaar tekstoverzicht per UZB (afwijkingen per factuur, excl. reeds gecrediteerd) ----
    if alle:
        genereer_mailoverzicht(map_pad, alle, credits)

    # ---- Werklijst: per factuur, per afwijking uitgesplitst + besluit-kolom ----
    if alle:
        datum_map = {fi['factuurnr']: fi.get('factuurdatum') for fi in factuur_info}
        genereer_afwijkingen_werklijst(map_pad, alle, credits, datum_map, factuurkaart,
                                       snoop_data=snoop_data, tariefkaart_cache=tariefkaart_cache)


def _afwijking_oordeel(diff_eur, soort, niet_doorgegeven=False):
    """(voordeel-tekst, celkleur, voorstel-tekst) op basis van richting van het verschil.
    Conventie (afspraak Dieter): + = wij betaalden TE VEEL (rood, terugvorderen, voordeel Baas);
    - = wij betaalden TE WEINIG (geel, lage prio, UZB kan bijfactureren)."""
    if diff_eur > 0.01:
        voordeel = 'Baas — wij betaalden te veel (terugvorderen)'
        fill = FILL_ROOD
        if soort == 'tarief':
            voorstel = 'Credit te ontvangen — eerst HR (tarief vs inschaling)'
        elif niet_doorgegeven:
            voorstel = 'Credit te ontvangen — gefactureerd zonder doorgegeven uren'
        else:
            voorstel = 'Credit te ontvangen — te veel uren gefactureerd'
    elif diff_eur < -0.01:
        voordeel = 'UZB — wij betaalden te weinig (UZB kan bijfactureren)'
        fill = FILL_GEEL
        voorstel = ('Mogelijk bijfacturen UZB — eerst HR (tarief vs inschaling)' if soort == 'tarief'
                    else 'Lage prio — te weinig gefactureerd (voordeel Baas)')
    else:
        voordeel = '—'
        fill = FILL_OK
        voorstel = 'Geen materieel verschil'
    return voordeel, fill, voorstel


def genereer_afwijkingen_werklijst(map_pad, alle, credits, datum_map=None, factuurkaart=None,
                                   snoop_data=None, tariefkaart_cache=None):
    """Eén werklijst (Afwijkingen_werklijst.xlsx): per factuur, één regel per afwijking, volledig
    uitgesplitst (gefactureerd vs verwacht, verschil, in wiens voordeel) + een Besluit-dropdown
    zodat per afwijking bepaald kan worden of er gecrediteerd moet worden (positief/negatief) of niet."""
    from openpyxl.worksheet.datavalidation import DataValidation
    datum_map = datum_map or {}
    credited = set()
    for cr in credits:
        for w in cr['weken']:
            credited.add((cr['code'], str(int(w)) if str(w).isdigit() else str(w)))

    wb = Workbook(); ws = wb.active; ws.title = 'Afwijkingen'
    ws.merge_cells('A1:T1')
    ws['A1'] = 'Afwijkingen-werklijst — per factuur, per afwijking beslissen'
    ws['A1'].font = FONT_TITLE; ws['A1'].fill = FILL_TITLE; ws['A1'].alignment = ALIGN_C
    ws.merge_cells('A2:T2')
    ws['A2'] = ('Eén regel per afwijking. "Gefactureerd" = wat het UZB rekende; "Verwacht" = wat volgens onze '
                'inschaling (tarief) of doorgegeven uren hoorde. Verschil € POSITIEF = wij betaalden TE VEEL '
                '(terugvorderen — voordeel Baas, ROOD); NEGATIEF = wij betaalden TE WEINIG (UZB kan bijfactureren, GEEL). '
                'Tarief-afwijkingen zijn t.o.v. SNOOP-inschaling → eerst HR bevestigen. Vul per regel "Besluit" in (dropdown).')
    ws['A2'].font = Font(name='Arial', size=9, italic=True); ws['A2'].alignment = ALIGN_L; ws.row_dimensions[2].height = 60
    hdr = ['Bureau', 'Week', 'Factuur', 'Factuurdatum', 'Tag', 'Medewerker', 'Soort', 'Categorie',
           'Gefact. uren', 'Gefact. tarief', 'Gefact. bedrag',
           'Verwacht uren', 'Verwacht tarief', 'Verwacht bedrag',
           'Verschil €', 'Voordeel voor', 'Al gecrediteerd?', 'Voorstel', 'Besluit', 'Opmerking']
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=4, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER

    rr = 5
    n_rood = n_geel = 0
    tot_teveel = tot_teweinig = 0.0
    for prof, week, jaar, rijen in sorted(alle, key=lambda x: (x[0].code, int(x[1]) if str(x[1]).isdigit() else 999)):
        wk = str(int(week)) if str(week).isdigit() else str(week)
        al_cred = 'Ja' if (prof.code, wk) in credited else 'Nee'
        for r in sorted(rijen, key=lambda x: str(x.get('naam', ''))):
            afw_rows = []
            # 1) tarief-afwijkingen (per categorie, t.o.v. inschaling)
            for bd in r.get('tarief_breakdown', []):
                uren = bd['uren']; ft = bd['factuur_tarief']; st = bd['schaal_tarief']
                gfb = round(uren * ft, 2); vwb = round(uren * st, 2); diff = round(gfb - vwb, 2)
                afw_rows.append(('Tarief', bd['categorie'], round(uren, 2), round(ft, 3), gfb,
                                 round(uren, 2), round(st, 3), vwb, diff, 'tarief', False))
            # 2) uren-afwijking (gefactureerd vs doorgegeven)
            if abs(r.get('verschil', 0.0)) > TOLERANTIE_UREN:
                tarief = r.get('tarief')
                fac_u = r.get('fac_uren', 0.0); dg_u = r.get('dg_uren', 0.0)
                niet_dg = 'NIET DOORGEGEVEN' in r.get('status', '')
                if isinstance(tarief, (int, float)):
                    gfb = round(fac_u * tarief, 2); vwb = round(dg_u * tarief, 2); diff = round(gfb - vwb, 2)
                    tw = round(tarief, 3)
                else:
                    gfb = vwb = diff = None; tw = ''
                afw_rows.append(('Uren', 'gewerkte uren (totaal)', round(fac_u, 2), tw, gfb,
                                 round(dg_u, 2), tw, vwb, diff, 'uren', niet_dg))
            for soort, cat, gu, gt, gb, vu, vt, vb, diff, skind, niet_dg in afw_rows:
                voordeel, fill, voorstel = _afwijking_oordeel(diff if diff is not None else 0.0, skind, niet_dg)
                if diff is not None and diff > 0.01:
                    n_rood += 1; tot_teveel += diff
                elif diff is not None and diff < -0.01:
                    n_geel += 1; tot_teweinig += diff
                datum = _datum_str(datum_map.get(r.get('factuurnr')))
                opm = (f'verwacht tarief uit {r.get("tarief_bron")}' if soort == 'Tarief' and r.get('tarief_bron') else '')
                vals = [prof.code, wk, r.get('factuurnr', ''), datum, r.get('tag', ''), r.get('naam', ''),
                        soort, cat, gu, gt, gb, vu, vt, vb, diff, voordeel, al_cred, voorstel, '', opm]
                for i, v in enumerate(vals, start=1):
                    c = ws.cell(row=rr, column=i, value=v); c.font = FONT_N; c.border = BORDER; c.fill = fill
                    c.alignment = ALIGN_R if i in (9, 10, 11, 12, 13, 14, 15) else ALIGN_C if i == 4 else ALIGN_L
                    if i in (10, 13):
                        c.number_format = '€ #,##0.000'
                    if i in (11, 14, 15):
                        c.number_format = '€ #,##0.00;-€ #,##0.00'
                rr += 1

    if rr == 5:
        ws.cell(row=5, column=1, value='Geen afwijkingen aangetroffen in de gecontroleerde facturen.')
    else:
        # Besluit-dropdown op kolom S (19), autofilter + bevriezing
        dv = DataValidation(type='list',
                            formula1='"Crediteren (terugvorderen),Crediteren (bijbetalen),Niet crediteren,Te verifieren HR,Geen actie"',
                            allow_blank=True)
        dv.add(f'S5:S{rr - 1}')
        ws.add_data_validation(dv)
        ws.auto_filter.ref = f'A4:T{rr - 1}'
        ws.freeze_panes = 'A5'
    for i, w in enumerate([7, 6, 12, 13, 7, 24, 8, 22, 11, 13, 14, 12, 13, 14, 13, 40, 14, 46, 26, 30]):
        ws.column_dimensions[chr(65 + i)].width = w

    # ── Tabblad 2: schaal-vergelijking SNOOP vs Werkbestand vs UZB, per uitzendkracht ──
    # Voor elke unieke (code, tag, naam) uit de weekcontroles: vergelijk de schaal volgens
    # SNOOP, het HR-werkbestand (Factuurcalculatie) en het UZB (omgekeerd opgezocht uit het
    # gefactureerde normaal-tarief in de tariefkaart). 'JA' als er actie nodig is, anders 'NEE'.
    snoop_data = snoop_data or {}
    tariefkaart_cache = tariefkaart_cache or {}
    base_key = {'L1': 't_100_135', 'LP': 't_100_135', 'SW': 't_100_135', 'CK': 't_100'}

    def _schaal_uzb(tarief, code):
        """Reverse-lookup: welke schaal hoort bij dit (normaal-)tarief volgens de tariefkaart?"""
        tk = tariefkaart_cache.get(code) or {}
        if tarief is None or not tk:
            return ''
        key = base_key.get(code, 't_100_135')
        tr = round(float(tarief), 2)
        for schaal, rec in tk.items():
            v = rec.get(key)
            if v is not None and round(float(v), 2) == tr:
                return schaal
        return ''

    # Per (code, tag, naam, schaal_snoop, schaal_wb): aparte regel — zodat wisselaars
    # (iemand die halverwege een andere schaal krijgt) als twee regels verschijnen, elk met
    # hun eigen weken. Tarief wordt per combo geaggregeerd (modus = meest voorkomende).
    uzk = {}
    for prof, week, jaar, rijen in alle:
        for r in rijen:
            sn = r.get('schaal_snoop_raw') or ''
            wb_s = r.get('schaal_werkbestand_raw') or ''
            k = (prof.code, str(r.get('tag', '')), r.get('naam', ''), sn, wb_s)
            rec = uzk.setdefault(k, {'tariefs': Counter(), 'weken': set()})
            t = r.get('tarief')
            if isinstance(t, (int, float)):
                rec['tariefs'][round(float(t), 3)] += 1
            rec['weken'].add(str(week))

    ws2 = wb.create_sheet(title='UZK zonder SNOOP')
    ws2.merge_cells('A1:J1')
    ws2['A1'] = 'Inschaling-controle per UZK — SNOOP vs Werkbestand vs UZB'
    ws2['A1'].font = FONT_TITLE; ws2['A1'].fill = FILL_TITLE; ws2['A1'].alignment = ALIGN_C
    ws2.merge_cells('A2:J2')
    ws2['A2'] = ('Per medewerker drie schaal-bronnen naast elkaar: SNOOP, HR-werkbestand (Factuurcalculatie) en de '
                 '"UZB-schaal" = de schaal die hoort bij het gefactureerde normaal-tarief (omgekeerd opgezocht via de '
                 'tariefkaart). "Aanpassing vereist" = JA zodra er iets afwijkt of ontbreekt; NEE als alle drie gelijk zijn. '
                 'De actie-kolom geeft aan WAT er moet gebeuren: SNOOP bijwerken / overleg met UZB / Factuurcalculatie aanvullen.')
    ws2['A2'].font = Font(name='Arial', size=9, italic=True); ws2['A2'].alignment = ALIGN_L; ws2.row_dimensions[2].height = 60
    hdr2 = ['Naam', 'UZB', 'Tag', 'Schaal SNOOP', 'Schaal Werkbestand', 'Schaal UZB',
            'aanpassing vereist (Werkbestand=UZB)', None, 'Actie', 'Weken gezien']
    for i, h in enumerate(hdr2, start=1):
        c = ws2.cell(row=4, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = ALIGN_C; c.border = BORDER

    gr = 5
    n_ja = n_nee = n_3way = 0
    for (code, tag, naam, s_snoop, s_wb), info in sorted(uzk.items(), key=lambda kv: (kv[0][0], kv[0][2], kv[0][3], kv[0][4])):
        # tarief (modus over alle weken in deze combo) -> schaal UZB via reverse-lookup
        t_uzb = info['tariefs'].most_common(1)[0][0] if info['tariefs'] else None
        s_uzb = _schaal_uzb(t_uzb, code)
        # Aanpassing vereist + actie
        sn = s_snoop or None; wb_s = s_wb or None; uz = s_uzb or None
        if sn and wb_s and uz and sn == wb_s == uz:
            ja_nee, actie, kl = 'NEE', 'correct', FILL_OK
        else:
            ja_nee = 'JA'
            # Volgorde van prioriteit: 0) alle drie bekend én alle drie verschillend (driewegmismatch);
            # 1) W != UZB (UZB-discrepantie = hardste signaal), 2) SNOOP ontbreekt,
            # 3) W = UZB maar SNOOP wijkt af, 4) W ontbreekt, 5) edge.
            if sn and wb_s and uz and sn != wb_s and sn != uz and wb_s != uz:
                actie = f'DRIEWEGMISMATCH — alle bronnen verschillen (SNOOP {sn} / werkbestand {wb_s} / UZB {uz}); uitzoeken'
                kl = FILL_ROOD
            elif wb_s and uz and wb_s != uz:
                extra = f' (SNOOP ook leeg)' if not sn else (f' (SNOOP zegt {sn})' if sn != wb_s and sn != uz else '')
                actie = f'overleg met UZB — werkbestand {wb_s}, UZB-tarief past bij {uz}{extra}'
                kl = FILL_ROOD
            elif not sn and wb_s and uz and wb_s == uz:
                actie = f'vul SNOOP-inschaling aan ({wb_s})'
                kl = FILL_GEEL
            elif not sn and wb_s:
                actie = f'vul SNOOP-inschaling aan ({wb_s} volgens werkbestand)'
                kl = FILL_GEEL
            elif not sn and uz and not wb_s:
                actie = f'vul SNOOP én Werkbestand aan ({uz} volgens UZB-tarief)'
                kl = FILL_BLAUW
            elif not sn and not wb_s and not uz:
                actie = 'overleg HR — geen enkele bron heeft een inschaling'
                kl = FILL_BLAUW
            elif wb_s and uz and wb_s == uz and sn != wb_s:
                actie = f'pas SNOOP-schaal aan ({wb_s}; nu staat {sn})'
                kl = FILL_ROOD
            elif sn and uz and not wb_s and sn == uz:
                actie = f'vul HR-werkbestand aan ({sn} volgens SNOOP/UZB)'
                kl = FILL_GEEL
            elif not wb_s:
                actie = 'vul HR-werkbestand aan'
                kl = FILL_GEEL
            else:
                actie = 'controleren — handmatig'
                kl = FILL_GEEL
        if ja_nee == 'JA':
            n_ja += 1
            if actie.startswith('DRIEWEGMISMATCH'):
                n_3way += 1
        else:
            n_nee += 1
        tag_disp = '' if tag in ('', '?', None) else tag
        wkn = ', '.join(sorted(info['weken'], key=lambda x: int(x) if str(x).isdigit() else 999))
        vals = [naam, code, tag_disp, s_snoop, s_wb, s_uzb, ja_nee, None, actie, wkn]
        for i, v in enumerate(vals, start=1):
            c = ws2.cell(row=gr, column=i, value=v); c.font = FONT_N; c.border = BORDER; c.fill = kl
            c.alignment = ALIGN_C if i in (2, 3, 4, 5, 6, 7) else ALIGN_L
        gr += 1
    if not uzk:
        ws2.cell(row=5, column=1, value='Geen medewerkers gevonden op gecontroleerde facturen.')
    else:
        ws2.auto_filter.ref = f'A4:J{gr - 1}'
        ws2.freeze_panes = 'A5'
    for i, w in enumerate([28, 7, 8, 14, 18, 14, 30, 4, 50, 28]):
        ws2.column_dimensions[chr(65 + i)].width = w
    print(f'   Inschaling-controle: {len(uzk)} UZK ({n_ja} aanpassing vereist, waarvan {n_3way} driewegmismatch; '
          f'{n_nee} correct) -> tabblad "UZK zonder SNOOP"')

    # ── Tabblad 3: DAG-bewuste tag-conflicten ──
    # Tag is uniek per DAG (Nitea hergebruikt tags over dagen/weken — normaal). Echt conflict =
    # zelfde tag, ZELFDE DAG, 2 verschillende personen (mag nooit voorkomen).
    occ = bouw_dag_occupatie(map_pad)
    conflicten = []
    for (tag, jaar, week, d), namen in occ.items():
        if len(namen) > 1:
            dag = DAGNAMEN[d] if 0 <= d < len(DAGNAMEN) else f'd{d}'
            conflicten.append((tag, jaar, week, dag, sorted(namen)))
    conflicten.sort(key=lambda x: (str(x[0]), x[1] or '', x[2] or 0))
    ws3 = wb.create_sheet(title='Tag-conflicten')
    ws3.merge_cells('A1:E1')
    ws3['A1'] = 'Tag-conflicten — zelfde tag op dezelfde dag bij meerdere personen'
    ws3['A1'].font = FONT_TITLE; ws3['A1'].fill = FILL_TITLE; ws3['A1'].alignment = ALIGN_C
    ws3.merge_cells('A2:E2')
    ws3['A2'] = ('Tagnummers worden in Nitea HERGEBRUIKT en zijn uniek per DAG (niet over tijd): dezelfde tag bij '
                 'meerdere personen op verschillende dagen/weken is NORMAAL. Hieronder alleen ECHTE conflicten: '
                 'zelfde tag, dezelfde dag, twee personen — dat hoort nooit en duidt op een registratiefout. '
                 f'Gecontroleerd over {len(occ)} dag-registraties uit de doorgegeven-bestanden.')
    ws3['A2'].font = Font(name='Arial', size=9, italic=True); ws3['A2'].alignment = ALIGN_L
    ws3.row_dimensions[2].height = 44
    for i, h in enumerate(['Tag', 'Jaar', 'Week', 'Dag', 'Personen (zelfde dag)'], start=1):
        c = ws3.cell(row=4, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
    if conflicten:
        for ri, (tag, jaar, week, dag, namen) in enumerate(conflicten, start=5):
            vals = [tag, jaar, week, dag, ' | '.join(namen)]
            for i, v in enumerate(vals, start=1):
                c = ws3.cell(row=ri, column=i, value=v); c.font = FONT_N; c.border = BORDER; c.fill = FILL_ROOD
                c.alignment = ALIGN_C if i in (1, 2, 3, 4) else ALIGN_L
        ws3.auto_filter.ref = f'A4:E{4 + len(conflicten)}'
    else:
        c = ws3.cell(row=5, column=1, value='Geen tag-conflicten gevonden — elke tag is per dag uniek. ✔')
        c.font = FONT_N; c.fill = FILL_OK
        ws3.merge_cells('A5:E5')
    ws3.freeze_panes = 'A5'
    for i, w in enumerate([10, 8, 8, 8, 60]):
        ws3.column_dimensions[chr(65 + i)].width = w
    print(f'   Tag-conflicten (zelfde dag): {len(conflicten)} -> tabblad "Tag-conflicten"')

    out = os.path.join(map_pad, 'Afwijkingen_werklijst.xlsx')
    try:
        wb.save(out)
        print(f'-> {os.path.basename(out)}: {rr - 5} afwijkingen ({n_rood} te veel €{tot_teveel:,.2f} / '
              f'{n_geel} te weinig €{tot_teweinig:,.2f})')
    except PermissionError:
        wb.save(out.replace('.xlsx', '_v2.xlsx'))
        print(f'-> Afwijkingen_werklijst_v2.xlsx (origineel stond open)')


def genereer_mailoverzicht(map_pad, alle, credits):
    """Schrijft per UZB een platte-tekst-bestand met alle afwijkingen per factuur,
    behalve weken waarvoor al een creditnota is ontvangen. Klaar om te kopiëren/mailen."""
    # weken waarvoor al een credit/correctie is binnen (per bureau-code)
    credited = set()
    for cr in credits:
        for w in cr['weken']:
            credited.add((cr['code'], str(int(w)) if str(w).isdigit() else str(w)))
    # groepeer per bureau
    per_bureau = {}
    for prof, week, jaar, rijen in alle:
        per_bureau.setdefault(prof, []).append((week, jaar, rijen))

    geschreven = []
    for prof, weken in per_bureau.items():
        lines = [f'OVERZICHT AFWIJKINGEN — {prof.naam}',
                 'Facturen waarvoor al een credit/correctie is ontvangen, zijn weggelaten.',
                 'Tarief-afwijkingen zijn t.o.v. SNOOP-inschaling — eerst intern (HR) bevestigen.',
                 '=' * 70, '']
        bureau_tot = 0.0
        iets = False
        for week, jaar, rijen in sorted(weken, key=lambda x: int(x[0]) if str(x[0]).isdigit() else 999):
            wk = str(int(week)) if str(week).isdigit() else str(week)
            if (prof.code, wk) in credited:
                lines.append(f'Week {wk}: credit/correctie reeds ontvangen — overgeslagen.')
                lines.append('')
                continue
            uren_afw = [r for r in rijen if 'UREN-AFWIJKING' in r['status'] or 'NIET DOORGEGEVEN' in r['status']]
            tar_afw = [r for r in rijen if r['tarief_status'] == 'TARIEF WIJKT AF']
            if not uren_afw and not tar_afw:
                continue
            iets = True
            fnrs = ', '.join(sorted({r['factuurnr'] for r in (uren_afw + tar_afw)}))
            lines.append(f'WEEK {wk} {jaar}  (factuur {fnrs})')
            if uren_afw:
                lines.append('  Uren:')
                for r in uren_afw:
                    lines.append(f"    - {r['naam']} (tag {r['tag']}): doorgegeven {r['dg_uren']}h, "
                                 f"gefactureerd {r['fac_uren']}h ({r['verschil']:+.2f}h)")
            if tar_afw:
                lines.append('  Tarief:')
                for r in sorted(tar_afw, key=lambda x: -x['euro_impact']):
                    bureau_tot += r['euro_impact']
                    det = '; '.join(f"{bd['uren']:.2f}u {bd['categorie']} €{bd['factuur_tarief']:.2f} i.p.v. €{bd['schaal_tarief']:.2f}"
                                    for bd in r['tarief_breakdown'])
                    ri = 'te veel' if r['euro_impact'] > 0 else 'te weinig'
                    lines.append(f"    - {r['naam']} (tag {r['tag']}, schaal {r['schaal']}): {det} "
                                 f"= €{r['euro_impact']:+.2f} ({ri})")
            lines.append('')
        if not iets:
            lines.append('Geen openstaande afwijkingen.')
            lines.append('')
        lines.append('-' * 70)
        lines.append(f'Totaal tarief-effect (excl. reeds gecrediteerd): €{round(bureau_tot, 2):+.2f}')
        lines.append('(positief = mogelijk te veel gefactureerd / credit aanvragen)')
        out = os.path.join(map_pad, f'Overzicht_afwijkingen_{prof.code}.txt')
        with open(out, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        geschreven.append(os.path.basename(out))
    print(f'\nKopieerbare overzichten per UZB: {", ".join(geschreven)}')


# Per-bureau: factuurcategorie -> kolom in de schaal-tariefkaart (zelfde mapping als tarief_checks)
CAT_SCHAALKEY = {
    'L1': {'normaal': 't_100_135', 'ow135': 't_100_135', 'ow150': 't_150', 'bijzonder150': 't_bijzonder'},
    'LP': {'normaal': 't_100_135', 'ow135': 't_100_135', 'ow150': 't_150', 'bijzonder150': 't_bijzonder'},
    'SW': {'normaal': 't_100_135', 'ow135': 't_100_135', 'ow150': 't_150', 'nacht50': 't_nacht50'},
    'CK': {'normaal': 't_100', 'ow135': 't_135', 'ow150': 't_150'},
}
CAT_LABEL = {'normaal': 'normale uren', 'ow135': 'OT 135%', 'ow150': 'OT 150%',
             'bijzonder150': 'bijzondere uren 150%', 'nacht50': '50% nacht'}


def _afletter_medewerkers(credits, snoop_data, tariefkaart_cache, factuurkaart=None):
    """Per medewerker-regel op elke credit/correctiefactuur: vergelijk het (nieuwe) tarief
    met de SNOOP-inschaling (of, als die ontbreekt, met de HR-factuurkaart) en geef een oordeel.
    Doel: bepalen of iedereen die bij ons werkte conform tarief betaald wordt en of de
    correcties de eerder gevlagde afwijkingen écht oplossen."""
    snoop_data = snoop_data or {}
    tariefkaart_cache = tariefkaart_cache or {}
    factuurkaart = factuurkaart or {}
    snoop_namen = list(snoop_data.keys())

    def snoop_lookup(nn):
        if nn in snoop_data:
            return snoop_data[nn]
        k = difflib.get_close_matches(nn, snoop_namen, n=1, cutoff=0.88)
        return snoop_data[k[0]] if k else None

    rijen = []
    for cr in credits:
        code = cr['code']
        tk = tariefkaart_cache.get(code, {})
        catmap = CAT_SCHAALKEY.get(code, {})
        for regel in cr.get('regels', []):
            nn = regel.naam_norm
            wklabel = f'{regel.jaar}-{regel.week}' if getattr(regel, 'jaar', None) else str(regel.week)
            srec = snoop_lookup(nn)
            schaal = snoop_schaal_voor_week(srec, regel.jaar, regel.week) if srec else None
            sch = tk.get(schaal) if schaal else None
            if sch is None:  # fallback naar HR-factuurkaart (op naam; credits hebben geen tag)
                frec = _fk_lookup(factuurkaart.get(code), None, nn)
                if frec:
                    sch = frec
                    schaal = schaal or frec.get('schaal')
            for cat, info in regel.categorieen.items():
                ntar = info.get('tarief')
                if ntar is None:
                    continue
                netto_uren = round(info.get('uren', 0.0), 2)
                key = catmap.get(cat)
                star = sch.get(key) if (sch and key) else None
                if schaal is None:
                    oordeel, kl, dlt = 'Geen inschaling in SNOOP — verifiëren (HR)', 'blauw', None
                elif star is None:
                    oordeel, kl, dlt = f'Geen schaaltarief voor {schaal}/{cat} — verifiëren (HR)', 'blauw', None
                else:
                    dlt = round(round(float(ntar), 2) - round(float(star), 2), 2)
                    if tarief_match(ntar, star):
                        oordeel, kl = '✅ Tarief klopt nu met inschaling', 'groen'
                    else:
                        oordeel, kl = f'⚠️ Wijkt af van inschaling ({dlt:+.3f}/u) — noteren', 'rood'
                rijen.append({
                    'factuur': cr['bestand'], 'code': code, 'bureau': cr['bureau'],
                    'week': wklabel, 'naam': regel.naam_factuur, 'schaal': schaal or '?',
                    'categorie': CAT_LABEL.get(cat, cat), 'netto_uren': netto_uren,
                    'nieuw_tarief': round(float(ntar), 2),
                    'inschaling_tarief': round(float(star), 2) if star is not None else None,
                    'delta': dlt, 'oordeel': oordeel, 'kleur': kl,
                })
    return rijen


CREDIT_HDR = ['Factuur', 'Factuurdatum', 'Type', 'Bureau', 'Week(en)', 'Hoofdweek', 'Netto uren',
              'Saldo (excl. BTW)', 'Richting', 'Omschrijving / waar', 'Aansluiting register / actie']


def _datum_str(d):
    return d.strftime('%d-%m-%Y') if isinstance(d, datetime.date) else ''


def _schrijf_credit_sheet(ws, titel, credits_subset, open_posten):
    """Vult één werkblad met de credit-/correctieregels, gesorteerd op factuurdatum (nieuwste boven).
    Retourneert (saldo_ontvangen, saldo_betalen)."""
    ws.merge_cells('A1:K1')
    ws['A1'] = titel
    ws['A1'].font = FONT_TITLE; ws['A1'].fill = FILL_TITLE; ws['A1'].alignment = ALIGN_C
    ws.merge_cells('A2:K2')
    ws['A2'] = ('Gesorteerd op factuurdatum (nieuwste boven). Hoofdweek = de week die voor de meeste '
                'uitzendkrachten op de factuur staat. Credit-/correctie-/tariefcorrectie-facturen tellen NIET mee '
                'in de gewone weekcontrole. Saldo: NEGATIEF = wij ontvangen terug, POSITIEF = wij betalen bij. '
                'Tariefcorrectie = zelfde uren tegen oud (–) én nieuw (+) tarief, uren netto 0.')
    ws['A2'].font = Font(name='Arial', size=9, italic=True); ws['A2'].alignment = ALIGN_L
    ws.row_dimensions[2].height = 52
    for i, h in enumerate(CREDIT_HDR, start=1):
        c = ws.cell(row=4, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
    rr = 5
    saldo_ontv = saldo_bet = 0.0
    gesorteerd = sorted(credits_subset, key=lambda c: (c.get('factuurdatum') or datetime.date.min), reverse=True)
    for cr in gesorteerd:
        matches = [r for r in open_posten if str(r.get('Leverancier', '')).startswith(cr['bureau'][:10])]
        wk_match = [r for r in matches if any(str(r.get('Week', '')).endswith(f'-{int(w):02d}') for w in cr['weken'] if str(w).isdigit())]
        bedr = cr['netto_bedrag']
        richting = 'te ontvangen' if bedr < -0.01 else ('bij te betalen' if bedr > 0.01 else 'saldo €0')
        if bedr < -0.01:
            saldo_ontv += bedr
        elif bedr > 0.01:
            saldo_bet += bedr
        if cr['type'] == 'Tariefcorrectie' and wk_match:
            actie = f'Sluit aan op {len(wk_match)} tarief-post(en) week {cr["weken"]} → register op "Ontvangen" zetten.'
            fill = FILL_OK
        elif cr['type'] == 'Creditnota' and wk_match:
            actie = f'{len(wk_match)} post(en) zelfde week → afletteren, register op "Ontvangen".'
            fill = FILL_OK
        elif matches:
            actie = f'{len(matches)} openstaande post(en) bij dit bureau — controleer aansluiting, dan afletteren.'
            fill = FILL_GEEL
        else:
            actie = 'Geen openstaande post — handmatig nalopen.'
            fill = FILL_ROOD
        oms = cr.get('omschrijving', '') or ('— geen namen op factuur —' if cr['type'] == 'Tariefcorrectie' else '')
        vals = [cr['bestand'][:46], _datum_str(cr.get('factuurdatum')), cr['type'], cr['bureau'],
                ', '.join(cr['weken']) or '?', cr.get('hoofdweek', ''),
                cr.get('netto_uren', ''), bedr, richting, oms, actie]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=rr, column=i, value=v); c.font = FONT_N; c.border = BORDER; c.fill = fill
            c.alignment = ALIGN_R if i in (7, 8) else ALIGN_C if i in (2, 6) else ALIGN_L
            if i in (10, 11):
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if i == 8:
                c.number_format = '€ #,##0.00;-€ #,##0.00'
        rr += 1
    # totaalregel
    tc = ws.cell(row=rr, column=7, value='Saldo'); tc.font = FONT_HEADER; tc.alignment = ALIGN_R; tc.border = BORDER
    ts = ws.cell(row=rr, column=8, value=round(saldo_ontv + saldo_bet, 2))
    ts.font = FONT_HEADER; ts.border = BORDER; ts.alignment = ALIGN_R; ts.number_format = '€ #,##0.00;-€ #,##0.00'
    rcell = ws.cell(row=rr, column=9, value=f'ontv €{saldo_ontv:,.2f} / bij €{saldo_bet:,.2f}')
    rcell.font = FONT_HEADER; rcell.border = BORDER; rcell.alignment = ALIGN_L
    for i, w in enumerate([40, 13, 16, 20, 16, 10, 11, 18, 14, 46, 48]):
        ws.column_dimensions[chr(65 + i)].width = w
    return saldo_ontv, saldo_bet


def verwerk_credits(map_pad, credits, snoop_data=None, tariefkaart_cache=None, factuurkaart=None):
    """Lettert creditfacturen af tegen het openstaande-posten-register en schrijft een overzicht
    met een apart tabblad per UZB + een samenvattingstabblad + een per-medewerker afletter-tab."""
    reg_pad = os.path.join(map_pad, REGISTER_NAAM)
    regels = lees_register(reg_pad)
    open_posten = [r for r in regels if str(r.get('Status', '')).lower().startswith(('open', 'te verif'))]

    print(f'\n=== CREDIT-/CORRECTIE-/TARIEFCORRECTIE-FACTUREN ({len(credits)}) ===')

    # groepeer per bureau (code), behoud volgorde van eerste voorkomen
    per_bureau = {}
    for cr in credits:
        per_bureau.setdefault(cr['code'], []).append(cr)

    wb = Workbook()
    # Samenvattingstabblad eerst
    ws_sum = wb.active
    ws_sum.title = 'Samenvatting'
    ws_sum.merge_cells('A1:D1')
    ws_sum['A1'] = 'Creditfacturen — samenvatting per UZB'
    ws_sum['A1'].font = FONT_TITLE; ws_sum['A1'].fill = FILL_TITLE; ws_sum['A1'].alignment = ALIGN_C
    for i, h in enumerate(['UZB', 'Aantal', 'Te ontvangen (–)', 'Bij te betalen (+)'], start=1):
        c = ws_sum.cell(row=3, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER

    sum_rr = 4
    tot_ontv = tot_bet = 0.0
    for code, subset in per_bureau.items():
        naam = subset[0]['bureau']
        # tabbladnaam mag max 31 tekens en geen rare tekens
        tabnaam = code if code else naam[:31]
        ws = wb.create_sheet(title=tabnaam[:31])
        ontv, bet = _schrijf_credit_sheet(ws, f'Creditfacturen — {naam}', subset, open_posten)
        tot_ontv += ontv; tot_bet += bet
        for cr in subset:
            bedr = cr['netto_bedrag']
            richting = 'te ontvangen' if bedr < -0.01 else ('bij te betalen' if bedr > 0.01 else 'saldo €0')
            oms = cr.get('omschrijving', '') or ('— geen namen op factuur —' if cr['type'] == 'Tariefcorrectie' else '')
            print(f"  [{code}] [{cr['type']:<15}] {cr['bestand'][:46]:<46} saldo €{bedr:>10.2f} ({richting}) | {oms[:50]}")
        vals = [naam, len(subset), round(ontv, 2), round(bet, 2)]
        for i, v in enumerate(vals, start=1):
            c = ws_sum.cell(row=sum_rr, column=i, value=v); c.font = FONT_N; c.border = BORDER
            c.alignment = ALIGN_R if i in (2, 3, 4) else ALIGN_L
            if i in (3, 4):
                c.number_format = '€ #,##0.00;-€ #,##0.00'
        sum_rr += 1
    # totaalregel samenvatting
    tc = ws_sum.cell(row=sum_rr, column=1, value='TOTAAL'); tc.font = FONT_HEADER; tc.border = BORDER
    ac = ws_sum.cell(row=sum_rr, column=2, value=len(credits)); ac.font = FONT_HEADER; ac.border = BORDER; ac.alignment = ALIGN_R
    for i, v in ((3, round(tot_ontv, 2)), (4, round(tot_bet, 2))):
        c = ws_sum.cell(row=sum_rr, column=i, value=v); c.font = FONT_HEADER; c.border = BORDER
        c.alignment = ALIGN_R; c.number_format = '€ #,##0.00;-€ #,##0.00'
    for i, w in enumerate([24, 10, 18, 18]):
        ws_sum.column_dimensions[chr(65 + i)].width = w

    if not per_bureau:
        ws_sum.cell(row=4, column=1, value='Geen credit-/correctiefacturen aangetroffen.')

    # ── Per-medewerker afletter-tab: tarief op de correctie vs SNOOP-inschaling ──
    afl = _afletter_medewerkers(credits, snoop_data, tariefkaart_cache, factuurkaart)
    if afl:
        wsa = wb.create_sheet(title='Aansluiting medewerkers')
        wsa.merge_cells('A1:K1')
        wsa['A1'] = 'Per medewerker — tarief op correctie vs SNOOP-inschaling'
        wsa['A1'].font = FONT_TITLE; wsa['A1'].fill = FILL_TITLE; wsa['A1'].alignment = ALIGN_C
        n_ok = sum(1 for r in afl if r['kleur'] == 'groen')
        n_afw = sum(1 for r in afl if r['kleur'] == 'rood')
        n_ver = sum(1 for r in afl if r['kleur'] == 'blauw')
        wsa.merge_cells('A2:K2')
        wsa['A2'] = (f'Doel: controleren of iedereen conform tarief wordt betaald. '
                     f'✅ klopt met inschaling: {n_ok}  ·  ⚠️ wijkt af: {n_afw}  ·  '
                     f'🔵 niet te verifiëren (geen SNOOP-inschaling): {n_ver}. '
                     f'"Nieuw tarief" = het tarief ná de correctie; vergeleken op 2 decimalen met de inschaling.')
        wsa['A2'].font = Font(name='Arial', size=9, italic=True); wsa['A2'].alignment = ALIGN_L
        wsa.row_dimensions[2].height = 40
        hdr = ['Bureau', 'Week', 'Medewerker', 'Schaal (SNOOP)', 'Categorie', 'Netto uren',
               'Nieuw tarief', 'Inschaling-tarief', 'Δ/u', 'Oordeel', 'Factuur']
        for i, h in enumerate(hdr, start=1):
            c = wsa.cell(row=4, column=i, value=h); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
        kleurmap = {'groen': FILL_OK, 'rood': FILL_ROOD, 'blauw': FILL_BLAUW}
        rr = 5
        # sorteer: afwijkingen eerst, dan te verifiëren, dan ok; daarbinnen bureau/week/naam
        prio = {'rood': 0, 'blauw': 1, 'groen': 2}
        for r in sorted(afl, key=lambda x: (prio[x['kleur']], x['code'], x['week'], x['naam'], x['categorie'])):
            fill = kleurmap[r['kleur']]
            vals = [r['code'], r['week'], r['naam'], r['schaal'], r['categorie'], r['netto_uren'],
                    r['nieuw_tarief'], r['inschaling_tarief'], r['delta'], r['oordeel'], r['factuur'][:46]]
            for i, v in enumerate(vals, start=1):
                c = wsa.cell(row=rr, column=i, value=v); c.font = FONT_N; c.border = BORDER; c.fill = fill
                c.alignment = ALIGN_R if i in (6, 7, 8, 9) else ALIGN_L
                if i in (7, 8):
                    c.number_format = '€ #,##0.000'
                if i == 9:
                    c.number_format = '+€ #,##0.000;-€ #,##0.000'
            rr += 1
        for i, w in enumerate([8, 9, 26, 13, 18, 10, 12, 14, 10, 46, 46]):
            wsa.column_dimensions[chr(65 + i)].width = w
        print(f'  Aansluiting medewerkers: {n_ok} klopt, {n_afw} wijkt af, {n_ver} te verifieren ({len(afl)} regels)')

    out = os.path.join(map_pad, 'Credits_overzicht.xlsx')
    try:
        wb.save(out)
        print(f'-> {os.path.basename(out)} (tabblad per UZB + samenvatting)')
    except PermissionError:
        wb.save(out.replace('.xlsx', '_v2.xlsx'))


if __name__ == '__main__':
    main()
